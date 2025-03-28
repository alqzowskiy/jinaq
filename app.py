# app.py
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from firebase_admin import credentials, initialize_app, auth, firestore, storage
import firebase_admin
from functools import wraps
from werkzeug.utils import secure_filename
import uuid
import datetime
from flask import jsonify, request, abort
import json
import re
import os
from flask import jsonify, url_for
from firebase_admin import firestore
from dotenv import load_dotenv
import gunicorn
import geocoder
import requests
from flask import jsonify
from geopy.geocoders import Nominatim
import google.generativeai as genai
from firebase_admin import firestore
import traceback
import requests
import os
import datetime
import uuid
from flask import request, jsonify, url_for, redirect, flash, session, render_template, abort
import random
import string
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
from io import BytesIO
from datetime import datetime
import matplotlib.pyplot as plt
import base64
from werkzeug.security import generate_password_hash
from flask import send_file
import datetime
from datetime import timedelta
import pandas as pd
import transliterate
import random
import string
from transliterate.base import TranslitLanguagePack, registry
from transliterate.discover import autodiscover
import openai



app = Flask(__name__)
app.secret_key = 'mega-secret-key-yeah'  
app.config['LOGO_SVG_PATH'] = 'jinaq_logo.svg'
firebase_creds_path = os.getenv('FIREBASE_CREDENTIALS_PATH')
openai.api_key = os.getenv("OPENAI_API_KEY")

ADMIN_IDS = os.getenv("ADMIN_IDS")

load_dotenv()

password_reset_tokens = {}
try:
    firebase_creds_path = os.getenv('FIREBASE_CREDENTIALS_PATH')
    with open(firebase_creds_path, 'r') as f:
        firebase_credentials = json.load(f)

    

    required_fields = ['type', 'project_id', 'private_key', 'client_email']
    for field in required_fields:
        if field not in firebase_credentials:
            raise ValueError(f"Missing required field: {field}")
    
    print("Firebase credentials fully validated")
except json.JSONDecodeError:
    print("ERROR: Invalid JSON in Firebase credentials")
except ValueError as e:
    print(f"Credentials validation error: {e}")


def check_openai_key_on_start():
    print("checking api key")
    api_key = openai.api_key

    if not api_key:
        print("not found")
        return False

    try:
        openai.Model.list()
        print("openai is on my guy")
        return True
    except openai.AuthenticationError:
        print("aint working my guy")
        return False
    except Exception as e:
        print(f"error is: {str(e)}")
        return False


NOTIFICATION_TYPES = {
    'project_comment': {
        'icon': '💬',
        'label': 'Project Comment'
    },
    'like_comment': {
        'icon': '❤️',
        'label': 'Comment Liked'
    },
    'reply_comment': {
        'icon': '💬',
        'label': 'Comment Replied'
    },
    'profile_comment': {
        'icon': '📝',
        'label': 'Profile Comment'
    },
    'account_change': {
        'icon': '🌐',
        'label': 'System Notification'
    },
    'verification': {
        'icon': '🏆',
        'label': 'Account Verified'
    },
    'system': {
        'icon': '🌐',
        'label': 'System Notification'
    },
    'project_collaboration': {
        'icon': '👥',
        'label': 'Project Collaboration'
    },
    'project_update': {
        'icon': '🔄',
        'label': 'Project Update'
    },
    'project_invitation': {
        'icon': '📩',
        'label': 'Project Invitation'
    }
}

ENHANCED_NOTIFICATION_TYPES = {
    **NOTIFICATION_TYPES,
    'system': { 
        'icon': '🌐',
        'label': 'System Notification',
        'priority': 'high',
        'sender': {
            'username': 'Jinaq',
            'verified': True,
            'verification_type': 'system'
        }
    },
    'project_comment': {
        'icon': '💬',
        'label': 'Project Comment',
        'priority': 'normal',
    },
    'admin_message': {
        'icon': '🏛️', 
        'label': 'Administrative Message', 
        'priority': 'critical',
        'sender': {
            'username': 'Jinaq Admin',
            'verified': True,
            'verification_type': 'official'
        }
    },
    'important': {
        'icon': '❗', 
        'label': 'Important Notification', 
        'priority': 'high',
        'sender': {
            'username': 'Jinaq Alert',
            'verified': True,
            'verification_type': 'system'
        }
    }
}

cred = credentials.Certificate(firebase_credentials)
firebase_admin.initialize_app(cred, {
    'storageBucket': 'jinaq-1b755.firebasestorage.app'
})

VERIFICATION_TYPES = {
    'official': {
        'icon': '🏛️',
        'color': 'blue',
        'label': 'Официальный аккаунт'
    },
    'creator': {
        'icon': '🎨',
        'color': 'purple',
        'label': 'Создатель контента'
    },
    'business': {
        'icon': '💼',
        'color': 'green',
        'label': 'Бизнес-аккаунт'
    }
}

GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY')
genai.configure(api_key=GOOGLE_API_KEY)



db = firestore.client()
bucket = storage.bucket()
import firebase_admin
from firebase_admin import credentials, initialize_app, auth, firestore, storage
import os
import json


@app.route('/debug/firebase_status')
def firebase_status():
    try: 

        firestore_client = firestore.client()
        auth_client = auth
        storage_client = storage.bucket()
        
        return jsonify({
            'firestore': 'Connected' if firestore_client else 'Not Connected',
            'auth': 'Initialized' if auth_client else 'Not Initialized',
            'storage': 'Connected' if storage_client else 'Not Connected'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
def initialize_firebase():
    try:

        if not firebase_admin._apps:
            firebase_creds_str = os.getenv('FIREBASE_PRIVATE_KEY')
            if not firebase_creds_str:
                raise ValueError("Firebase credentials not found in environment")
            
            firebase_credentials = json.loads(firebase_creds_str)
            
            # Создание credentials объекта
            cred = credentials.Certificate(firebase_credentials)
            
            # Инициализация с явным указанием параметров
            firebase_admin.initialize_app(cred, {
                'projectId': firebase_credentials['project_id'],
                'storageBucket': 'jinaq-1b755.firebasestorage.app'
            })
            
            print("Firebase successfully initialized")
        else:
            print("Firebase already initialized")
    except Exception as e:
        print(f"Firebase initialization error: {e}")
        raise

# Вызов функции инициализации
initialize_firebase()
@app.route('/debug/firebase_config')
def debug_firebase_config():
    try:
        firebase_creds_str = os.getenv('FIREBASE_PRIVATE_KEY')
        credentials_dict = json.loads(firebase_creds_str)
        
        # Маскировка чувствительных данных
        safe_credentials = {
            'type': credentials_dict.get('type'),
            'project_id': credentials_dict.get('project_id'),
            'client_email': credentials_dict.get('client_email'),
            'private_key_exists': bool(credentials_dict.get('private_key'))
        }
        
        return jsonify({
            'status': 'success',
            'credentials': safe_credentials
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
def get_current_user_avatar():
    """Helper function to get current user's avatar"""
    if 'user_id' not in session:
        return None
    
    try:
        current_user_id = session['user_id']
        current_user_doc = db.collection('users').document(current_user_id).get()
        current_user_data = current_user_doc.to_dict()
        
        return generate_avatar_url(current_user_data)
    except Exception as e:
        print(f"Error getting current user avatar: {e}")
        return None





def get_current_username():
    """Helper function to get current user's username"""
    if 'username' in session:
        return session['username']
    return None
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function
# Add this route to handle skills updates
@app.route('/update-skills', methods=['POST'])
@login_required
def update_skills():
    user_id = session['user_id']
    data = request.json
    
    try:
        # Clean and validate skills
        skills = [skill.strip() for skill in data.get('skills', []) if skill.strip()]
        
        # Update user document
        db.collection('users').document(user_id).update({
            'skills': skills,
            'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
        })
        
        return jsonify({'success': True, 'message': 'Skills updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/update_academic_portfolio', methods=['POST'])
@login_required
def update_academic_portfolio():
    user_id = session['user_id']
    
    # Detailed logging
    print("Full Request Data:", request.get_data(as_text=True))
    print("Request Headers:", dict(request.headers))
    print("Content Type:", request.content_type)
    

    try:
        data = request.get_json(force=True)
        

        if not isinstance(data, dict):
            print("Invalid data type:", type(data))
            return jsonify({
                'success': False, 
                'error': 'Invalid data format'
            }), 400


        required_keys = ['gpa', 'sat_score', 'toefl_score', 'ielts_score', 'languages', 'achievements']
        for key in required_keys:
            if key not in data:
                print(f"Missing key: {key}")
                return jsonify({
                    'success': False, 
                    'error': f'Missing required key: {key}'
                }), 400


        academic_info = {
            'gpa': str(data.get('gpa', '')).strip(),
            'sat_score': str(data.get('sat_score', '')).strip(),
            'toefl_score': str(data.get('toefl_score', '')).strip(),
            'ielts_score': str(data.get('ielts_score', '')).strip(),
            'languages': data.get('languages', []),
            'achievements': data.get('achievements', [])
        }

        db.collection('users').document(user_id).update({
            'academic_info': academic_info
        })

        return jsonify({'success': True, 'message': 'Academic portfolio updated'})

    except Exception as e:
        print(f"Academic Portfolio Update Error: {e}")
        return jsonify({
            'success': False, 
            'error': str(e)
        }), 400


@app.route('/settings')
@login_required
def settings():
    user_id = session['user_id']
    user_doc = db.collection('users').document(user_id).get()
    user_data = user_doc.to_dict() if user_doc.exists else {}
    

    current_user_avatar = get_current_user_avatar()
    current_username = get_current_username()
    
    return render_template(
        'settings.html',
        user_data=user_data,
        current_user_avatar=current_user_avatar,
        current_username=current_username,
        avatar_url=generate_avatar_url(user_data)
    )
@app.route('/update-email', methods=['POST'])
@login_required
def update_email():
    user_id = session['user_id']
    current_password = request.json.get('currentPassword')
    new_email = request.json.get('newEmail')
    
    try:

        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict()
        current_email = user_data['email']
        
        
        user = auth.update_user(
            user_id,
            email=new_email
        )
        
        
        db.collection('users').document(user_id).update({
            'email': new_email,
            'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
        })

        create_notification(
            user_id, 
            'account_change', 
            {
                'action': 'email_updated',
                'new_email': new_email,
                'message': f'Your email has been updated to {new_email}'
            }
        )

        return jsonify({'success': True, 'message': 'Email updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/')
def index():
    current_username = get_current_username()
    current_user_avatar = get_current_user_avatar()
    notifications_count = 0
    
    if 'user_id' in session:
        user_id = session['user_id']
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict() if user_doc.exists else None
        
        avatar_url = generate_avatar_url(user_data) if user_data else None
        notifications_count = get_unread_notifications_count(user_id)
        
        return render_template('index.html', 
                               user_data=user_data, 
                               avatar_url=avatar_url,
                               current_user_avatar=current_user_avatar,
                               current_username=current_username,
                               notifications_count=notifications_count)
    
    return render_template('index.html', 
                           user_data=None, 
                           avatar_url=None,
                           current_user_avatar=current_user_avatar,
                           current_username=current_username,
                           notifications_count=notifications_count)



@app.route('/onboarding')
@login_required
def onboarding():
    user_id = session['user_id']
    
    # Check if user has already completed onboarding
    user_doc = db.collection('users').document(user_id).get()
    user_data = user_doc.to_dict() or {}
    
    if user_data.get('onboarding_completed', False):
        # If onboarding already done, redirect to profile
        return redirect(url_for('profile'))
    
    # Check if this is a student with a school_id
    school_data = None
    if user_data.get('school_id'):
        try:
            school_id = user_data.get('school_id')
            school_doc = db.collection('schools').document(school_id).get()
            
            if school_doc.exists:
                school_data = {
                    'id': school_id,
                    'name': school_doc.to_dict().get('name', 'Unknown School'),
                    'logo_url': school_doc.to_dict().get('logo_url')
                }
        except Exception as e:
            print(f"Error getting school data: {e}")
    
    # Render the onboarding template with school data
    return render_template('onboarding.html', user_data=user_data, school_data=school_data)
@app.route('/complete_onboarding', methods=['POST'])
@login_required
def complete_onboarding():
    user_id = session['user_id']
    
    try:
        # Get form data from request
        data = request.get_json()
        
        # Process the onboarding data
        profile_data = {
            'full_name': data.get('full_name', ''),
            'specialty': data.get('specialty', ''),
            'education': data.get('education', ''),
            'bio': data.get('bio', ''),
            'goals': data.get('goals', ''),
            'skills': data.get('skills', []),
            'onboarding_completed': True,
            'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
        }
        
        # Update user document with onboarding data
        db.collection('users').document(user_id).update(profile_data)
        
        # Create a welcome notification
        create_notification(
            user_id, 
            'system', 
            {
                'message': 'Welcome to Jinaq! Your profile has been set up successfully.',
                'action': 'onboarding_completed'
            }
        )
        
        # Check if the user is from a school
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict()
        is_school_student = user_data.get('school_id') is not None
        
        # Set response data
        response_data = {
            'success': True,
            'redirect': url_for('profile'),
            'showCareerTest': True,  # Flag to show career test modal
            'careerTestMandatory': is_school_student  # Flag if it's mandatory
        }
        
        if is_school_student:
            # For school students, mark the test as mandatory in their profile
            db.collection('users').document(user_id).update({
                'requires_career_test': True
            })
        
        # Log successful completion
        print(f"User {user_id} completed onboarding successfully")
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error completing onboarding: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/show-career-test')
@login_required
def show_career_test_popup():
    """Route to re-show the career test popup"""
    user_id = session['user_id']
    
    # Check if user is from a school
    user_doc = db.collection('users').document(user_id).get()
    user_data = user_doc.to_dict()
    is_school_student = user_data.get('school_id') is not None
    
    # Set session flag for mandatory status
    is_mandatory = is_school_student and user_data.get('requires_career_test', False)
    
    # Return a small JS snippet that will show the modal
    return """
    <script>
        localStorage.setItem('showCareerTestModal', 'true');
        localStorage.setItem('careerTestMandatory', '{0}');
        window.location.href = '{1}';
    </script>
    """.format('true' if is_mandatory else 'false', url_for('profile'))

@app.before_request
def check_career_test_requirement():
    """Middleware to check if school student needs to complete the career test"""
    # Skip for non-authenticated users
    if 'user_id' not in session:
        return
        
    # Skip for certain paths
    excluded_paths = [
        '/career-test', 
        '/static',
        '/logout',
        '/login',
        '/register',
        '/favicon.ico',
        '/api',
        # Add any other paths that should be accessible without test completion
    ]
    
    if any(request.path.startswith(path) for path in excluded_paths):
        return
    
    user_id = session['user_id']
    
    try:
        # Check if the user is a school student who needs to complete the test
        user_doc = db.collection('users').document(user_id).get()
        if not user_doc.exists:
            return
            
        user_data = user_doc.to_dict()
        
        # Only enforce for school students with requires_career_test flag
        if not user_data.get('school_id') or not user_data.get('requires_career_test', False):
            return
            
        # Check if user has completed the career test
        test_completed = False
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        if test_progress.exists:
            progress_data = test_progress.to_dict()
            completed_stages = progress_data.get('completed_stages', [])
            
            # If all stages completed, mark test as completed
            if len(completed_stages) >= 4:
                test_completed = True
                # Remove the requirement flag
                db.collection('users').document(user_id).update({
                    'requires_career_test': False
                })
        
        # If test not completed, redirect to career test
        if not test_completed:
            return redirect(url_for('career_test'))
                
    except Exception as e:
        print(f"Error checking career test requirement: {e}")
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        display_username = request.form['username'] 
        username = display_username.lower()

        try:
            # Additional server-side validation for username
            if ' ' in username:
                flash('Username cannot contain spaces')
                return redirect(url_for('register'))
                
            # Check for non-Latin characters (including Cyrillic)
            if not all(ord(c) < 128 for c in username) or not username.isalnum() and '_' not in username:
                flash('Username must contain only Latin letters, numbers, and underscores')
                return redirect(url_for('register'))

            # Check username uniqueness
            users_ref = db.collection('users')
            username_query = users_ref.where('username', '==', username).get()
            if len(list(username_query)) > 0:
                flash('Username already taken')
                return redirect(url_for('register'))

            # Create user in Firebase Auth
            user = auth.create_user(
                email=email,
                password=password,
                display_name=display_username
            )

            # Create user document in Firestore
            user_data = {
                'username': username,
                'display_username': display_username, 
                'email': email,
                'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
                'uid': user.uid,
                'verified': False,
                'verification_type': None,
                'verified_by': None,
                'verified_at': None,
                'onboarding_completed': False,
                'academic_info': {
                    'gpa': '',
                    'sat_score': '',
                    'toefl_score': '',
                    'ielts_score': '',
                    'languages': [],
                    'achievements': []
                }
            }
            db.collection('users').document(user.uid).set(user_data)

            # Log the user in
            session['user_id'] = user.uid
            session['username'] = display_username
            
            # Redirect to onboarding
            return redirect(url_for('onboarding'))
            
        except Exception as e:
            flash(f'Error: {str(e)}')
            return redirect(url_for('register'))

    return render_template('register.html')

def check_onboarding_required():
    """Middleware to check if user needs to complete onboarding"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Skip check for onboarding, static, and certain other routes
            if (request.path.startswith('/onboarding') or 
                request.path.startswith('/static') or
                request.path == '/complete_onboarding' or
                request.path == '/logout'):
                return f(*args, **kwargs)
                
            if 'user_id' in session:
                user_id = session['user_id']
                
                # Skip for school admins
                user_doc = db.collection('schools').document(user_id).get()
                if user_doc.exists:
                    return f(*args, **kwargs)
                
                # Check if regular user has completed onboarding
                user_doc = db.collection('users').document(user_id).get()
                if user_doc.exists:
                    user_data = user_doc.to_dict()
                    if not user_data.get('onboarding_completed', False):
                        return redirect(url_for('onboarding'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator
@app.before_request
def check_user_onboarding():
    # Skip the onboarding check entirely - make it optional
    return    

@app.route('/<username>/comments/<comment_id>/reply', methods=['POST'])
@login_required
def reply_to_comment(username, comment_id):
    try:
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username.lower()).limit(1).stream()
        user_doc = next(query, None)

        if user_doc is None:
            return jsonify({'error': 'User not found'}), 404

        target_user_id = user_doc.id
        reply_text = request.form.get('reply')
        original_comment_id = request.form.get('original_comment_id', comment_id)

        if not reply_text or len(reply_text.strip()) == 0:
            return jsonify({'error': 'Reply text is required'}), 400


        comment_ref = db.collection('users').document(target_user_id).collection('comments').document(original_comment_id)
        comment_doc = comment_ref.get()
        
        if not comment_doc.exists:
            return jsonify({'error': 'Comment not found'}), 404
            
        comment_data = comment_doc.to_dict()
        

        if session['user_id'] != comment_data['author_id']:
            create_notification(
                comment_data['author_id'],
                'reply_comment',
                {
                    'original_comment_text': comment_data['text'],
                    'reply_text': reply_text.strip(),
                    'reply_chain': comment_data.get('reply_chain', [])
                },
                sender_id=session['user_id'],
                related_id=original_comment_id
            )


        reply_chain = comment_data.get('reply_chain', [])
        if 'parent_id' in comment_data:
            reply_chain = comment_data.get('reply_chain', [])
        reply_chain.append(original_comment_id)


        reply_data = {
            'author_id': session['user_id'],
            'author_username': session['username'],
            'text': reply_text.strip(),
            'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
            'parent_id': comment_id, 
            'reply_chain': reply_chain,
            'reply_to_username': comment_data['author_username'],
            'reply_level': len(reply_chain),
            'likes': []
        }

        reply_ref = db.collection('users').document(target_user_id).collection('comments').add(reply_data)


        author_doc = db.collection('users').document(session['user_id']).get()
        author_data = author_doc.to_dict()


        response_data = {
            'status': 'success',
            'reply': {
                'id': reply_ref[1].id,
                **reply_data,
                'author_avatar': generate_avatar_url(author_data),
                'author_verified': author_data.get('verified', False),
                'author_verification_type': author_data.get('verification_type'),
                'likes_count': 0,
                'is_liked': False
            }
        }

        return jsonify(response_data), 201

    except Exception as e:
        print(f"Error in reply_to_comment: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/<username>/comments/<comment_id>/like', methods=['POST'])
@login_required
def like_comment(username, comment_id):
    try:
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username.lower()).limit(1).stream()
        user_doc = next(query, None)

        if user_doc is None:
            return jsonify({'error': 'User not found'}), 404

        target_user_id = user_doc.id
        current_user_id = session['user_id']

        comment_ref = db.collection('users').document(target_user_id).collection('comments').document(comment_id)
        comment_doc = comment_ref.get()
        comment_data = comment_doc.to_dict()
        
        if comment_data['author_id'] != current_user_id:
            create_notification(
                comment_data['author_id'], 
                'like_comment', 
                {
                    'comment_text': comment_data['text']
                },
                sender_id=current_user_id,
                related_id=comment_id
            )

        if not comment_doc.exists:
            return jsonify({'error': 'Comment not found'}), 404

        comment_data = comment_doc.to_dict()
        likes = comment_data.get('likes', [])

        if current_user_id in likes:
            likes.remove(current_user_id)
        else:
            likes.append(current_user_id)




        comment_ref.update({'likes': likes})

        return jsonify({
            'status': 'success',
            'likes_count': len(likes),
            'is_liked': current_user_id in likes
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        identifier = request.form['identifier']  
        password = request.form['password']
        
        try:
            # Debugging
            print(f"Login attempt with identifier: {identifier}")
            
            # First try to find by email
            try:
                user = auth.get_user_by_email(identifier)
                email = identifier
                print(f"Found user by email: {user.uid}")
            except Exception as e:
                print(f"Email lookup failed: {e}")
                
                # Next try to find by username in users collection
                users_ref = db.collection('users')
                query = users_ref.where('username', '==', identifier.lower()).limit(1).get()
                
                if not query or len(query) == 0:
                    print("Username not found in users collection, checking schools collection")
                    
                    # Try to find in schools collection
                    schools_ref = db.collection('schools')
                    school_query = schools_ref.where('username', '==', identifier).limit(1).get()
                    
                    if not school_query or len(school_query) == 0:
                        print("Username not found in schools collection either")
                        flash('User not found')
                        return redirect(url_for('login'))
                    
                    school_data = school_query[0].to_dict()
                    email = school_data['email']
                    print(f"Found school by username, email: {email}")
                else:
                    user_data = query[0].to_dict()
                    email = user_data['email']
                    print(f"Found user by username, email: {email}")

                # Now get the user from Firebase Auth
                user = auth.get_user_by_email(email)
            
            # Try to authenticate with Firebase Auth (validation happens with Firebase)
            # Since we can't verify password directly, we'll rely on Firebase's auth methods
            # For additional security, you would implement sign-in with Firebase SDK
            
            # Check if user is a school
            school_doc = db.collection('schools').document(user.uid).get()
            if school_doc.exists:
                print(f"User {user.uid} is a school account")
                school_data = school_doc.to_dict()
                
                # Check if school is blocked
                if school_data.get('blocked', False):
                    flash('This account has been blocked. Please contact support.')
                    return redirect(url_for('login'))
                
                # Set session data for school
                session['user_id'] = user.uid
                session['username'] = school_data.get('name', 'School Admin')
                session['is_school'] = True
                
                # Redirect to appropriate page based on onboarding status
                if school_data.get('onboarding_completed', False):
                    return redirect(url_for('school_dashboard'))
                else:
                    return redirect(url_for('school_onboarding'))
            else:
                # Regular user login
                user_doc = db.collection('users').document(user.uid).get()
                user_data = user_doc.to_dict()
                
                if not user_data:
                    flash('User data not found')
                    return redirect(url_for('login'))
                
                if user_data.get('blocked', False):
                    flash('This account has been blocked. Please contact support.')
                    return redirect(url_for('login'))
                
                session['user_id'] = user.uid
                session['username'] = user_data.get('display_username', user_data['username'])
                session['is_school'] = False
                
                return redirect(url_for('profile'))

        except Exception as e:
            print(f"Login error (detailed): {str(e)}")
            import traceback
            traceback.print_exc()
            flash('Login failed: Invalid credentials')
            return redirect(url_for('login'))

    return render_template('login.html')
@app.route('/delete-certificate/<cert_id>', methods=['DELETE'])
@login_required
def delete_certificate(cert_id):
    try:
        user_id = session['user_id']
        

        cert_ref = db.collection('users').document(user_id).collection('certificates').document(cert_id)
        cert_doc = cert_ref.get()
        
        if not cert_doc.exists:
            return jsonify({'error': 'Certificate not found'}), 404
            
        cert_data = cert_doc.to_dict()
        

        if 'file_url' in cert_data:
            try:

                file_path = cert_data['file_url'].split('/')[-1]
                blob = bucket.blob(f'certificates/{user_id}/{file_path}')
                blob.delete()
            except Exception as e:
                print(f"Error deleting file from storage: {e}")
        

        cert_ref.delete()
        
        return jsonify({'status': 'success'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500




@app.route('/<username>/comments', methods=['GET', 'POST'])

def comments(username):
    if request.method == 'POST':

        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401

        comment_text = request.form.get('comment')
        if not comment_text or len(comment_text.strip()) == 0:
            return jsonify({'error': 'Comment text is required'}), 400
        
        if len(comment_text) > 500:
            return jsonify({'error': 'Comment is too long'}), 400

        try:

            users_ref = db.collection('users')
            query = users_ref.where('username', '==', username.lower()).limit(1).stream()
            user_doc = next(query, None)

            if user_doc is None:
                return jsonify({'error': 'User not found'}), 404

            target_user_id = user_doc.id

            comment_data = {
                'author_id': session['user_id'],
                'author_username': session['username'],
                'text': comment_text.strip(),
                'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
                'likes': []
            }


            comment_ref = db.collection('users').document(target_user_id).collection('comments').add(comment_data)
            

            author_doc = db.collection('users').document(session['user_id']).get()
            author_data = author_doc.to_dict()
            

            if session['user_id'] != target_user_id:
                create_notification(
                    target_user_id,
                    'profile_comment',
                    {
                        'commenter_username': session['username'],
                        'comment_text': comment_text.strip()
                    },
                    sender_id=session['user_id'],
                    related_id=comment_ref[1].id
                )


            response_data = {
                'status': 'success',
                'comment': {
                    'id': comment_ref[1].id,
                    **comment_data,
                    'author_avatar': generate_avatar_url(author_data),
                    'likes_count': 0,
                    'is_liked': False
                }
            }

            return jsonify(response_data), 201

        except Exception as e:
            print(f"Error creating comment: {str(e)}")
            return jsonify({'error': str(e)}), 500


    try:

        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username.lower()).limit(1).stream()
        user_doc = next(query, None)

        if user_doc is None:
            return jsonify({'error': 'User not found'}), 404

        target_user_id = user_doc.id


        comments_ref = db.collection('users').document(target_user_id).collection('comments')
        comments_query = comments_ref.order_by('created_at', direction=firestore.Query.DESCENDING)

        all_comments = comments_query.stream()
        

        main_comments = {}
        replies = {}


        for comment_doc in all_comments:
            comment = comment_doc.to_dict()
            comment['id'] = comment_doc.id

            author_doc = db.collection('users').document(comment['author_id']).get()
            author_data = author_doc.to_dict()
            
            comment['author_avatar'] = generate_avatar_url(author_data)
            comment['likes_count'] = len(comment.get('likes', []))
            comment['is_liked'] = session.get('user_id') in comment.get('likes', [])

            if 'parent_id' in comment:
                if comment['parent_id'] not in replies:
                    replies[comment['parent_id']] = []
                replies[comment['parent_id']].append(comment)
            else:

                main_comments[comment_doc.id] = comment


        for comment_id, comment in main_comments.items():
            if comment_id in replies:

                sorted_replies = sorted(replies[comment_id], key=lambda x: x['created_at'])
                
                
                def process_nested_replies(reply_list):
                    for reply in reply_list:
                        nested_replies = [r for r in replies.get(reply['id'], []) if r.get('parent_id') == reply['id']]
                        if nested_replies:
                            reply['replies'] = sorted(nested_replies, key=lambda x: x['created_at'])
                            process_nested_replies(reply['replies'])
                    return reply_list
                
                comment['replies'] = process_nested_replies(sorted_replies)

        return jsonify(list(main_comments.values()))

    except Exception as e:
        print(f"Error in comments route: {str(e)}")
        return jsonify({'error': str(e)}), 500
def prepare_project_data(project_doc, current_user_id=None):
    """Prepare project data with proper creator and collaborator information"""
    project_data = project_doc.to_dict()
    project_data['id'] = project_doc.id
    
    # Determine the user's role in this project
    is_creator = current_user_id == project_data.get('created_by')
    is_collaborator = any(c.get('user_id') == current_user_id for c in project_data.get('collaborators', []))
    
    project_data['role'] = 'creator' if is_creator else 'collaborator' if is_collaborator else 'viewer'
    
    # If user is a collaborator, add creator info
    if is_collaborator and 'created_by' in project_data:
        creator_id = project_data['created_by']
        try:
            creator_doc = db.collection('users').document(creator_id).get()
            if creator_doc.exists:
                creator_data = creator_doc.to_dict()
                project_data['creator'] = {
                    'user_id': creator_id,
                    'username': creator_data.get('display_username', creator_data.get('username', '')),
                    'avatar': generate_avatar_url(creator_data)
                }
        except Exception as e:
            print(f"Error fetching creator info: {e}")
    
    # If current user is a collaborator, filter themselves out of the collaborators list
    if current_user_id and is_collaborator and 'collaborators' in project_data:
        project_data['collaborators'] = [
            c for c in project_data['collaborators'] if c.get('user_id') != current_user_id
        ]
    
    return project_data
@app.route('/<username>/comments/<comment_id>/replies', methods=['GET'])
def get_comment_replies(username, comment_id):
    try:
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username.lower()).limit(1).stream()
        user_doc = next(query, None)

        if user_doc is None:
            return jsonify({'error': 'User not found'}), 404

        target_user_id = user_doc.id
        replies = db.collection('users').document(target_user_id).collection('comments').where('parent_id', '==', comment_id).order_by('created_at').stream()
        
        replies_list = []
        for reply_doc in replies:
            reply = reply_doc.to_dict()
            reply['id'] = reply_doc.id
            
            author_doc = db.collection('users').document(reply['author_id']).get()
            author_data = author_doc.to_dict()
            
            reply['author_avatar'] = author_data.get('avatar_url', url_for('static', filename='default-avatar.png'))
            reply['likes_count'] = len(reply.get('likes', []))
            reply['is_liked'] = session.get('user_id') in reply.get('likes', [])
            
            replies_list.append(reply)
            
        return jsonify(replies_list)
        
    except Exception as e:
        print(f"Error in get_comment_replies: {str(e)}")
        return jsonify({'error': str(e)}), 500    
@app.route('/<username>/comments/<comment_id>', methods=['DELETE'])
@login_required
def delete_comment(username, comment_id):
    try:
     
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username).limit(1).stream()
        user_doc = next(query, None)

        if user_doc is None:
            return jsonify({'error': 'User not found'}), 404

        target_user_id = user_doc.id
        
      
        comment_ref = db.collection('users').document(target_user_id).collection('comments').document(comment_id)
        comment_doc = comment_ref.get()
        
        if not comment_doc.exists:
            return jsonify({'error': 'Comment not found'}), 404
        
        comment_data = comment_doc.to_dict()
        
        if comment_data['author_id'] != session['user_id'] and session['user_id'] != target_user_id:
            return jsonify({'error': 'Unauthorized to delete this comment'}), 403
        
        comment_ref.delete()
        return jsonify({'status': 'success'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    

def generate_avatar_url(user_data):
    """Generate avatar URL for user"""
    if not user_data:
        return "https://ui-avatars.com/api/?name=U&background=random&color=fff&size=128"
    
    # If avatar_url exists in user data, use it
    if user_data.get('avatar_url'):
        return user_data['avatar_url']
    
    # Generate avatar from username
    display_name = user_data.get('display_username', user_data.get('username', 'U'))
    initials = ''.join(word[0].upper() for word in display_name.split()[:2])
    
    return f"https://ui-avatars.com/api/?name={initials}&background=random&color=fff&size=128"
@app.route('/debug/my-user-id')
def debug_user_id():
    if 'user_id' not in session:
        return "Not logged in"
    return session['user_id']
@app.route('/test-maintenance')
def test_maintenance_mode():
    try:
        # Check if maintenance document exists and is properly set
        settings_ref = db.collection('system_settings').document('maintenance')
        settings_doc = settings_ref.get()
        
        maintenance_data = {
            'exists': settings_doc.exists,
            'data': settings_doc.to_dict() if settings_doc.exists else None
        }
        
        return jsonify(maintenance_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/migrate_all_users_onboarding', methods=['GET'])
@login_required
def migrate_all_users_onboarding():
    # Ensure only admin can access
    if session.get('user_id') != 'vVbXL4LKGidXtrKnvqa21gWRY3V2':  # Your admin ID
        return "Unauthorized", 403
    
    try:
        # Get all users
        users_ref = db.collection('users')
        users = users_ref.stream()
        
        updated_count = 0
        
        for user_doc in users:
            user_data = user_doc.to_dict()
            
            # Only update if onboarding_completed is False or missing
            if not user_data.get('onboarding_completed', False):
                db.collection('users').document(user_doc.id).update({
                    'onboarding_completed': True
                })
                updated_count += 1
        
        return f"Successfully updated {updated_count} users. All existing users now have onboarding_completed set to True."
    
    except Exception as e:
        return f"Error during migration: {str(e)}", 500    
@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    # Ensure only admin can access
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    # Get sort order parameter
    sort_by = request.args.get('sort_by', 'created_at_desc')  # Default sort by creation date descending
    
    # Fetch users with registration date
    users_ref = db.collection('users')
    users = users_ref.stream()
    
    user_list = []
    for user_doc in users:
        user_data = user_doc.to_dict()
        user_data['id'] = user_doc.id
        # Convert created_at to datetime object for sorting if it exists
        if 'created_at' in user_data:
            # Store the original created_at for display
            user_data['created_at_display'] = user_data['created_at']
            # Format created_at for display
            if isinstance(user_data['created_at'], datetime.datetime):
                user_data['created_at_formatted'] = user_data['created_at'].strftime('%Y-%m-%d %H:%M')
            else:
                user_data['created_at_formatted'] = 'Unknown'
        else:
            user_data['created_at'] = datetime.datetime.min
            user_data['created_at_formatted'] = 'Unknown'
        
        user_list.append(user_data)
    
    # Sort users based on selected sort order
    if sort_by == 'created_at_asc':
        user_list.sort(key=lambda x: x.get('created_at', datetime.datetime.min))
    elif sort_by == 'created_at_desc':
        user_list.sort(key=lambda x: x.get('created_at', datetime.datetime.min), reverse=True)
    elif sort_by == 'username_asc':
        user_list.sort(key=lambda x: x.get('username', '').lower())
    elif sort_by == 'username_desc':
        user_list.sort(key=lambda x: x.get('username', '').lower(), reverse=True)
    
    # Calculate metrics
    total_users = len(user_list)
    verified_users = len([u for u in user_list if u.get('verified')])
    blocked_users = len([u for u in user_list if u.get('blocked')])
    
    # Check maintenance mode status
    maintenance_enabled = False
    try:
        settings_ref = db.collection('system_settings').document('maintenance')
        settings_doc = settings_ref.get()
        
        if settings_doc.exists:
            maintenance_data = settings_doc.to_dict()
            maintenance_enabled = maintenance_data.get('enabled', False)
    except Exception as e:
        print(f"Error getting maintenance status: {e}")
    
    return render_template('admin_dashboard.html', 
        users=user_list,
        total_users=total_users,
        verified_users=verified_users,
        blocked_users=blocked_users,
        user_growth_percentage=10,  # You can calculate this dynamically
        verified_percentage=(verified_users / total_users * 100) if total_users > 0 else 0,
        blocked_percentage=(blocked_users / total_users * 100) if total_users > 0 else 0,
        active_users=total_users - blocked_users,
        active_user_percentage=((total_users - blocked_users) / total_users * 100) if total_users > 0 else 0,
        generate_avatar_url=generate_avatar_url,
        maintenance_enabled=maintenance_enabled,
        current_sort=sort_by
    )

@app.route('/admin/delete_user/<user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    # Ensure only admin can access
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Get user document
        user_doc = db.collection('users').document(user_id).get()
        
        if not user_doc.exists:
            return jsonify({'success': False, 'error': 'User not found'}), 404
            
        user_data = user_doc.to_dict()
        username = user_data.get('display_username', user_data.get('username', 'Unknown User'))
        
        # Delete user's data in Firestore
        
        # 1. Delete user's certificates
        certificates_ref = db.collection('users').document(user_id).collection('certificates')
        delete_collection(certificates_ref, 100)
        
        # 2. Delete user's comments
        comments_ref = db.collection('users').document(user_id).collection('comments')
        delete_collection(comments_ref, 100)
        
        # 3. Delete user's notifications
        notifications_ref = db.collection('users').document(user_id).collection('notifications')
        delete_collection(notifications_ref, 100)
        
        # 4. Delete user's projects
        # First get all projects created by this user
        projects_created = db.collection('projects').where('created_by', '==', user_id).stream()
        for project in projects_created:
            # Delete project's comments
            project_comments_ref = db.collection('projects').document(project.id).collection('comments')
            delete_collection(project_comments_ref, 100)
            # Delete the project document
            db.collection('projects').document(project.id).delete()
        
        # 5. Delete user's avatars and other storage files
        try:
            # Delete avatar
            if user_data.get('avatar_url'):
                try:
                    avatar_path = f"avatars/{user_id}"
                    blobs = bucket.list_blobs(prefix=avatar_path)
                    for blob in blobs:
                        blob.delete()
                except Exception as e:
                    print(f"Error deleting avatar: {e}")
            
            # Delete certificate files
            cert_path = f"certificates/{user_id}"
            cert_blobs = bucket.list_blobs(prefix=cert_path)
            for blob in cert_blobs:
                blob.delete()
                
            # Delete project files
            project_path = f"projects/{user_id}"
            project_blobs = bucket.list_blobs(prefix=project_path)
            for blob in project_blobs:
                blob.delete()
        except Exception as e:
            print(f"Error deleting user files: {e}")
        
        # 6. Delete the user document
        db.collection('users').document(user_id).delete()
        
        # 7. Delete user from Firebase Auth
        try:
            auth.delete_user(user_id)
        except Exception as e:
            print(f"Error deleting user from Firebase Auth: {e}")
        
        # Log the action
        log_admin_action('delete_user', {
            'user_id': user_id,
            'username': username,
            'admin_id': session['user_id']
        })
        
        return jsonify({
            'success': True,
            'message': f"User '{username}' and all associated data deleted successfully."
        })
        
    except Exception as e:
        print(f"Error deleting user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/batch_delete_users', methods=['POST'])
@login_required
def batch_delete_users():
    # Ensure only admin can access
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        user_ids = data.get('user_ids', [])
        
        if not user_ids:
            return jsonify({'success': False, 'error': 'No user IDs provided'}), 400
        
        # Keep track of successful and failed deletions
        successful_deletions = 0
        failed_deletions = 0
        usernames = []
        
        # Process each user for deletion
        for user_id in user_ids:
            try:
                # Get user document
                user_doc = db.collection('users').document(user_id).get()
                
                if not user_doc.exists:
                    failed_deletions += 1
                    continue
                
                user_data = user_doc.to_dict()
                username = user_data.get('display_username', user_data.get('username', 'Unknown User'))
                usernames.append(username)
                
                # Delete user's data in Firestore
                
                # 1. Delete user's certificates
                certificates_ref = db.collection('users').document(user_id).collection('certificates')
                delete_collection(certificates_ref, 100)
                
                # 2. Delete user's comments
                comments_ref = db.collection('users').document(user_id).collection('comments')
                delete_collection(comments_ref, 100)
                
                # 3. Delete user's notifications
                notifications_ref = db.collection('users').document(user_id).collection('notifications')
                delete_collection(notifications_ref, 100)
                
                # 4. Delete user's projects
                # First get all projects created by this user
                projects_created = db.collection('projects').where('created_by', '==', user_id).stream()
                for project in projects_created:
                    # Delete project's comments
                    project_comments_ref = db.collection('projects').document(project.id).collection('comments')
                    delete_collection(project_comments_ref, 100)
                    # Delete the project document
                    db.collection('projects').document(project.id).delete()
                
                # 5. Delete user's avatars and other storage files
                try:
                    # Delete avatar
                    if user_data.get('avatar_url'):
                        try:
                            avatar_path = f"avatars/{user_id}"
                            blobs = bucket.list_blobs(prefix=avatar_path)
                            for blob in blobs:
                                blob.delete()
                        except Exception as e:
                            print(f"Error deleting avatar: {e}")
                    
                    # Delete certificate files
                    cert_path = f"certificates/{user_id}"
                    cert_blobs = bucket.list_blobs(prefix=cert_path)
                    for blob in cert_blobs:
                        blob.delete()
                        
                    # Delete project files
                    project_path = f"projects/{user_id}"
                    project_blobs = bucket.list_blobs(prefix=project_path)
                    for blob in project_blobs:
                        blob.delete()
                except Exception as e:
                    print(f"Error deleting user files: {e}")
                
                # 6. Delete the user document
                db.collection('users').document(user_id).delete()
                
                # 7. Delete user from Firebase Auth
                try:
                    auth.delete_user(user_id)
                except Exception as e:
                    print(f"Error deleting user from Firebase Auth: {e}")
                
                successful_deletions += 1
                
            except Exception as e:
                print(f"Error deleting user {user_id}: {e}")
                failed_deletions += 1
        
        # Log the action
        log_admin_action('batch_delete_users', {
            'user_ids': user_ids,
            'usernames': usernames,
            'successful_deletions': successful_deletions,
            'failed_deletions': failed_deletions,
            'admin_id': session['user_id']
        })
        
        return jsonify({
            'success': True,
            'message': f"Successfully deleted {successful_deletions} users. {failed_deletions} deletions failed.",
            'successful_deletions': successful_deletions,
            'failed_deletions': failed_deletions
        })
        
    except Exception as e:
        print(f"Error in batch delete users: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

        
def get_user_location():
    try:
        # Try to get IP address
        if request.headers.get('X-Forwarded-For'):
            ip = request.headers.get('X-Forwarded-For').split(',')[0]
        else:
            ip = request.remote_addr
            
        # Get location data from IP
        g = geocoder.ip(ip)
        if g.ok:
            return {
                'city': g.city,
                'country': g.country,
                'coordinates': {
                    'lat': g.lat,
                    'lng': g.lng
                }
            }
    except Exception as e:
        print(f"Error getting location: {e}")
    return None

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user_id = session['user_id']
    
    try:
        # === Base Profile Data Retrieval ===
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict() or {}
        
        # Check basic validation
        if user_data.get('blocked', False):
            abort(404)
        if not user_data.get('onboarding_completed', False):
            print(f"User {user_id} redirected to onboarding from profile page")
            return redirect(url_for('onboarding'))
        
        # Set up avatar
        avatar_url = generate_avatar_url(user_data)
        current_user_avatar = avatar_url
        
        # === Career Test Status & Results ===
        test_completed = False
        test_results = []
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        if test_progress.exists:
            progress_data = test_progress.to_dict()
            completed_stages = progress_data.get('completed_stages', [])
            
            # Test is completed if all 4 stages are done
            if len(completed_stages) >= 4:
                test_completed = True
                
                # Get recommendations - first try from progress data
                if 'recommendations' in progress_data:
                    test_results = progress_data.get('recommendations', [])
                else:
                    # Then check dedicated recommendations collection
                    recommendations_ref = db.collection('career_recommendations').document(user_id)
                    recommendations_doc = recommendations_ref.get()
                    
                    if recommendations_doc.exists:
                        rec_data = recommendations_doc.to_dict()
                        test_results = rec_data.get('recommendations', [])
        
        # === Location Data ===
        if not user_data.get('location'):
            location = get_user_location()
            if location:
                user_data['location'] = location
                db.collection('users').document(user_id).update({
                    'location': location
                })
        
        # === School Data (for students) ===
        if user_data.get('school_id'):
            school_id = user_data.get('school_id')
            school_doc = db.collection('schools').document(school_id).get()
            
            if school_doc.exists:
                school_data = school_doc.to_dict()
                user_data['school_name'] = school_data.get('name', 'Unknown School')
                user_data['school_logo_url'] = school_data.get('logo_url')
                
                # Automatically set education field if not already set
                if not user_data.get('education'):
                    user_data['education'] = user_data['school_name']
                    db.collection('users').document(user_id).update({
                        'education': user_data['school_name']
                    })
        
        # === Initialize Academic Info If Missing ===
        if 'academic_info' not in user_data:
            user_data['academic_info'] = {
                'gpa': '',
                'sat_score': '',
                'toefl_score': '',
                'ielts_score': '',
                'languages': [],
                'achievements': []
            }

        # === Form Submission Handling (POST request) ===
        if request.method == 'POST':
            try:
                projects = []
                
                # Get created projects
                created_projects_refs = list(db.collection('projects').where('created_by', '==', user_id).stream())
                for project_doc in created_projects_refs:
                    project_data = prepare_project_data(project_doc, user_id)
                    projects.append(project_data)
                
                # Get collaborative projects
                potential_collab_projects = db.collection('projects').stream()
                for project_doc in potential_collab_projects:
                    project_data = project_doc.to_dict()
                    
                    # Skip if user is the creator (already added above)
                    if project_data.get('created_by') == user_id:
                        continue
                        
                    # Check if user is in collaborators
                    collaborators = project_data.get('collaborators', [])
                    is_collaborator = any(c.get('user_id') == user_id for c in collaborators)
                    
                    if is_collaborator:
                        project_data = prepare_project_data(project_doc, user_id)
                        projects.append(project_data)
                
                # Sort projects by last updated time
                projects.sort(key=lambda x: x.get('last_updated', x.get('created_at', datetime.datetime.min)), reverse=True)
                # Handle JSON updates (AJAX requests)
                if request.headers.get('Content-Type') == 'application/json':
                    data = request.get_json()
                    
                    # Handle links update
                    if data.get('action') == 'update_links':
                        links = data.get('links', [])
                        processed_links = []
                        
                        for link in links:
                            if link.get('title') and link.get('url'):
                                url = link['url']
                                if not url.startswith(('http://', 'https://')):
                                    url = 'https://' + url
                                processed_links.append({
                                    'title': link['title'],
                                    'url': url
                                })
                        
                        db.collection('users').document(user_id).update({
                            'links': processed_links,
                            'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
                        })
                        return jsonify({'success': True})

                # Handle certificate upload
                if 'certificate' in request.files:
                    file = request.files['certificate']
                    title = request.form.get('title', file.filename)
                    
                    if file and file.filename:
                        file_extension = file.filename.rsplit('.', 1)[1].lower()
                        filename = f"certificates/{user_id}/{str(uuid.uuid4())}.{file_extension}"
                        
                        blob = bucket.blob(filename)
                        blob.upload_from_string(
                            file.read(),
                            content_type=file.content_type
                        )
                        
                        blob.make_public()
                        
                        cert_data = {
                            'title': title,
                            'file_url': blob.public_url,
                            'uploaded_at': datetime.datetime.now(tz=datetime.timezone.utc)
                        }
                        
                        cert_ref = db.collection('users').document(user_id).collection('certificates').add(cert_data)
                        
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return jsonify({
                                'success': True,
                                'user_data': user_data
                            })
                        
                        flash('Profile updated successfully!')
                        return redirect(url_for('profile'))

                # Handle avatar upload
                if 'avatar' in request.files:
                    avatar_file = request.files['avatar']
                    if avatar_file and avatar_file.filename:
                        # Delete old avatar if exists
                        try:
                            if user_data.get('avatar_url'):
                                old_avatar_path = user_data['avatar_url'].split('/')[-1]
                                old_blob = bucket.blob(f'avatars/{user_id}/{old_avatar_path}')
                                old_blob.delete()
                        except Exception as e:
                            print(f"Error deleting old avatar: {e}")

                        # Upload new avatar
                        file_extension = avatar_file.filename.rsplit('.', 1)[1].lower()
                        filename = f"{str(uuid.uuid4())}.{file_extension}"
                        full_path = f"avatars/{user_id}/{filename}"
                        
                        blob = bucket.blob(full_path)
                        blob.upload_from_string(
                            avatar_file.read(),
                            content_type=avatar_file.content_type
                        )
                        
                        blob.make_public()
                        
                        profile_data = {
                            'avatar_url': blob.public_url,
                            'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
                        }
                        db.collection('users').document(user_id).update(profile_data)

                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return jsonify({
                                'success': True,
                                'avatar_url': blob.public_url
                            })

                # Handle regular form submission
                if request.form:
                    # Get current user data to preserve academic info
                    current_user_data = db.collection('users').document(user_id).get().to_dict()
                    
                    profile_data = {
                        'full_name': request.form.get('full_name', ''),
                        'age': request.form.get('age', ''),
                        'specialty': request.form.get('specialty', ''),
                        'goals': request.form.get('goals', ''),
                        'bio': request.form.get('bio', ''),
                        'education': request.form.get('education', ''),
                        'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
                    }

                    # Preserve existing academic_info
                    if 'academic_info' in current_user_data:
                        profile_data['academic_info'] = current_user_data['academic_info']

                    # Convert age to integer if provided
                    if profile_data['age']:
                        try:
                            profile_data['age'] = int(profile_data['age'])
                        except ValueError:
                            del profile_data['age']

                    # Add social media and contact links
                    social_media = {
                        'linkedin': request.form.get('linkedin', ''),
                        'github': request.form.get('github', ''),
                        'twitter': request.form.get('twitter', ''),
                        'website': request.form.get('website', '')
                    }
                    profile_data['social_media'] = {k: v for k, v in social_media.items() if v}

                    # Update all profile data
                    db.collection('users').document(user_id).update(profile_data)
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': True})
                    
                    flash('Profile updated successfully!')
                    return redirect(url_for('profile'))

            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'error': str(e)})
                flash(f'Error updating profile: {str(e)}')
                return redirect(url_for('profile'))

        # === Projects Data ===
        projects = []
        
        # Get created projects
        created_projects_refs = list(db.collection('projects').where('created_by', '==', user_id).stream())
        for project_doc in created_projects_refs:
            project_data = project_doc.to_dict()
            project_data['id'] = project_doc.id
            project_data['role'] = 'creator'
            projects.append(project_data)
        
        # Get collaborative projects
        potential_collab_projects = db.collection('projects').stream()
        for project_doc in potential_collab_projects:
            project_data = project_doc.to_dict()
            if project_data.get('created_by') == user_id:
                continue
            collaborators = project_data.get('collaborators', [])
            is_collaborator = any(c.get('user_id') == user_id for c in collaborators)
            if is_collaborator:
                project_data['id'] = project_doc.id
                project_data['role'] = 'collaborator'
                projects.append(project_data)
        
        # Sort projects by last updated time
        projects.sort(key=lambda x: x.get('last_updated', x.get('created_at', datetime.datetime.min)), reverse=True)
        
        # === Certificates & Notifications ===
        certificates = list(db.collection('users').document(user_id).collection('certificates').stream())
        
        notifications_ref = db.collection('users').document(user_id).collection('notifications')
        unread_notifications = notifications_ref.where('is_read', '==', False).get()
        notifications_count = len(list(unread_notifications))
        
        # Get username for JS usage
        username = user_data.get('display_username', user_data.get('username', ''))

        # === Render Template with All Data ===
        return render_template('profile.html',
                              user_data=user_data,
                              avatar_url=avatar_url,
                              current_user_avatar=current_user_avatar,
                              certificates=certificates,
                              notifications_count=notifications_count,
                              social_media=user_data.get('social_media', {}),
                              academic_info=user_data.get('academic_info', {}),
                              projects=projects,
                              current_username=username,
                              test_completed=test_completed,
                              test_results=test_results)

    except Exception as e:
        print(f"Error in profile route: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while loading your profile.')
        return redirect(url_for('index'))
@app.route('/check-username', methods=['POST'])
def check_username():
    """API endpoint to check if a username is already taken."""
    try:
        data = request.get_json()
        username = data.get('username', '').lower()
        
        # Basic validation
        if not username:
            return jsonify({'valid': False, 'message': 'Username is required'})
            
        # Check for spaces
        if ' ' in username:
            return jsonify({'valid': False, 'message': 'Username cannot contain spaces'})
            
        # Check for non-Latin characters (including Cyrillic)
        if not all(ord(c) < 128 for c in username):
            return jsonify({'valid': False, 'message': 'Username must contain only Latin characters (a-z, 0-9, _)'})
        
        # Check for availability in database
        users_ref = db.collection('users')
        username_query = users_ref.where('username', '==', username).get()
        
        if len(list(username_query)) > 0:
            return jsonify({'valid': False, 'message': 'Username is already taken'})
        
        return jsonify({'valid': True, 'message': 'Username is available'})
        
    except Exception as e:
        print(f"Error checking username: {e}")
        return jsonify({'valid': False, 'message': 'Error checking username'}), 500
@app.route('/update-status', methods=['POST'])
@login_required
def update_status():
    user_id = session['user_id']
    data = request.json
    
    try:
        # Clean and validate status
        status_text = data.get('status', '').strip()
        
        # Check if exceeds character limit (280 characters like Twitter)
        if len(status_text) > 280:
            return jsonify({'success': False, 'error': 'Status exceeds the 280 character limit'}), 400
        
        # Update user document
        db.collection('users').document(user_id).update({
            'status': {
                'text': status_text,
                'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
            },
            'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
        })
        
        return jsonify({'success': True, 'message': 'Status updated successfully'})
    except Exception as e:
        print(f"Error updating status: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/<username>')
def public_profile(username):
    try:
        print(f"Attempting to access profile for username: {username}")
        
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username.lower()).limit(1).stream()
        user_doc = next(query, None)


        if user_doc is None:
            print(f"No user found with username: {username}")
            abort(404, description=f"User '{username}' not found")

        viewed_user_data = user_doc.to_dict()
        

        print(f"User data: {viewed_user_data}")
        

        if viewed_user_data.get('blocked', False):
            print(f"Profile for {username} is blocked")
            abort(404, description="User profile not available")
        
        viewed_user_avatar = generate_avatar_url(viewed_user_data)
        

        current_user_avatar = None
        if 'user_id' in session:
            current_user = db.collection('users').document(session['user_id']).get()
            if current_user.exists:
                current_user_data = current_user.to_dict()
                current_user_avatar = generate_avatar_url(current_user_data)
        
        certificates = list(db.collection('users').document(user_doc.id).collection('certificates').stream())
        certificates_count = len(certificates)
        user_id = user_doc.id
        projects = []
        projects_refs = list(db.collection('projects').where('created_by', '==', user_id).stream())
        
        for project_doc in projects_refs:
            project_data = project_doc.to_dict()
            project_data['id'] = project_doc.id
            projects.append(project_data)
        
        # Sort projects by last updated time
        projects.sort(key=lambda x: x.get('last_updated', x.get('created_at')), reverse=True)
        
        user_id = user_doc.id
        
        # Get projects (both created and collaborative)
        projects = []
        
        # Step 1: Get projects created by the user
        created_projects_refs = list(db.collection('projects').where('created_by', '==', user_id).stream())
        for project_doc in created_projects_refs:
            project_data = project_doc.to_dict()
            project_data['id'] = project_doc.id
            project_data['role'] = 'creator'
            projects.append(project_data)
        
        # Step 2: Get projects where user is a collaborator
        potential_collab_projects = db.collection('projects').stream()
        
        for project_doc in potential_collab_projects:
            project_data = project_doc.to_dict()
            
            # Skip if user is the creator (already added above)
            if project_data.get('created_by') == user_id:
                continue
                
            # Check if user is in collaborators
            collaborators = project_data.get('collaborators', [])
            is_collaborator = any(c.get('user_id') == user_id for c in collaborators)
            
            if is_collaborator:
                project_data['id'] = project_doc.id
                project_data['role'] = 'collaborator'
                projects.append(project_data)
        
        # Sort projects by last updated time
        projects.sort(key=lambda x: x.get('last_updated', x.get('created_at', datetime.datetime.min)), reverse=True)
        
        return render_template('public_profile.html',
                             user_data=viewed_user_data, 
                             avatar_url=viewed_user_avatar,
                             current_user_avatar=current_user_avatar,
                             current_username=session.get('username'),
                             certificates=certificates,
                             certificates_count=certificates_count,
                             projects=projects)
                             
    except Exception as e:
        print(f"Comprehensive error in public_profile: {e}")
        import traceback
        traceback.print_exc()
        

        raise
from firebase_admin import auth

@app.route('/update_links', methods=['POST'])
def update_links():
    try:

        id_token = request.headers.get('Authorization')
        if not id_token:
            return jsonify(success=False, error="No token provided"), 401
            

        if id_token.startswith('Bearer '):
            id_token = id_token.split('Bearer ')[1]
            

        decoded_token = auth.verify_id_token(id_token)
        user_id = decoded_token['uid']
        

        data = request.get_json()
        links = data.get('links', [])
        

        if not isinstance(links, list):
            return jsonify(success=False, error="Invalid data format"), 400
            

        for link in links:
            if not isinstance(link, dict) or 'title' not in link or 'url' not in link:
                return jsonify(success=False, error="Invalid link format"), 400
        

        db.collection('users').document(user_id).update({
            'links': links
        })
        
        return jsonify(success=True)
    except auth.InvalidIdTokenError:
        return jsonify(success=False, error="Invalid token"), 401
    except Exception as e:
        print(f"Error updating links: {str(e)}")
        return jsonify(success=False, error=str(e)), 500
@app.route('/admin/migrate_usernames', methods=['GET'])
@login_required  
def migrate_usernames():

    if session.get('user_id') != 'admin_user_id':  
        return "Unauthorized", 403
    
    users_ref = db.collection('users')
    users = users_ref.stream()
    
    for user_doc in users:
        user_data = user_doc.to_dict()
        

        if 'display_username' not in user_data:
            users_ref.document(user_doc.id).update({
                'username': user_data['username'].lower(),
                'display_username': user_data['username']
            })
    
    return "Migration completed successfully"    


ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'default_secret_password')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Incorrect admin password', 'error')
    
    # Add current date to be used in template
    from datetime import datetime
    return render_template('admin_login.html', current_date=datetime.now())


@app.route('/update-username', methods=['POST'])
@login_required
def update_username():
    user_id = session['user_id']
    new_username = request.json.get('username').lower()
    display_username = request.json.get('username') 
    
    try:
        # First check if this is a student account
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict() if user_doc.exists else {}
        
        is_student = user_data.get('school_id') is not None
        
        # Check if username already exists
        users_ref = db.collection('users')
        username_query = users_ref.where('username', '==', new_username).get()
        if len(list(username_query)) > 0:
            return jsonify({'error': 'Username already taken'}), 400
        
        user_ref = db.collection('users').document(user_id)
        
        # If it's a student, we handle differently
        if is_student:
            # Update username but keep the base part for school records
            user_ref.update({
                'username': new_username,
                'display_username': display_username,
                'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
            })
            
            # Update session with new display name
            session['username'] = display_username
            
            return jsonify({'success': True, 'message': 'Username updated successfully'})
        else:
            # Regular user - normal update
            user_ref.update({
                'username': new_username,
                'display_username': display_username,
                'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
            })
            
            session['username'] = display_username
            
            return jsonify({'success': True, 'message': 'Username updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/update-profile', methods=['POST'])
@login_required
def update_profile():
    user_id = session['user_id']
    data = request.json
    
    try:
        profile_data = {}
        
        if 'full_name' in data:
            profile_data['full_name'] = data['full_name']
            profile_data['full_name_lower'] = data['full_name'].lower()
            
            # Check if this is a student account
            user_doc = db.collection('users').document(user_id).get()
            user_data = user_doc.to_dict()
            
            # If this is a student, only update display name but keep original name for school
            if user_data and user_data.get('school_id'):
                # We don't update 'original_full_name' - it stays the same for school records
                pass
        
        if 'bio' in data:
            profile_data['bio'] = data['bio']
            
        if 'education' in data:
            # Check if this is a student account before allowing education update
            user_doc = db.collection('users').document(user_id).get()
            user_data = user_doc.to_dict()
            
            if not user_data or not user_data.get('school_id'):
                # Only update education if not a student account
                profile_data['education'] = data['education']
            
        profile_data['updated_at'] = datetime.datetime.now(tz=datetime.timezone.utc)
        
        db.collection('users').document(user_id).update(profile_data)
        
        return jsonify({'success': True, 'message': 'Profile updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update-password', methods=['POST'])
@login_required
def update_password():
    user_id = session['user_id']
    current_password = request.json.get('current_password')
    new_password = request.json.get('new_password')
    create_notification(
        user_id, 
        'account_change', 
        {
            'action': 'password_changed'
        }
    )
    try:

        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict()
        

        user = auth.get_user_by_email(user_data['email'])
        

        auth.update_user(
            user_id,
            password=new_password
        )
        
        return jsonify({'success': True, 'message': 'Password updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/delete-account', methods=['POST'])
@login_required
def delete_account():
    user_id = session['user_id']
    password = request.json.get('password')
    
    try:

        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict()
        
        user = auth.get_user_by_email(user_data['email'])
        
        certificates_ref = db.collection('users').document(user_id).collection('certificates')
        comments_ref = db.collection('users').document(user_id).collection('comments')
        

        certs = certificates_ref.stream()
        for cert in certs:
            cert.reference.delete()

        comments = comments_ref.stream()
        for comment in comments:
            comment.reference.delete()
            

        db.collection('users').document(user_id).delete()
        

        auth.delete_user(user_id)
        

        session.clear()
        
        return jsonify({'success': True, 'message': 'Account deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

import requests
import os
from dotenv import load_dotenv

# Make sure load_dotenv() is called at the beginning of your app.py
load_dotenv()


import uuid
import time

# Import necessary Firebase modules
from firebase_admin import auth

# Modify the custom reset handler to work with Firebase's reset link
@app.route('/custom-reset-handler')
def custom_reset_handler():
    try:
        # Get the oobCode parameter from the URL (Firebase's password reset code)
        oob_code = request.args.get('oobCode')
        mode = request.args.get('mode')
        
        print(f"Reset handler called with mode: {mode}, code: {oob_code}")
        
        if not oob_code or mode != 'resetPassword':
            flash('Invalid or missing reset code')
            return redirect(url_for('login'))
        
        # Verify the code is valid (will throw exception if not)
        try:
            email = auth.verify_password_reset_code(oob_code)
            print(f"Valid reset code for email: {email}")
            
            # Redirect to our custom reset page with the code
            return render_template('custom_reset_password.html', 
                                   oob_code=oob_code, 
                                   email=email)
            
        except Exception as verify_error:
            print(f"Error verifying reset code: {verify_error}")
            flash('Invalid or expired reset link')
            return redirect(url_for('login'))
            
    except Exception as e:
        print(f"Error in custom reset handler: {e}")
        flash('Error processing reset link')
        return redirect(url_for('login'))

@app.route('/custom-reset-password')
def custom_reset_password_page():
    # Get token from URL
    token = request.args.get('token')
    
    # Check if token exists and hasn't expired
    if not token or token not in password_reset_tokens:
        flash('Invalid or expired password reset link')
        return redirect(url_for('login'))
    
    # Check token expiration
    token_data = password_reset_tokens[token]
    if token_data['expires_at'] < time.time():
        # Remove expired token
        del password_reset_tokens[token]
        flash('Password reset link has expired')
        return redirect(url_for('login'))
    
    # Pass token to template
    return render_template('custom_reset_password.html', token=token)


@app.route('/search_users')
def search_users():
    query = request.args.get('query', '').lower()
    category = request.args.get('category')
    subcategory = request.args.get('subcategory')
    
    try:
        users_ref = db.collection('users')
        results = []
        
        # Если есть запрос и он достаточно длинный, или указана категория/подкатегория
        if (query and len(query) >= 2) or subcategory:
            # Ищем по всем возможным полям, если есть запрос
            if query and len(query) >= 2:
                # Подготовим список для хранения всех найденных документов
                all_user_docs = []
                
                # 1. Поиск по username (который должен быть в латинице и строчными буквами)
                username_query = users_ref.where('username', '>=', query).where('username', '<=', query + '\uf8ff').limit(20).stream()
                all_user_docs.extend(list(username_query))
                
                # 2. Поиск по display_username (может содержать кириллицу)
                # Приведем запрос к нижнему регистру для поиска
                display_username_query = []
                try:
                    # Попробуем найти пользователей, где display_username содержит запрос
                    # Это сложнее, так как Firestore не поддерживает поиск по подстроке
                    # Поэтому мы получим больше пользователей и отфильтруем их на стороне сервера
                    all_users = users_ref.limit(100).stream()  # Ограничим выборку для производительности
                    for user_doc in all_users:
                        user_data = user_doc.to_dict()
                        display_username = user_data.get('display_username', '').lower()
                        if query in display_username:
                            display_username_query.append(user_doc)
                except Exception as e:
                    print(f"Error in display_username search: {e}")
                
                all_user_docs.extend(display_username_query)
                
                # 3. Поиск по full_name_lower, если такое поле существует
                try:
                    # Получим первый документ, чтобы проверить наличие поля
                    first_user = next(users_ref.limit(1).stream(), None)
                    if first_user and 'full_name_lower' in first_user.to_dict():
                        full_name_query = users_ref.where('full_name_lower', '>=', query).where('full_name_lower', '<=', query + '\uf8ff').limit(20).stream()
                        all_user_docs.extend(list(full_name_query))
                except Exception as e:
                    print(f"Error in full_name search: {e}")
                
                # 4. Если ищем по имени/фамилии, но поля full_name_lower нет, 
                # попробуем поискать через полную выборку (для небольших наборов данных)
                if not any('full_name_lower' in doc.to_dict() for doc in all_user_docs if doc):
                    try:
                        # Получим всех пользователей и отфильтруем на стороне сервера
                        all_users = users_ref.limit(100).stream()  # Ограничим выборку для производительности
                        for user_doc in all_users:
                            user_data = user_doc.to_dict()
                            full_name = user_data.get('full_name', '').lower()
                            if query in full_name:
                                all_user_docs.append(user_doc)
                    except Exception as e:
                        print(f"Error in fallback full_name search: {e}")
                
                # Удаляем дубликаты, используя ID документа
                unique_docs = {}
                for doc in all_user_docs:
                    if doc and doc.id not in unique_docs:
                        unique_docs[doc.id] = doc
                
                # Обрабатываем каждый уникальный документ
                for doc_id, user_doc in unique_docs.items():
                    user_data = user_doc.to_dict()
                    
                    # Пропускаем заблокированных пользователей
                    if user_data.get('blocked', False):
                        continue
                    
                    # Применяем фильтр по подкатегории, если указан
                    if subcategory:
                        if subcategory == 'developers' and 'developer' not in user_data.get('specialty', '').lower():
                            continue
                        elif subcategory == 'designers' and 'design' not in user_data.get('specialty', '').lower():
                            continue
                        elif subcategory == 'students' and not user_data.get('education', ''):
                            continue
                        elif subcategory == 'verified' and not user_data.get('verified', False):
                            continue
                    
                    # Добавляем пользователя в результаты с правильными полями
                    results.append({
                        'user_id': doc_id,
                        'username': user_data.get('username', ''),  # Для @username
                        'display_username': user_data.get('display_username', ''),  # Отображаемое имя (может быть на кириллице)
                        'full_name': user_data.get('full_name', ''),
                        'avatar': generate_avatar_url(user_data),
                        'verified': user_data.get('verified', False)
                    })
            
            # Если нет запроса, но есть подкатегория, отфильтровываем всех пользователей
            elif subcategory:
                all_users_query = users_ref.limit(50).stream()
                
                for user_doc in all_users_query:
                    user_data = user_doc.to_dict()
                    
                    # Пропускаем заблокированных пользователей
                    if user_data.get('blocked', False):
                        continue
                    
                    # Фильтруем по подкатегории
                    if subcategory == 'developers' and 'developer' not in user_data.get('specialty', '').lower():
                        continue
                    elif subcategory == 'designers' and 'design' not in user_data.get('specialty', '').lower():
                        continue
                    elif subcategory == 'students' and not user_data.get('education', ''):
                        continue
                    elif subcategory == 'verified' and not user_data.get('verified', False):
                        continue
                    
                    results.append({
                        'user_id': user_doc.id,
                        'username': user_data.get('username', ''),
                        'display_username': user_data.get('display_username', ''),
                        'full_name': user_data.get('full_name', ''),
                        'avatar': generate_avatar_url(user_data),
                        'verified': user_data.get('verified', False)
                    })
        
        # Если нет запроса и подкатегории, но указана категория 'people'
        elif not query and not subcategory and category == 'people':
            # Ограничиваем до 20 пользователей для функции "View all"
            all_users_query = users_ref.limit(20).stream()
            
            for user_doc in all_users_query:
                user_data = user_doc.to_dict()
                
                # Пропускаем заблокированных пользователей
                if user_data.get('blocked', False):
                    continue
                
                results.append({
                    'user_id': user_doc.id,
                    'username': user_data.get('username', ''),
                    'display_username': user_data.get('display_username', ''),
                    'full_name': user_data.get('full_name', ''),
                    'avatar': generate_avatar_url(user_data),
                    'verified': user_data.get('verified', False)
                })
        
        # Улучшенная сортировка результатов по релевантности
        def sort_key(user):
            username = user.get('username', '').lower()
            display_username = user.get('display_username', '').lower()
            full_name = user.get('full_name', '').lower()
            
            # Создаем числовой вес для сортировки (меньше = выше в результатах)
            score = 100  # Начальное значение
            
            # Проверяем точные совпадения
            if query and username == query:
                score -= 50
            elif query and display_username == query:
                score -= 40
            elif query and full_name == query:
                score -= 30
                
            # Проверяем совпадения по началу строки
            elif query and username.startswith(query):
                score -= 25
            elif query and display_username.startswith(query):
                score -= 20
            elif query and full_name.startswith(query):
                score -= 15
                
            # Проверяем вхождения где-либо в строке
            elif query and query in username:
                score -= 10
            elif query and query in display_username:
                score -= 8
            elif query and query in full_name:
                score -= 5
                
            # Верифицированные пользователи показываются выше
            if user.get('verified', False):
                score -= 3
                
            return score
        
        # Применяем сортировку
        if query:
            results.sort(key=sort_key)
        else:
            # Если нет запроса, сортируем по имени
            results.sort(key=lambda x: x.get('display_username', '').lower())
        
        return jsonify(results[:20])  # Ограничиваем до 20 результатов
    
    except Exception as e:
        print(f"Error searching users: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([]), 500

@app.before_request
def check_user_exists():
    """Проверяет, существует ли пользователь перед каждым запросом"""
    # Пропускаем для публичных маршрутов и статических файлов
    public_routes = ['login', 'register', 'index', 'logout']
    if (request.endpoint in public_routes or 
        request.endpoint is None or 
        request.path.startswith('/static')):
        return
    
    # Если пользователь в сессии, проверяем, существует ли он
    if 'user_id' in session:
        user_id = session['user_id']
        
        try:
            # Проверяем существование пользователя в Firebase Auth
            try:
                # Пробуем получить пользователя в Firebase Auth
                auth.get_user(user_id)
            except auth.UserNotFoundError:
                # Если пользователь не найден в Firebase Auth, выходим из системы
                print(f"User {user_id} not found in Firebase Auth, logging out")
                session.clear()
                flash('Your account has been deleted or disabled. Please log in again.')
                return redirect(url_for('login'))
            
            # Также проверяем существование документа пользователя
            if session.get('is_school'):
                # Для школьных аккаунтов проверяем коллекцию schools
                school_doc = db.collection('schools').document(user_id).get()
                if not school_doc.exists:
                    print(f"School {user_id} document not found, logging out")
                    session.clear()
                    flash('Your school account has been deleted. Please contact support.')
                    return redirect(url_for('login'))
            else:
                # Для обычных пользователей проверяем коллекцию users
                user_doc = db.collection('users').document(user_id).get()
                if not user_doc.exists:
                    print(f"User {user_id} document not found, logging out")
                    session.clear()
                    flash('Your account data has been deleted. Please log in again.')
                    return redirect(url_for('login'))
                
                # Проверяем, не заблокирован ли пользователь
                user_data = user_doc.to_dict()
                if user_data.get('blocked', False):
                    print(f"User {user_id} is blocked, logging out")
                    session.clear()
                    flash('Your account has been blocked. Please contact support.')
                    return redirect(url_for('login'))
            
        except Exception as e:
            # В случае любых ошибок при проверке, выходим из системы
            print(f"Error checking user {user_id} existence: {e}")
            session.clear()
            flash('Session expired. Please log in again.')
            return redirect(url_for('login'))
@app.errorhandler(500)
def internal_server_error(e):
    # Если ошибка связана с отсутствием пользователя, очищаем сессию
    error_msg = str(e)
    if 'auth/user-not-found' in error_msg or 'No such user' in error_msg:
        session.clear()
        flash('Your session has expired. Please log in again.')
        return redirect(url_for('login'))
    
    print(f"500 Error: {e}")
    print(f"Request URL: {request.url}")
    print(f"Request Method: {request.method}")
    
    return render_template('500.html'), 500        
@app.route('/admin/migrate_fullnames', methods=['GET'])
@login_required
def migrate_fullnames():
    if session.get('user_id') != 'vVbXL4LKGidXtrKnvqa21gWRY3V2': 
        return "Unauthorized", 403
        
    try:
        users_ref = db.collection('users')
        users = users_ref.stream()
        
        for user_doc in users:
            user_data = user_doc.to_dict()
            
            if 'full_name' in user_data and 'full_name_lower' not in user_data:
                users_ref.document(user_doc.id).update({
                    'full_name_lower': user_data['full_name'].lower()
                })
        
        return "Migration completed successfully"
    except Exception as e:
        return f"Error during migration: {str(e)}", 500
@app.route('/admin/add_admin', methods=['POST'])
@login_required
def add_admin():

    SUPER_ADMIN_IDS = ['vVbXL4LKGidXtrKnvqa21gWRY3V2']

    if session['user_id'] not in SUPER_ADMIN_IDS:
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    user_id = data.get('user_id')
    
    try:

        db.collection('users').document(user_id).update({
            'is_admin': True,
            'admin_added_by': session['user_id'],
            'admin_added_at': firestore.SERVER_TIMESTAMP
        })


        create_notification(
            user_id, 
            'account_change', 
            {
                'message': 'You have been granted admin access.'
            },
            sender_id=session['user_id']
        )

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/admin/verify_user', methods=['POST'])
@login_required
def verify_user():

    ADMIN_IDS = ['vVbXL4LKGidXtrKnvqa21gWRY3V2']

    # Debug logging
    print(f"Verification attempt by user: {session.get('user_id')}")

    if session['user_id'] not in ADMIN_IDS:
        print("Unauthorized verification attempt")
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:

        data = request.get_json()
        

        user_id = data.get('user_id')
        verification_type = data.get('type', 'official')
        
        if not user_id:
            print("No user ID provided")
            return jsonify({'success': False, 'error': 'User ID is required'}), 400

        valid_types = ['official', 'creator', 'business', 'remove']
        if verification_type not in valid_types:
            print(f"Invalid verification type: {verification_type}")
            return jsonify({'success': False, 'error': 'Invalid verification type'}), 400


        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            print(f"User not found: {user_id}")
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        user_data = user_doc.to_dict()
        

        update_data = {
            'verified': verification_type != 'remove',
            'verification_type': None if verification_type == 'remove' else verification_type,
            'verified_by': session['user_id'] if verification_type != 'remove' else None,
            'verified_at': firestore.SERVER_TIMESTAMP if verification_type != 'remove' else None
        }
        

        try:
            user_ref.update(update_data)
        except Exception as update_error:
            print(f"Database update error: {update_error}")
            return jsonify({'success': False, 'error': 'Failed to update user verification'}), 500


        if verification_type == 'remove':
            create_notification(
                user_id, 
                'account_change', 
                {
                    'message': 'Your account verification has been revoked.',
                    'action': 'verification_revoked'
                },
                sender_id=session['user_id']
            )
        else:
            create_notification(
                user_id, 
                'verification', 
                {
                    'message': f'Your account has been verified as a {verification_type} account.',
                    'verification_type': verification_type
                },
                sender_id=session['user_id']
            )


        print(f"User {user_id} verified as {verification_type}")
        
        return jsonify({
            'success': True, 
            'message': f'User verified as {verification_type}',
            'verification_type': verification_type
        })
    
    except Exception as e:

        print(f"Unexpected error in user verification: {e}")
        return jsonify({
            'success': False, 
            'error': 'An error occurred while verifying the user.',
            'details': str(e)
        }), 500

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))
def create_notification(user_id, type, content, sender_id=None, related_id=None):
    try:
        # Get sender info if sender_id is provided
        if sender_id:
            try:
                sender_doc = db.collection('users').document(sender_id).get()
                sender_data = sender_doc.to_dict()
                sender_info = {
                    'username': sender_data.get('display_username', sender_data.get('username', 'Jinaq')),
                    'avatar_url': generate_avatar_url(sender_data),
                    'verified': sender_data.get('verified', False),
                    'verification_type': sender_data.get('verification_type')
                }
            except Exception as e:
                print(f"Error fetching sender info: {e}")
                sender_info = None
        else:
            # For system notifications, use default sender info
            type_config = ENHANCED_NOTIFICATION_TYPES.get(type, {})
            sender_info = type_config.get('sender', {
                'username': 'Jinaq',
                'avatar_url': generate_avatar_url({'username': 'Jinaq'}),
                'verified': True,
                'verification_type': 'system'
            })

        # Get notification type configuration
        notification_config = ENHANCED_NOTIFICATION_TYPES.get(type, {})
        
        # Enrich content based on notification type
        enriched_content = {
            'type_label': notification_config.get('label', 'Notification'),
            'icon': notification_config.get('icon', '🔔'),
            'priority': notification_config.get('priority', 'normal'),
            **content
        }
        
        # Process specific notification types to add more context
        if type == 'project_collaboration':
            project_id = content.get('project_id')
            if project_id:
                try:
                    project_doc = db.collection('projects').document(project_id).get()
                    if project_doc.exists:
                        project_data = project_doc.to_dict()
                        if 'project_title' not in content:
                            enriched_content['project_title'] = project_data.get('title', 'Unknown Project')
                        if 'project_image' not in content and 'thumbnail' in project_data:
                            enriched_content['project_image'] = project_data.get('thumbnail')
                except Exception as e:
                    print(f"Error enriching project collaboration notification: {e}")
        
        elif type == 'project_update':
            project_id = content.get('project_id')
            if project_id:
                try:
                    project_doc = db.collection('projects').document(project_id).get()
                    if project_doc.exists:
                        project_data = project_doc.to_dict()
                        if 'project_title' not in content:
                            enriched_content['project_title'] = project_data.get('title', 'Unknown Project')
                except Exception as e:
                    print(f"Error enriching project update notification: {e}")
        
        elif type == 'profile_comment':
            if 'commenter_username' in content and 'username' not in content:
                enriched_content['username'] = content['commenter_username']
        
        elif type == 'reply_comment':
            # Get the original comment text if not included
            if 'original_comment_text' not in content and 'related_id' in locals() and related_id:
                try:
                    if content.get('original_comment'):
                        enriched_content['original_comment_text'] = content['original_comment'].get('text', '')
                    elif user_id and related_id:
                        comment_doc = db.collection('users').document(user_id).collection('comments').document(related_id).get()
                        if comment_doc.exists:
                            comment_data = comment_doc.to_dict()
                            enriched_content['original_comment_text'] = comment_data.get('text', '')
                except Exception as e:
                    print(f"Error enriching reply comment notification: {e}")
        
        # Create notification data structure
        notification_data = {
            'type': type,
            'content': enriched_content,
            'sender_info': sender_info,
            'sender_id': sender_id,
            'related_id': related_id,
            'is_read': False,
            'created_at': firestore.SERVER_TIMESTAMP,
            'recipient_id': user_id,
            'metadata': {
                'timestamp': datetime.datetime.now(tz=datetime.timezone.utc),
                'source': 'server',
                'version': '1.2'
            }
        }

        # Set pinned status for critical notifications
        if notification_config.get('priority') == 'critical':
            notification_data['is_pinned'] = True

        # Create the notification in Firestore
        notification_ref = db.collection('users').document(user_id).collection('notifications').document()
        notification_ref.set(notification_data)

        return notification_ref.id

    except Exception as e:
        print(f"Comprehensive Notification Creation Error: {e}")
        return None


@app.route('/notifications/<notification_id>/details', methods=['GET'])
@login_required
def get_notification_details(notification_id):
    """Get full details for a notification including all related data"""
    user_id = session['user_id']
    
    try:
        notification_ref = db.collection('users').document(user_id).collection('notifications').document(notification_id)
        notification = notification_ref.get()
        
        if not notification.exists:
            return jsonify({'error': 'Notification not found'}), 404
            
        notification_data = notification.to_dict()
        

        if notification_data['type'] == 'reply_comment' and notification_data.get('related_id'):
            comment_id = notification_data['related_id']
            comment_ref = db.collection('users').document(user_id).collection('comments').document(comment_id)
            comment_data = comment_ref.get().to_dict()
            
            if comment_data:
                notification_data['comment_details'] = {
                    'text': comment_data['text'],
                    'created_at': comment_data['created_at'],
                    'author_username': comment_data['author_username']
                }

        return jsonify(notification_data)
    except Exception as e:
        print(f"Error getting notification details: {e}")
        return jsonify({'error': 'Failed to get notification details'}), 500
@app.route('/search_projects')
def search_projects():
    query = request.args.get('query', '').lower()
    category = request.args.get('category')
    subcategory = request.args.get('subcategory')
    
    try:
        projects_ref = db.collection('projects')
        results = []
        
        # Если есть запрос и он достаточно длинный, или указана подкатегория
        if (query and len(query) >= 2) or subcategory:
            # Поиск по названию, если есть запрос
            if query and len(query) >= 2:
                # Сначала получаем все проекты и фильтруем их
                title_query = projects_ref.limit(50).stream()
                
                for doc in title_query:
                    project_data = doc.to_dict()
                    project_title = project_data.get('title', '').lower()
                    
                    # Пропускаем удаленные проекты
                    if project_data.get('deleted', False):
                        continue
                    
                    # Ищем по названию
                    if query not in project_title:
                        continue
                    
                    # Применяем фильтр по подкатегории
                    if subcategory:
                        tags = [tag.lower() for tag in project_data.get('tags', [])]
                        
                        if subcategory == 'web' and 'web' not in tags and 'website' not in tags:
                            continue
                        elif subcategory == 'mobile' and 'mobile' not in tags and 'app' not in tags:
                            continue
                        elif subcategory == 'design' and 'design' not in tags and 'ui' not in tags and 'ux' not in tags:
                            continue
                        elif subcategory == 'popular' and len(project_data.get('likes', [])) < 5:
                            continue
                    
                    results.append({
                        'id': doc.id,
                        'title': project_data.get('title', ''),
                        'description': project_data.get('description', '')[:100] + '...' if project_data.get('description', '') else '',
                        'thumbnail': project_data.get('thumbnail', None),
                        'tags': project_data.get('tags', []),
                        'created_by': project_data.get('created_by', '')
                    })
            
            # Только фильтрация по подкатегории, без запроса
            elif subcategory:
                subcategory_query = projects_ref.limit(50).stream()
                
                for doc in subcategory_query:
                    project_data = doc.to_dict()
                    
                    # Пропускаем удаленные проекты
                    if project_data.get('deleted', False):
                        continue
                    
                    # Применяем фильтр по подкатегории
                    tags = [tag.lower() for tag in project_data.get('tags', [])]
                    
                    if subcategory == 'web' and 'web' not in tags and 'website' not in tags:
                        continue
                    elif subcategory == 'mobile' and 'mobile' not in tags and 'app' not in tags:
                        continue
                    elif subcategory == 'design' and 'design' not in tags and 'ui' not in tags and 'ux' not in tags:
                        continue
                    elif subcategory == 'popular' and len(project_data.get('likes', [])) < 5:
                        continue
                    
                    results.append({
                        'id': doc.id,
                        'title': project_data.get('title', ''),
                        'description': project_data.get('description', '')[:100] + '...' if project_data.get('description', '') else '',
                        'thumbnail': project_data.get('thumbnail', None),
                        'tags': project_data.get('tags', []),
                        'created_by': project_data.get('created_by', '')
                    })
        
        # Если нет запроса и подкатегории, но указана категория 'projects'
        elif not query and not subcategory and category == 'projects':
            # Ограничиваем до 20 проектов для функции "View all"
            all_projects_query = projects_ref.limit(20).stream()
            
            for doc in all_projects_query:
                project_data = doc.to_dict()
                
                # Пропускаем удаленные проекты
                if project_data.get('deleted', False):
                    continue
                
                results.append({
                    'id': doc.id,
                    'title': project_data.get('title', ''),
                    'description': project_data.get('description', '')[:100] + '...' if project_data.get('description', '') else '',
                    'thumbnail': project_data.get('thumbnail', None),
                    'tags': project_data.get('tags', []),
                    'created_by': project_data.get('created_by', '')
                })
        
        return jsonify(results[:20])
    
    except Exception as e:
        print(f"Error searching projects: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([]), 500
@app.context_processor
def inject_global_variables():
    """Make common variables available to all templates"""
    context = {}
    
    # Add current user avatar
    if 'user_id' in session:
        try:
            current_user_id = session['user_id']
            current_user_doc = db.collection('users').document(current_user_id).get()
            current_user_data = current_user_doc.to_dict() if current_user_doc.exists else {}
            
            # Get avatar
            avatar_url = generate_avatar_url(current_user_data)
            context['current_user_avatar'] = avatar_url
            
            # Get username
            context['current_username'] = current_user_data.get('display_username', 
                                                              current_user_data.get('username', ''))
            
            # Get notifications count
            try:
                notifications_ref = db.collection('users').document(current_user_id).collection('notifications')
                unread_notifications = notifications_ref.where('is_read', '==', False).get()
                context['notifications_count'] = len(list(unread_notifications))
            except Exception as e:
                print(f"Error getting notification count: {e}")
                context['notifications_count'] = 0
        except Exception as e:
            print(f"Error in global context processor: {e}")
            context['current_user_avatar'] = None
            context['current_username'] = None
            context['notifications_count'] = 0
    else:
        context['current_user_avatar'] = None
        context['current_username'] = None
        context['notifications_count'] = 0
    
    return context
@app.route('/notifications', methods=['GET'])
@login_required
def get_notifications():
    user_id = session['user_id']
    
    try:
        notifications_ref = db.collection('users').document(user_id).collection('notifications')
        notifications = notifications_ref.order_by('created_at', direction=firestore.Query.DESCENDING).limit(20).stream()
        
        notification_list = []
        for notification_doc in notifications:
            notification = notification_doc.to_dict()
            notification['id'] = notification_doc.id
            

            notification_type = notification['type']
            notification_type_info = ENHANCED_NOTIFICATION_TYPES.get(notification_type, {})
            

            notification['icon'] = notification_type_info.get('icon', '🔔')
            notification['type_label'] = notification_type_info.get('label', 'Notification')
            

            if notification_type in ['system', 'admin_message', 'important']:
                sender_config = notification_type_info.get('sender', {})
                notification['sender_info'] = {
                    'username': sender_config.get('username', 'Jinaq'),
                    'verified': sender_config.get('verified', True),
                    'verification_type': sender_config.get('verification_type', 'system'),
                    'avatar_url': url_for('static', filename='jinaq_logo.svg')
                }
            

            elif notification.get('sender_info'):
                sender_info = notification['sender_info']
                notification['sender_info']['avatar_url'] = sender_info.get('avatar_url') or generate_avatar_url({
                    'username': sender_info['username']
                })
                notification['sender_verified'] = sender_info.get('verified', False)
                notification['sender_verification_type'] = sender_info.get('verification_type')
            
            notification_list.append(notification)
        
        return jsonify(notification_list)
    except Exception as e:
        print(f"Error retrieving notifications: {e}")
        return jsonify({'error': 'Failed to retrieve notifications'}), 500

@app.route('/notifications/<notification_id>', methods=['DELETE'])
@login_required
def delete_notification(notification_id):
    """Delete a specific notification"""
    user_id = session['user_id']
    
    try:
        notification_ref = db.collection('users').document(user_id).collection('notifications').document(notification_id)
        notification_ref.delete()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error deleting notification: {e}")
        return jsonify({'error': 'Failed to delete notification'}), 500

@app.route('/notifications/mark_read/<notification_id>', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    """Mark a specific notification as read"""
    user_id = session['user_id']
    
    try:
        notification_ref = db.collection('users').document(user_id).collection('notifications').document(notification_id)
        notification_ref.update({'is_read': True})
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error marking notification as read: {e}")
        return jsonify({'error': 'Failed to mark notification as read'}), 500

@app.route('/admin/send_system_notification', methods=['POST'])
@login_required
def send_system_notification():

    ADMIN_IDS = ['vVbXL4LKGidXtrKnvqa21gWRY3V2'] 
    

    if session.get('user_id') not in ADMIN_IDS:
        return jsonify({'error': 'Unauthorized access'}), 403
    

    data = request.json
    
    recipient_type = data.get('recipient_type', 'all')
    message_type = data.get('message_type', 'system')
    message_text = data.get('message_text', '')
    selected_users = data.get('selected_users', [])
    
    if not message_text:
        return jsonify({'error': 'Message cannot be empty'}), 400
    
    try:

        admin_doc = db.collection('users').document(session['user_id']).get()
        admin_data = admin_doc.to_dict()
        

        users_ref = db.collection('users')
        
        if recipient_type == 'verified':
            query = users_ref.where('verified', '==', True)
        elif recipient_type == 'unverified':
            query = users_ref.where('verified', '==', False)
        elif recipient_type == 'selected':

            if not selected_users:
                return jsonify({'error': 'No users selected'}), 400
            
            notification_count = 0
            for user_id in selected_users:
                create_notification(
                    user_id, 
                    message_type, 
                    {
                        'message': message_text,
                        'sender': admin_data.get('display_username', 'Admin')
                    },
                    sender_id=session['user_id']
                )
                notification_count += 1
            
            return jsonify({
                'success': True, 
                'notifications_sent': notification_count
            })
        else:

            query = users_ref
        

        users = query.stream()
        
        notification_count = 0
        for user_doc in users:
            create_notification(
                user_doc.id, 
                message_type, 
                {
                    'message': message_text,
                    'sender': admin_data.get('display_username', 'Admin')
                },
                sender_id=session['user_id']
            )
            notification_count += 1
        
        return jsonify({
            'success': True, 
            'notifications_sent': notification_count
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/admin/toggle_user_block', methods=['POST'])
@login_required
def toggle_user_block():

    SUPER_ADMIN_IDS = ['vVbXL4LKGidXtrKnvqa21gWRY3V2']

    if session['user_id'] not in SUPER_ADMIN_IDS:
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    data = request.json
    user_id = data.get('user_id')
    
    try:

        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        user_data = user_doc.to_dict()
        

        current_blocked_status = user_data.get('blocked', False)
        new_blocked_status = not current_blocked_status
        

        user_ref.update({
            'blocked': new_blocked_status,
            'blocked_at': firestore.SERVER_TIMESTAMP if new_blocked_status else None,
            'blocked_by': session['user_id'] if new_blocked_status else None
        })
        

        create_notification(
            user_id, 
            'account_change', 
            {
                'message': f'Your account has been {"blocked" if new_blocked_status else "unblocked"} by an admin.',
                'action': 'account_blocked' if new_blocked_status else 'account_unblocked'
            },
            sender_id=session['user_id']
        )
        
        return jsonify({
            'success': True, 
            'blocked': new_blocked_status,
            'message': f'User has been {"blocked" if new_blocked_status else "unblocked"} successfully.'
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
# Add these routes to your Flask app
@app.route('/update-location', methods=['POST'])
@login_required
def update_location():
    try:
        data = request.get_json()
        coordinates = data.get('coordinates')
        
        if not coordinates:
            return jsonify({'success': False, 'error': 'No coordinates provided'}), 400

        geolocator = Nominatim(user_agent="your_app_name")
        location = geolocator.reverse(f"{coordinates['lat']}, {coordinates['lng']}")

        if location:
            location_data = {
                'city': location.raw.get('address', {}).get('city') or 
                       location.raw.get('address', {}).get('town') or 
                       location.raw.get('address', {}).get('municipality'),
                'country': location.raw.get('address', {}).get('country'),
                'coordinates': coordinates
            }

            # Update user's location in database
            db.collection('users').document(session['user_id']).update({
                'location': location_data
            })

            return jsonify({
                'success': True,
                'location': location_data
            })

        return jsonify({'success': False, 'error': 'Location not found'}), 404

    except Exception as e:
        print(f"Error updating location: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-ip-location')
@login_required
def get_ip_location():
    try:
        # Get IP address
        ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        
        # Use ip-api.com for IP geolocation (free tier)
        response = requests.get(f'http://ip-api.com/json/{ip}')
        data = response.json()
        
        if data['status'] == 'success':
            location_data = {
                'city': data.get('city'),
                'country': data.get('country'),
                'coordinates': {
                    'lat': data.get('lat'),
                    'lng': data.get('lon')
                }
            }

            # Update user's location in database
            db.collection('users').document(session['user_id']).update({
                'location': location_data
            })

            return jsonify({
                'success': True,
                'location': location_data
            })

        return jsonify({'success': False, 'error': 'Location not found'}), 404

    except Exception as e:
        print(f"Error getting IP location: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/update_header', methods=['POST'])
@login_required
def update_header():
    try:
        user_id = session['user_id']

        if 'header_image' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400

        file = request.files['header_image']
        position = request.form.get('position', '50% 50%')

        if not file or not file.filename:
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        # Get user doc to check for existing header
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict()

        # Delete old header if exists
        if user_data.get('header_image', {}).get('url'):
            try:
                old_path = user_data['header_image']['url'].split('/')[-1]
                old_blob = bucket.blob(f'headers/{user_id}/{old_path}')
                old_blob.delete()
            except Exception as e:
                print(f"Error deleting old header: {e}")

        # Upload new header
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{str(uuid.uuid4())}.{file_extension}"
        full_path = f"headers/{user_id}/{filename}"

        blob = bucket.blob(full_path)
        blob.upload_from_string(
            file.read(),
            content_type=file.content_type
        )

        blob.make_public()

        # Update user document with new header info
        header_data = {
            'header_image': {
                'url': blob.public_url,
                'position': position,
                'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
            }
        }

        db.collection('users').document(user_id).update(header_data)

        return jsonify({
            'success': True,
            'header': header_data['header_image']
        })

    except Exception as e:
        print(f"Error updating header: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500






@app.route('/<username>/like', methods=['POST'])
@login_required
def like_profile(username):
    try:
        # Find the target user
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username.lower()).limit(1).stream()
        user_doc = next(query, None)

        if user_doc is None:
            return jsonify({'success': False, 'error': 'User not found'}), 404

        target_user_id = user_doc.id
        current_user_id = session['user_id']

        # Prevent self-liking
        if target_user_id == current_user_id:
            return jsonify({'success': False, 'error': 'Cannot like your own profile'}), 400

        # Reference to target user document
        user_ref = db.collection('users').document(target_user_id)
        
        # Create a transaction
        transaction = db.transaction()

        # Firestore transaction to safely update likes
        @firestore.transactional
        def update_likes(transaction):
            snapshot = user_ref.get(transaction=transaction)
            
            # Retrieve likes directly from snapshot dictionary
            current_likes = snapshot.to_dict().get('likes', [])
            
            # Ensure current_likes is a list
            if not isinstance(current_likes, list):
                current_likes = []
            
            # Check if user has already liked
            if current_user_id in current_likes:
                # Unlike
                current_likes.remove(current_user_id)
            else:
                # Like
                current_likes.append(current_user_id)
            
            # Update document with new likes
            transaction.update(user_ref, {
                'likes': current_likes,
                'likes_count': len(current_likes)
            })
            
            return len(current_likes)

        # Execute the transaction
        likes_count = update_likes(transaction)

        # Create notification if profile is newly liked
        user_data = user_doc.to_dict()
        likes = user_data.get('likes', [])
        
        if current_user_id not in likes:
            create_notification(
                target_user_id, 
                'like_profile', 
                {
                    'message': 'Your profile was liked'
                },
                sender_id=current_user_id
            )

        return jsonify({
            'success': True, 
            'likes_count': likes_count
        })

    except Exception as e:
        print(f"Error in like_profile: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def get_cached_university_image(university_name):
    """Get cached image URL if available"""
    try:
        cache_ref = db.collection('image_cache').document('universities')
        cache_doc = cache_ref.get()
        
        if not cache_doc.exists:
            return None
        
        cache_data = cache_doc.to_dict()
        if university_name not in cache_data:
            return None
        
        # Check if cache is still valid (less than 7 days old)
        cache_entry = cache_data[university_name]
        timestamp = cache_entry.get('timestamp')
        
        if isinstance(timestamp, datetime.datetime):
            age = datetime.datetime.now(tz=datetime.timezone.utc) - timestamp
            if age.days < 7:
                return cache_entry.get('url')
        
        return None
    except Exception as e:
        print(f"Error retrieving cached image: {e}")
        return None

def extract_universities(text):
    """Extract university names from recommendation text"""
    universities = []
    
    if not text:
        return universities
    
    lines = text.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Check for university headers
        if "**University:" in line:
            university_name = line.replace('**University:', '').replace('**', '').strip()
            universities.append(university_name)
    
    return universities
# Add this function to synchronize profile data with the recommendations sidebar
@app.route('/update-recommendation-sidebar', methods=['POST'])
@login_required
def update_recommendation_sidebar():
    try:
        user_id = session['user_id']
        
        # Get complete user data from the database
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict() or {}
        
        # Get academic information
        academic_info = user_data.get('academic_info', {})
        
        # Get certificates
        certificates = []
        certs_query = db.collection('users').document(user_id).collection('certificates').stream()
        for cert_doc in certs_query:
            cert_data = cert_doc.to_dict()
            cert_data['id'] = cert_doc.id
            certificates.append(cert_data)
        
        # Prepare the response data
        response_data = {
            'academic': {
                'gpa': academic_info.get('gpa', ''),
                'sat_score': academic_info.get('sat_score', ''),
                'toefl_score': academic_info.get('toefl_score', ''),
                'ielts_score': academic_info.get('ielts_score', '')
            },
            'personal': {
                'specialty': user_data.get('specialty', ''),
                'age': user_data.get('age', ''),
                'education': user_data.get('education', ''),
                'goals': user_data.get('goals', ''),
                'bio': user_data.get('bio', '')
            },
            'skills': user_data.get('skills', []),
            'certificates': [
                {
                    'id': cert.get('id', ''),
                    'title': cert.get('title', ''),
                    'file_url': cert.get('file_url', '')
                } for cert in certificates
            ]
        }
        
        return jsonify({
            'success': True,
            'profile_data': response_data
        })
        
    except Exception as e:
        print(f"Error updating recommendation sidebar: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/get-recommendation-history', methods=['GET'])
@login_required
def get_recommendation_history():
    user_id = session['user_id']
    try:
        recommendations = db.collection('university_recommendations')\
            .where('user_id', '==', user_id)\
            .order_by('created_at', direction=firestore.Query.DESCENDING)\
            .limit(5)\
            .stream()
            
        history = []
        for rec in recommendations:
            rec_data = rec.to_dict()
            history.append({
                'id': rec.id,
                'country': rec_data['country'],
                'created_at': rec_data['created_at'],
                'universities': [u for u in rec_data['images'].keys()]
            })
            
        return jsonify({'history': history})
    except Exception as e:
        print(f"Error getting recommendation history: {e}")
        return jsonify({'error': str(e)}), 500
def format_datetime(value, format='%b %d, %Y'):
    """Custom datetime filter for Jinja2"""
    if value is None:
        return ""
    if not isinstance(value, datetime.datetime):
        return str(value)
    return value.strftime(format)

# Register the filter with Jinja2
app.jinja_env.filters['datetime'] = format_datetime


# Add this to app.py after other imports
from functools import wraps
from flask import g, render_template, redirect, url_for

# Create a maintenance mode middleware function
def check_maintenance_mode():
    """Middleware to check if site is in maintenance mode"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Skip maintenance check for admin routes and the maintenance page itself
            if (request.path.startswith('/admin') or 
                request.path == '/maintenance' or 
                request.path.startswith('/static')):
                return f(*args, **kwargs)
            
            try:
                # Get maintenance status from Firebase
                settings_ref = db.collection('system_settings').document('maintenance')
                settings_doc = settings_ref.get()
                
                if settings_doc.exists:
                    maintenance_data = settings_doc.to_dict()
                    if maintenance_data.get('enabled', False):
                        # Store maintenance message for template
                        g.maintenance_message = maintenance_data.get('message', 'The site is currently under maintenance. Please try again later.')
                        g.maintenance_details = maintenance_data.get('details', '')
                        g.maintenance_eta = maintenance_data.get('eta', '')
                        # Redirect to maintenance page
                        return redirect(url_for('maintenance_page'))
            except Exception as e:
                print(f"Error checking maintenance mode: {e}")
                
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Apply the decorator to Flask's before_request
@app.before_request
def check_site_maintenance():
    # Skip maintenance check for specific paths
    if (request.path.startswith('/admin') or 
        request.path == '/maintenance' or 
        request.path == '/maintenance-status' or
        request.path.startswith('/static')):
        return
    
    try:
        # Get maintenance status from Firebase
        settings_ref = db.collection('system_settings').document('maintenance')
        settings_doc = settings_ref.get()
        
        if settings_doc.exists:
            maintenance_data = settings_doc.to_dict()
            if maintenance_data.get('enabled', False):
                # Store maintenance message for template
                g.maintenance_message = maintenance_data.get('message', 'The site is currently under maintenance. Please try again later.')
                g.maintenance_details = maintenance_data.get('details', '')
                g.maintenance_eta = maintenance_data.get('eta', '')
                
                # Only redirect if not already on maintenance page
                if request.path != '/maintenance':
                    return redirect(url_for('maintenance_page'))
    except Exception as e:
        print(f"Error checking maintenance mode: {e}")

@app.route('/maintenance-status', methods=['GET'])
def maintenance_status():
    try:
        # Get maintenance status from Firebase
        settings_ref = db.collection('system_settings').document('maintenance')
        settings_doc = settings_ref.get()
        
        maintenance_active = False
        if settings_doc.exists:
            maintenance_data = settings_doc.to_dict()
            maintenance_active = maintenance_data.get('enabled', False)
            eta = maintenance_data.get('eta', '')
            
            # Check if ETA has passed and maintenance is still enabled
            if maintenance_active and eta:
                try:
                    # Parse MM/DD/YYYY HH:MM format
                    parts = eta.split(' ')
                    if len(parts) == 2:
                        date_parts = parts[0].split('/')
                        time_parts = parts[1].split(':')
                        
                        if len(date_parts) == 3 and len(time_parts) == 2:
                            month = int(date_parts[0])
                            day = int(date_parts[1])
                            year = int(date_parts[2])
                            hour = int(time_parts[0])
                            minute = int(time_parts[1])
                            
                            # Create Kazakhstan time (UTC+5)
                            kz_tz = datetime.timezone(datetime.timedelta(hours=5))
                            eta_date = datetime.datetime(year, month, day, hour, minute, tzinfo=kz_tz)
                            
                            # Get current time in Kazakhstan timezone
                            now = datetime.datetime.now(tz=kz_tz)
                            
                            print(f"Current time (KZ): {now}, ETA (KZ): {eta_date}")
                            
                            # If the ETA has passed, disable maintenance mode
                            if now >= eta_date:
                                print("ETA has passed, auto-disabling maintenance mode")
                                settings_ref.update({
                                    'enabled': False,
                                    'auto_disabled_at': datetime.datetime.now(tz=datetime.timezone.utc),
                                    'auto_disabled_reason': 'Scheduled maintenance time ended'
                                })
                                maintenance_active = False
                    else:
                        print(f"Invalid ETA format: {eta}")
                except Exception as date_error:
                    print(f"Error parsing ETA date: {date_error}")
        
        return jsonify({
            'maintenance_active': maintenance_active,
            'timestamp': datetime.datetime.now(tz=datetime.timezone.utc).isoformat()
        })
    except Exception as e:
        print(f"Error checking maintenance status: {e}")
        # Default to maintenance active in case of error to be safe
        return jsonify({'maintenance_active': True, 'error': str(e)})
@app.route('/maintenance')
def maintenance_page():
    try:
        # Directly fetch the maintenance data from the database
        settings_ref = db.collection('system_settings').document('maintenance')
        settings_doc = settings_ref.get()
        
        message = 'The site is currently under maintenance. Please try again later.'
        details = ''
        eta = ''
        
        if settings_doc.exists:
            maintenance_data = settings_doc.to_dict()
            message = maintenance_data.get('message', message)
            details = maintenance_data.get('details', details)
            eta = maintenance_data.get('eta', eta)
            
            # Print for debugging
            print(f"Maintenance data for page: message={message}, details={details}, eta={eta}")
        
        # Use the values directly, not from g
        return render_template('maintenance.html', 
                            message=message,
                            details=details,
                            eta=eta)
    except Exception as e:
        print(f"Error rendering maintenance page: {e}")
        # Fallback content
        return render_template('maintenance.html',
                            message='The site is currently under maintenance.',
                            details='',
                            eta='')

# Add admin routes to manage maintenance mode
@app.route('/admin/maintenance', methods=['GET'])
@login_required
def admin_maintenance():
    print(f"User trying to access maintenance: {session.get('user_id')}")
    print(f"ADMIN_IDS: {ADMIN_IDS}")
    
    # For testing, let's use direct comparison
    if session.get('user_id') != 'vVbXL4LKGidXtrKnvqa21gWRY3V2':
        print(f"Admin authentication failed - user:{session.get('user_id')}, required: {ADMIN_IDS}")
        return redirect(url_for('admin_login'))
    
    try:
        # Get current maintenance settings
        settings_ref = db.collection('system_settings').document('maintenance')
        settings_doc = settings_ref.get()
        
        maintenance_data = {}
        if settings_doc.exists:
            maintenance_data = settings_doc.to_dict()
        
        return render_template('admin_maintenance.html',
                             enabled=maintenance_data.get('enabled', False),
                             message=maintenance_data.get('message', ''),
                             details=maintenance_data.get('details', ''),
                             eta=maintenance_data.get('eta', ''),
                             last_updated=maintenance_data.get('last_updated'),
                             updated_by=maintenance_data.get('updated_by'))
    
    except Exception as e:
        print(f"Error in admin maintenance view: {e}")
        flash('Error loading maintenance settings')
        return redirect(url_for('admin_dashboard'))
@app.route('/admin/maintenance/toggle', methods=['POST'])
@login_required
def toggle_maintenance():
    # Ensure only admin can access
    if session.get('user_id') != 'vVbXL4LKGidXtrKnvqa21gWRY3V2':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        enabled = data.get('enabled', False)
        message = data.get('message', 'The site is currently under maintenance. Please try again later.')
        details = data.get('details', '')
        eta = data.get('eta', '')
        
        # Print the received data for debugging
        print(f"Received maintenance data: enabled={enabled}, message={message}, details={details}, eta={eta}")
        
        # Validate eta format if provided
        if eta:
            try:
                # Handle MM/DD/YYYY HH:MM format
                month, day, year = eta.split(' ')[0].split('/')
                hour, minute = eta.split(' ')[1].split(':')
                
                # Create date object to validate (already in Kazakhstan time)
                valid_date = datetime.datetime(int(year), int(month), int(day), int(hour), int(minute))
                
                # Format eta consistently for storage, using MM/DD/YYYY format
                eta = valid_date.strftime('%m/%d/%Y %H:%M')
                
                print(f"Maintenance ETA set to: {eta} (Kazakhstan time, UTC+5)")
            except Exception as e:
                print(f"Error validating eta format: {e}")
                return jsonify({'success': False, 'error': 'Invalid time format. Use MM/DD/YYYY HH:MM'}), 400
        
        # Get admin user info
        admin_doc = db.collection('users').document(session['user_id']).get()
        admin_data = admin_doc.to_dict()
        admin_username = admin_data.get('display_username', admin_data.get('username', 'Admin'))
        
        # Create maintenance settings document data
        maintenance_data = {
            'enabled': enabled,
            'message': message,
            'details': details,
            'eta': eta,
            'last_updated': datetime.datetime.now(tz=datetime.timezone.utc),
            'updated_by': {
                'user_id': session['user_id'],
                'username': admin_username
            }
        }
        
        # Update maintenance settings
        settings_ref = db.collection('system_settings').document('maintenance')
        settings_ref.set(maintenance_data)
        
        # Verify the data was saved correctly
        updated_doc = settings_ref.get()
        if updated_doc.exists:
            updated_data = updated_doc.to_dict()
            print(f"Saved maintenance data: {updated_data}")
        
        # Log the action
        log_admin_action('maintenance_mode', {
            'action': 'enabled' if enabled else 'disabled',
            'message': message,
            'details': details,
            'eta': eta
        })
        
        return jsonify({'success': True})
    
    except Exception as e:
        print(f"Error toggling maintenance mode: {e}")
        import traceback
        traceback.print_exc()  # Print full stack trace for debugging
        return jsonify({'success': False, 'error': str(e)}), 500
def log_admin_action(action_type, details):
    try:
        log_data = {
            'action_type': action_type,
            'details': details,
            'admin_id': session['user_id'],
            'timestamp': datetime.datetime.now(tz=datetime.timezone.utc),
            'ip_address': request.remote_addr
        }
        
        db.collection('admin_logs').add(log_data)
    except Exception as e:
        print(f"Error logging admin action: {e}")

@app.route('/projects/create', methods=['POST'])
@login_required
def create_project():
    user_id = session['user_id']
    
    try:
        # Get form data
        title = request.form.get('title')
        description = request.form.get('description')
        github_url = request.form.get('github_url', '')
        project_url = request.form.get('project_url', '')

        # Parse tags and collaborators
        tags = json.loads(request.form.get('tags', '[]'))
        
        try:
            collaborators_json = request.form.get('collaborators', '[]')
            collaborators_data = json.loads(collaborators_json)
            if not isinstance(collaborators_data, list):
                collaborators_data = []
        except Exception as json_error:
            print(f"Error parsing collaborators JSON: {json_error}")
            collaborators_data = []
        
        # Validate required fields
        if not title or not description:
            return jsonify({'success': False, 'error': 'Title and description are required'}), 400
        
        # Get user info for creator
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict()
        
        creator_info = {
            'user_id': user_id,
            'username': user_data.get('username', ''),
            'full_name': user_data.get('full_name', ''),
            'avatar': generate_avatar_url(user_data)
        }
        
        # Process collaborators
        collaborators = []
        for collab in collaborators_data:
            if not collab.get('user_id') or not collab.get('username'):
                continue
                
            collaborator = {
                'user_id': collab.get('user_id', ''),
                'username': collab.get('username', ''),
                'full_name': collab.get('full_name', ''),
                'avatar': collab.get('avatar', '')
            }
            collaborators.append(collaborator)
        
        # Process images
        image_urls = []
        
        # Get existing images if present
        existing_images = request.form.getlist('existing_images')
        if existing_images:
            image_urls.extend(existing_images)
        
        # Process thumbnail
        thumbnail_url = None
        if 'thumbnail' in request.files:
            thumbnail_file = request.files['thumbnail']
            if thumbnail_file and thumbnail_file.filename:
                # Create a unique filename
                file_extension = thumbnail_file.filename.rsplit('.', 1)[1].lower()
                filename = f"thumbnail_{str(uuid.uuid4())}.{file_extension}"
                
                # Save for later since we need the project ID
                thumbnail_data = {
                    'file': thumbnail_file,
                    'filename': filename,
                    'extension': file_extension
                }
        
        # Process header image
        header_image_url = None
        if 'header_image' in request.files:
            header_file = request.files['header_image']
            if header_file and header_file.filename:
                # Create a unique filename
                file_extension = header_file.filename.rsplit('.', 1)[1].lower()
                filename = f"header_{str(uuid.uuid4())}.{file_extension}"
                
                # Save for later since we need the project ID
                header_data = {
                    'file': header_file,
                    'filename': filename,
                    'extension': file_extension
                }
        
        # Upload new images
        for key in request.files:
            if key.startswith('image_'):
                image_file = request.files[key]
                
                if image_file and image_file.filename:
                    # Create a unique filename
                    file_extension = image_file.filename.rsplit('.', 1)[1].lower()
                    filename = f"projects/{user_id}/{str(uuid.uuid4())}.{file_extension}"
                    
                    # Upload to Firebase Storage
                    blob = bucket.blob(filename)
                    blob.upload_from_string(
                        image_file.read(),
                        content_type=image_file.content_type
                    )
                    
                    # Make public and get URL
                    blob.make_public()
                    image_urls.append(blob.public_url)
        
        # Create project document
        project_data = {
            'title': title,
            'title_lower': title.lower(),  # Add lowercase version for case-insensitive search
            'description': description,
            'github_url': github_url,
            'project_url': project_url,
            'tags': tags,
            'images': image_urls,
            'creator': creator_info,
            'collaborators': collaborators,
            'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
            'created_by': user_id,
            'updates': [],
            'last_updated': datetime.datetime.now(tz=datetime.timezone.utc)
        }
        
        # Save to Firestore
        project_ref = db.collection('projects').document()
        project_ref.set(project_data)
        project_id = project_ref.id
        
        # Now upload thumbnail if exists
        if 'thumbnail' in request.files and request.files['thumbnail'] and request.files['thumbnail'].filename:
            thumbnail_file = request.files['thumbnail']
            blob_path = f"projects/thumbnails/{project_id}/{thumbnail_data['filename']}"
            
            blob = bucket.blob(blob_path)
            blob.upload_from_string(
                thumbnail_file.read(),
                content_type=thumbnail_file.content_type
            )
            
            # Make public and get URL
            blob.make_public()
            thumbnail_url = blob.public_url
            
            # Update project with thumbnail URL
            project_ref.update({
                'thumbnail': thumbnail_url
            })
        
        # Now upload header image if exists
        if 'header_image' in request.files and request.files['header_image'] and request.files['header_image'].filename:
            header_file = request.files['header_image']
            blob_path = f"projects/headers/{project_id}/{header_data['filename']}"
            
            blob = bucket.blob(blob_path)
            blob.upload_from_string(
                header_file.read(),
                content_type=header_file.content_type
            )
            
            # Make public and get URL
            blob.make_public()
            header_image_url = blob.public_url
            
            # Update project with header URL
            project_ref.update({
                'header_image': header_image_url
            })
        
        # Add project reference to user document
        user_projects_ref = db.collection('users').document(user_id).collection('projects').document(project_id)
        user_projects_ref.set({
            'project_id': project_id,
            'created_at': datetime.datetime.now(tz=datetime.timezone.utc)
        })
        
        # Create notifications for collaborators
        for collaborator in collaborators:
            create_notification(
                collaborator['user_id'],
                'project_collaboration',
                {
                    'project_id': project_id,
                    'project_title': title,
                    'message': f"{user_data.get('display_username', user_data.get('username', ''))} added you as a collaborator on {title}"
                },
                sender_id=user_id
            )
        collaborators_data = json.loads(request.form.get('collaborators', '[]'))
        collaborators = process_collaborators(collaborators_data)        
        return jsonify({
            'success': True,
            'project_id': project_id,
            'message': 'Project created successfully'
        })
        
    except Exception as e:
        print(f"Error creating project: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def process_collaborators(collaborators_data):
    """Process collaborator data and ensure avatars are set properly"""
    collaborators = []
    for collab in collaborators_data:
        # Skip invalid collaborators missing user_id
        if not collab.get('user_id'):
            print(f"Warning: Skipping collaborator with missing user_id: {collab}")
            continue
            
        # Fetch collaborator's user data to get proper information
        try:
            user_doc = db.collection('users').document(collab.get('user_id')).get()
            if user_doc.exists:
                user_data = user_doc.to_dict()
                
                # Generate proper avatar URL
                avatar_url = generate_avatar_url(user_data)
                
                collaborator = {
                    'user_id': collab.get('user_id', ''),
                    'username': user_data.get('display_username', user_data.get('username', collab.get('username', ''))),
                    'full_name': user_data.get('full_name', collab.get('full_name', '')),
                    'avatar': avatar_url
                }
                collaborators.append(collaborator)
            else:
                # Fallback if user doc doesn't exist
                collaborator = {
                    'user_id': collab.get('user_id', ''),
                    'username': collab.get('username', ''),
                    'full_name': collab.get('full_name', ''),
                    'avatar': f"https://ui-avatars.com/api/?name={collab.get('username', 'U')[0]}&background=random&color=fff&size=128"
                }
                collaborators.append(collaborator)
        except Exception as e:
            print(f"Error processing collaborator {collab.get('user_id')}: {e}")
            # Add with fallback avatar
            collaborator = {
                'user_id': collab.get('user_id', ''),
                'username': collab.get('username', ''),
                'full_name': collab.get('full_name', ''),
                'avatar': "https://ui-avatars.com/api/?name=U&background=random&color=fff&size=128"
            }
            collaborators.append(collaborator)
    
    return collaborators
@app.route('/projects/<project_id>/invitation/<action>', methods=['POST'])
@login_required
def handle_project_invitation(project_id, action):
    user_id = session['user_id']
    
    if action not in ['accept', 'decline']:
        return jsonify({'success': False, 'error': 'Invalid action'}), 400
    
    try:
        # Get project document
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        project_title = project_data.get('title', 'Project')
        project_owner = project_data.get('created_by')
        
        # Check if user is in the pending collaborators
        collaborators = project_data.get('collaborators', [])
        found = False
        
        for i, collab in enumerate(collaborators):
            if collab.get('user_id') == user_id and collab.get('status', 'pending') == 'pending':
                found = True
                
                if action == 'accept':
                    # Update status to accepted
                    collaborators[i]['status'] = 'accepted'
                    
                    # Update project document
                    project_ref.update({
                        'collaborators': collaborators,
                        'last_updated': datetime.datetime.now(tz=datetime.timezone.utc)
                    })
                    
                    # Notify project owner
                    create_notification(
                        project_owner,
                        'project_collaboration',
                        {
                            'project_id': project_id,
                            'project_title': project_title,
                            'message': f"{session.get('username', 'Someone')} accepted your invitation to collaborate"
                        },
                        sender_id=user_id
                    )
                    
                    # Add project reference to user's collaborations
                    user_collab_ref = db.collection('users').document(user_id).collection('project_collaborations').document(project_id)
                    user_collab_ref.set({
                        'project_id': project_id,
                        'project_title': project_title,
                        'added_at': datetime.datetime.now(tz=datetime.timezone.utc),
                        'status': 'accepted'
                    })
                    
                    return jsonify({
                        'success': True, 
                        'message': f'You are now a collaborator on {project_title}'
                    })
                else:  # decline
                    # Remove user from collaborators
                    collaborators.pop(i)
                    
                    # Update project document
                    project_ref.update({
                        'collaborators': collaborators,
                        'last_updated': datetime.datetime.now(tz=datetime.timezone.utc)
                    })
                    
                    # Notify project owner
                    create_notification(
                        project_owner,
                        'project_collaboration',
                        {
                            'project_id': project_id,
                            'project_title': project_title,
                            'message': f"{session.get('username', 'Someone')} declined your invitation to collaborate"
                        },
                        sender_id=user_id
                    )
                    
                    return jsonify({
                        'success': True, 
                        'message': f'You declined the invitation to {project_title}'
                    })
        
        if not found:
            return jsonify({'success': False, 'error': 'Invitation not found or already processed'}), 404
            
    except Exception as e:
        print(f"Error handling project invitation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/projects/<project_id>')
def get_project(project_id):
    try:
        # Get project document
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        project_data['id'] = project_id
        
        return jsonify(project_data)
        
    except Exception as e:
        print(f"Error retrieving project: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/projects/<project_id>/update', methods=['POST'])
@login_required
def update_project(project_id):
    user_id = session['user_id']
    
    try:
        # Get project document
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        
        # Check if user is creator or collaborator
        if user_id != project_data['created_by'] and not any(c['user_id'] == user_id and c.get('status', '') == 'accepted' for c in project_data.get('collaborators', [])):
            return jsonify({'success': False, 'error': 'Not authorized to update this project'}), 403
        
        # Check if the user is the creator (only creators can modify collaborators)
        is_creator = user_id == project_data['created_by']
        
        # Get form data
        title = request.form.get('title')
        description = request.form.get('description')
        github_url = request.form.get('github_url', '')
        project_url = request.form.get('project_url', '')
        
        # Parse tags and collaborators
        tags = json.loads(request.form.get('tags', '[]'))
        
        # Only process collaborators if the user is the creator
        if is_creator:
            collaborators_data = json.loads(request.form.get('collaborators', '[]'))
            
            # Get existing collaborators for comparison
            existing_collaborators = {c.get('user_id'): c for c in project_data.get('collaborators', [])}
            
            # Process collaborators with status handling
            collaborators = []
            for collab in collaborators_data:
                # Skip invalid collaborators missing user_id
                if not collab.get('user_id'):
                    print(f"Warning: Skipping collaborator with missing user_id: {collab}")
                    continue
                
                # Check if this is an existing collaborator
                if collab.get('user_id') in existing_collaborators:
                    # Preserve existing status and data
                    existing_collab = existing_collaborators[collab.get('user_id')]
                    collaborator = {
                        'user_id': collab.get('user_id', ''),
                        'username': collab.get('username', ''),
                        'full_name': collab.get('full_name', ''),
                        'avatar': collab.get('avatar', ''),
                        'status': existing_collab.get('status', 'accepted')  # Default to accepted for backward compatibility
                    }
                else:
                    # New collaborator - set as pending
                    collaborator = {
                        'user_id': collab.get('user_id', ''),
                        'username': collab.get('username', ''),
                        'full_name': collab.get('full_name', ''),
                        'avatar': collab.get('avatar', ''),
                        'status': 'pending'  # New collaborators start as pending
                    }
                collaborators.append(collaborator)
            
            # Check for removed collaborators to notify them
            old_collaborators = {c.get('user_id') for c in project_data.get('collaborators', []) if c.get('user_id')}
            new_collaborators = {c.get('user_id') for c in collaborators if c.get('user_id')}
            
            removed_collaborators = old_collaborators - new_collaborators
            added_collaborators = new_collaborators - old_collaborators
            
            # Notify removed collaborators
            for collab_id in removed_collaborators:
                if not collab_id:  # Skip empty IDs
                    continue
                    
                create_notification(
                    collab_id,
                    'project_collaboration',
                    {
                        'project_id': project_id,
                        'project_title': title,
                        'message': f"You were removed as a collaborator from {title}"
                    },
                    sender_id=user_id
                )
                
                # Also remove project reference from their document
                try:
                    collab_ref = db.collection('users').document(collab_id).collection('project_collaborations').document(project_id)
                    collab_ref.delete()
                except Exception as e:
                    print(f"Error removing project reference for user {collab_id}: {e}")
            
            # Send invitations to newly added collaborators
            for collab_id in added_collaborators:
                if not collab_id:  # Skip empty IDs
                    continue
                    
                # Find the collaborator's username
                collab_data = next((c for c in collaborators if c.get('user_id') == collab_id), None)
                
                create_notification(
                    collab_id,
                    'project_invitation',  # New notification type
                    {
                        'project_id': project_id,
                        'project_title': title,
                        'message': f"You've been invited to collaborate on {title}",
                        'action': 'invitation'
                    },
                    sender_id=user_id
                )
        else:
            # If not creator, keep existing collaborators
            collaborators = project_data.get('collaborators', [])
                    # Update project document with new data

        # Validate required fields
        if not title or not description:
            return jsonify({'success': False, 'error': 'Title and description are required'}), 400
        
        # Process images
        image_urls = []
        
        # Get existing images if present
        existing_images = request.form.getlist('existing_images')
        if existing_images:
            image_urls.extend(existing_images)
        
        # Process thumbnail if provided
        thumbnail_url = project_data.get('thumbnail')
        if 'thumbnail' in request.files:
            thumbnail_file = request.files['thumbnail']
            if thumbnail_file and thumbnail_file.filename:
                # Delete old thumbnail if exists
                if thumbnail_url:
                    try:
                        old_filename = thumbnail_url.split('/')[-1]
                        old_blob = bucket.blob(f"projects/thumbnails/{project_id}/{old_filename}")
                        old_blob.delete()
                    except Exception as e:
                        print(f"Error deleting old thumbnail: {e}")
                
                # Upload new thumbnail
                file_extension = thumbnail_file.filename.rsplit('.', 1)[1].lower()
                filename = f"thumbnail_{str(uuid.uuid4())}.{file_extension}"
                blob_path = f"projects/thumbnails/{project_id}/{filename}"
                
                blob = bucket.blob(blob_path)
                blob.upload_from_string(
                    thumbnail_file.read(),
                    content_type=thumbnail_file.content_type
                )
                
                # Make public and get URL
                blob.make_public()
                thumbnail_url = blob.public_url
        
        # Process header image if provided
        header_image_url = project_data.get('header_image')
        if 'header_image' in request.files:
            header_file = request.files['header_image']
            if header_file and header_file.filename:
                # Delete old header if exists
                if header_image_url:
                    try:
                        old_filename = header_image_url.split('/')[-1]
                        old_blob = bucket.blob(f"projects/headers/{project_id}/{old_filename}")
                        old_blob.delete()
                    except Exception as e:
                        print(f"Error deleting old header: {e}")
                
                # Upload new header
                file_extension = header_file.filename.rsplit('.', 1)[1].lower()
                filename = f"header_{str(uuid.uuid4())}.{file_extension}"
                blob_path = f"projects/headers/{project_id}/{filename}"
                
                blob = bucket.blob(blob_path)
                blob.upload_from_string(
                    header_file.read(),
                    content_type=header_file.content_type
                )
                
                # Make public and get URL
                blob.make_public()
                header_image_url = blob.public_url
        
        # Upload new images
        for key in request.files:
            if key.startswith('image_'):
                image_file = request.files[key]
                
                if image_file and image_file.filename:
                    # Create a unique filename
                    file_extension = image_file.filename.rsplit('.', 1)[1].lower()
                    filename = f"projects/{user_id}/{str(uuid.uuid4())}.{file_extension}"
                    
                    # Upload to Firebase Storage
                    blob = bucket.blob(filename)
                    blob.upload_from_string(
                        image_file.read(),
                        content_type=image_file.content_type
                    )
                    
                    # Make public and get URL
                    blob.make_public()
                    image_urls.append(blob.public_url)
        
        # Update project document
        update_data = {
                'title': title,
                'description': description,
                'github_url': github_url,
                'project_url': project_url,
                'tags': tags,
                'collaborators': collaborators,
                'last_updated': datetime.datetime.now(tz=datetime.timezone.utc)
            }
        
        # Add thumbnail and header if they exist
        if thumbnail_url:
            update_data['thumbnail'] = thumbnail_url
        
        if header_image_url:
            update_data['header_image'] = header_image_url
        
        project_ref.update(update_data)
        
        # Notify active collaborators about project update
        active_collaborators = [c.get('user_id') for c in collaborators 
                          if c.get('user_id') and c.get('user_id') != user_id 
                          and c.get('status', '') == 'accepted']
        
        # Also notify the creator if update was made by collaborator
        if not is_creator:
            active_collaborators.append(project_data['created_by'])
        
        # Send notifications
        for collab_id in active_collaborators:
            create_notification(
                collab_id,
                'project_update',
                {
                    'project_id': project_id,
                    'project_title': title,
                    'message': f"Project {title} has been updated"
                },
                sender_id=user_id
            )
        
        return jsonify({
            'success': True,
            'message': 'Project updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating project: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
# Add this route to handle optimizing small avatars and caching them for better display
@app.route('/optimized-avatar/<user_id>/<size>')
def optimized_avatar(user_id, size):
    """Serve optimized avatar for a specific user in the requested size
    Sizes: 'small' (32px), 'medium' (64px), 'large' (128px)"""
    try:
        # Get user data
        user_doc = db.collection('users').document(user_id).get()
        if not user_doc.exists:
            # Return fallback avatar
            fallback_url = f"https://ui-avatars.com/api/?name=U&background=random&color=fff&size=256&bold=true"
            return redirect(fallback_url)
        
        user_data = user_doc.to_dict()
        
        # Get avatar URL
        avatar_url = user_data.get('avatar_url')
        
        if not avatar_url:
            # Generate UI Avatar with higher resolution
            display_name = user_data.get('display_username', user_data.get('username', 'U'))
            initials = ''.join(word[0].upper() for word in display_name.split()[:2])
            fallback_url = f"https://ui-avatars.com/api/?name={initials}&background=random&color=fff&size=256&bold=true"
            return redirect(fallback_url)
        
        # For uploaded avatars, we could resize them here using PIL if needed
        # For now, just redirect to the original avatar URL
        return redirect(avatar_url)
        
    except Exception as e:
        print(f"Error serving optimized avatar: {e}")
        # Return fallback avatar
        return redirect(f"https://ui-avatars.com/api/?name=U&background=random&color=fff&size=256&bold=true")
@app.route('/admin/repair_project_collaborators', methods=['GET'])
@login_required
def repair_project_collaborators():
    # Only allow admin or specific users to run this
    if session.get('user_id') != 'vVbXL4LKGidXtrKnvqa21gWRY3V2':  # Your admin ID
        return "Unauthorized", 403
    
    try:
        # Get all projects
        projects_ref = db.collection('projects')
        projects = projects_ref.stream()
        
        fixed_count = 0
        
        for project_doc in projects:
            project_data = project_doc.to_dict()
            project_id = project_doc.id
            
            if 'collaborators' not in project_data or not project_data['collaborators']:
                continue
            
            # Process and fix collaborators
            fixed_collaborators = []
            changed = False
            
            for collab in project_data['collaborators']:
                if not collab.get('user_id'):
                    continue
                
                try:
                    user_doc = db.collection('users').document(collab['user_id']).get()
                    if user_doc.exists:
                        user_data = user_doc.to_dict()
                        
                        # Generate proper avatar URL
                        avatar_url = generate_avatar_url(user_data)
                        
                        # Check if avatar needs to be updated
                        if collab.get('avatar') != avatar_url:
                            changed = True
                        
                        fixed_collab = {
                            'user_id': collab['user_id'],
                            'username': user_data.get('display_username', user_data.get('username', collab.get('username', ''))),
                            'full_name': user_data.get('full_name', collab.get('full_name', '')),
                            'avatar': avatar_url
                        }
                    else:
                        # If user no longer exists, keep original with fallback avatar
                        fallback_avatar = f"https://ui-avatars.com/api/?name={collab.get('username', 'U')[0]}&background=random&color=fff&size=128"
                        
                        if collab.get('avatar') != fallback_avatar:
                            changed = True
                        
                        fixed_collab = {
                            'user_id': collab['user_id'],
                            'username': collab.get('username', ''),
                            'full_name': collab.get('full_name', ''),
                            'avatar': fallback_avatar
                        }
                except Exception as e:
                    print(f"Error repairing collaborator {collab.get('user_id')} for project {project_id}: {e}")
                    # Keep original but with fallback avatar if missing
                    fixed_collab = collab.copy()
                    if not collab.get('avatar'):
                        fixed_collab['avatar'] = "https://ui-avatars.com/api/?name=U&background=random&color=fff&size=128"
                        changed = True
                
                fixed_collaborators.append(fixed_collab)
            
            # Update the project if changes were made
            if changed:
                projects_ref.document(project_id).update({
                    'collaborators': fixed_collaborators
                })
                fixed_count += 1
        
        return f"Successfully repaired collaborator avatars in {fixed_count} projects."
    
    except Exception as e:
        print(f"Error repairing collaborators: {e}")
        import traceback
        traceback.print_exc()
        return f"Error during repair: {str(e)}", 500    
@app.route('/projects/<project_id>/update_header', methods=['POST'])
@login_required
def update_project_header(project_id):
    user_id = session['user_id']
    
    try:
        # Check if project exists and user has permission
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        
        # Check if user is creator or collaborator
        is_creator = user_id == project_data.get('created_by')
        is_collaborator = any(c.get('user_id') == user_id for c in project_data.get('collaborators', []))
        
        if not (is_creator or is_collaborator):
            return jsonify({'success': False, 'error': 'Not authorized to update this project'}), 403
        
        # Check if header_image is in request
        if 'header_image' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        header_file = request.files['header_image']
        if not header_file or not header_file.filename:
            return jsonify({'success': False, 'error': 'No file selected'}), 400
            
        # Check file size (5MB max)
        if len(header_file.read()) > 5 * 1024 * 1024:
            return jsonify({'success': False, 'error': 'File size exceeds 5MB limit'}), 400
            
        # Reset file pointer after reading
        header_file.seek(0)
        
        # Delete old header image if it exists
        if project_data.get('header_image'):
            try:
                # Extract filename from URL
                old_filename = project_data['header_image'].split('/')[-1]
                old_blob = bucket.blob(f'projects/headers/{project_id}/{old_filename}')
                old_blob.delete()
            except Exception as e:
                print(f"Error deleting old header: {e}")
        
        # Upload new header image
        file_extension = header_file.filename.rsplit('.', 1)[1].lower()
        filename = f"header_{str(uuid.uuid4())}.{file_extension}"
        blob_path = f"projects/headers/{project_id}/{filename}"
        
        blob = bucket.blob(blob_path)
        blob.upload_from_string(
            header_file.read(),
            content_type=header_file.content_type
        )
        
        # Make file publicly accessible
        blob.make_public()
        header_url = blob.public_url
        
        # Update project document
        project_ref.update({
            'header_image': header_url,
            'last_updated': datetime.datetime.now(tz=datetime.timezone.utc)
        })
        
        return jsonify({
            'success': True,
            'header_url': header_url,
            'message': 'Header image updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating project header: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/projects/<project_id>/update_thumbnail', methods=['POST'])
@login_required
def update_project_thumbnail(project_id):
    user_id = session['user_id']
    
    try:
        # Check if project exists and user has permission
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        
        # Check if user is creator or collaborator
        is_creator = user_id == project_data.get('created_by')
        is_collaborator = any(c.get('user_id') == user_id for c in project_data.get('collaborators', []))
        
        if not (is_creator or is_collaborator):
            return jsonify({'success': False, 'error': 'Not authorized to update this project'}), 403
        
        # Check if thumbnail is in request
        if 'thumbnail' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        thumbnail_file = request.files['thumbnail']
        if not thumbnail_file or not thumbnail_file.filename:
            return jsonify({'success': False, 'error': 'No file selected'}), 400
            
        # Check file size (2MB max)
        if len(thumbnail_file.read()) > 2 * 1024 * 1024:
            return jsonify({'success': False, 'error': 'File size exceeds 2MB limit'}), 400
            
        # Reset file pointer after reading
        thumbnail_file.seek(0)
        
        # Delete old thumbnail if it exists
        if project_data.get('thumbnail'):
            try:
                # Extract filename from URL
                old_filename = project_data['thumbnail'].split('/')[-1]
                old_blob = bucket.blob(f'projects/thumbnails/{project_id}/{old_filename}')
                old_blob.delete()
            except Exception as e:
                print(f"Error deleting old thumbnail: {e}")
        
        # Upload new thumbnail
        file_extension = thumbnail_file.filename.rsplit('.', 1)[1].lower()
        filename = f"thumbnail_{str(uuid.uuid4())}.{file_extension}"
        blob_path = f"projects/thumbnails/{project_id}/{filename}"
        
        blob = bucket.blob(blob_path)
        blob.upload_from_string(
            thumbnail_file.read(),
            content_type=thumbnail_file.content_type
        )
        
        # Make file publicly accessible
        blob.make_public()
        thumbnail_url = blob.public_url
        
        # Update project document
        project_ref.update({
            'thumbnail': thumbnail_url,
            'last_updated': datetime.datetime.now(tz=datetime.timezone.utc)
        })
        
        return jsonify({
            'success': True,
            'thumbnail_url': thumbnail_url,
            'message': 'Thumbnail updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating project thumbnail: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/projects/<project_id>/updates', methods=['POST'])
@login_required
def add_project_update(project_id):
    user_id = session['user_id']
    
    try:
        # Get project document
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        
        # Check if user is creator or collaborator
        is_collaborator = (user_id == project_data['created_by']) or \
                          any(c['user_id'] == user_id for c in project_data.get('collaborators', []))
                          
        if not is_collaborator:
            return jsonify({'success': False, 'error': 'Not authorized to update this project'}), 403
        
        # Get form data
        title = request.form.get('title')
        update_type = request.form.get('type', 'Update')
        content = request.form.get('content')
        
        # Validate
        if not title or not content:
            return jsonify({'success': False, 'error': 'Title and content are required'}), 400
        
        # Get user info
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict()
        
        # Create update object
        update_data = {
            'title': title,
            'type': update_type,
            'content': content,
            'author_id': user_id,
            'author_username': user_data.get('display_username', user_data.get('username', '')),
            'author_avatar': generate_avatar_url(user_data),
            'created_at': datetime.datetime.now(tz=datetime.timezone.utc)
        }
        
        # Add update to project
        updates = project_data.get('updates', [])
        updates.insert(0, update_data)  # Add to the beginning to show newest first
        
        project_ref.update({
            'updates': updates,
            'last_updated': datetime.datetime.now(tz=datetime.timezone.utc)
        })
        
        # Notify creator and collaborators about the update
        notified_users = set()
        
        # Add creator to notification list if not the author
        if project_data['created_by'] != user_id:
            notified_users.add(project_data['created_by'])
        
        # Add all collaborators except the author
        for collab in project_data.get('collaborators', []):
            if collab['user_id'] != user_id:
                notified_users.add(collab['user_id'])
        
        # Send notifications
        for recipient_id in notified_users:
            create_notification(
                recipient_id,
                'project_update',
                {
                    'project_id': project_id,
                    'project_title': project_data['title'],
                    'update_title': title,
                    'update_type': update_type,
                    'message': f"New {update_type.lower()} posted on {project_data['title']}: {title}"
                },
                sender_id=user_id
            )
        
        return jsonify({
            'success': True,
            'message': 'Update added successfully'
        })
        
    except Exception as e:
        print(f"Error adding project update: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
      
@app.route('/projects/<project_id>/delete', methods=['DELETE'])
@login_required
def delete_project(project_id):
    user_id = session['user_id']
    
    try:
        # Get project document
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        
        # Check if user is creator
        if user_id != project_data['created_by']:
            return jsonify({'success': False, 'error': 'Not authorized to delete this project'}), 403
        
        # Delete project images from storage
        for image_url in project_data.get('images', []):
            try:
                path = image_url.split('/')[-1]
                blob = bucket.blob(f"projects/{user_id}/{path}")
                blob.delete()
            except Exception as e:
                print(f"Error deleting project image: {e}")
        
        # Delete project reference from user document
        user_projects_ref = db.collection('users').document(user_id).collection('projects').document(project_id)
        user_projects_ref.delete()
        
        # Delete project document
        project_ref.delete()
        
        # Notify collaborators
        for collaborator in project_data.get('collaborators', []):
            create_notification(
                collaborator['user_id'],
                'project_collaboration',
                {
                    'project_title': project_data['title'],
                    'message': f"Project {project_data['title']} has been deleted by the creator"
                },
                sender_id=user_id
            )
        
        return jsonify({
            'success': True,
            'message': 'Project deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting project: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/view_project/<project_id>')
def view_project(project_id):
    try:
        # Get project document
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            flash('Project not found')
            return redirect(url_for('index'))
            
        project_data = project_doc.to_dict()
        project_data['id'] = project_id
        
        # Check if current user is a collaborator
        is_collaborator = False
        if 'user_id' in session:
            current_user_id = session['user_id']
            
            # User is collaborator if they are the creator or listed as a collaborator
            is_collaborator = (current_user_id == project_data['created_by']) or \
                             any(c['user_id'] == current_user_id for c in project_data.get('collaborators', []))
        
        return render_template('project_view.html',
                             project=project_data,
                             is_collaborator=is_collaborator,
                             current_user_id=session.get('user_id'))
        
    except Exception as e:
        print(f"Error viewing project: {e}")
        flash('Error loading project')
        return redirect(url_for('index'))


@app.route('/projects/<project_id>/comments', methods=['GET', 'POST'])
@login_required
def project_comments(project_id):
    try:
        # Check if project exists
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'error': 'Project not found'}), 404
            
        # POST request - add a new comment
        if request.method == 'POST':
            comment_text = request.form.get('comment')
            if not comment_text or len(comment_text.strip()) == 0:
                return jsonify({'error': 'Comment text is required'}), 400
            
            if len(comment_text) > 500:
                return jsonify({'error': 'Comment is too long'}), 400

            # Create comment data
            comment_data = {
                'author_id': session['user_id'],
                'author_username': session['username'],
                'text': comment_text.strip(),
                'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
                'likes': []
            }

            # Add comment to project's comments collection
            comment_ref = db.collection('projects').document(project_id).collection('comments').add(comment_data)
            
            # Get author info for response
            author_doc = db.collection('users').document(session['user_id']).get()
            author_data = author_doc.to_dict()
            
            # Notify project owner if the commenter is not the owner
            project_data = project_doc.to_dict()
            owner_id = project_data.get('created_by')
            
            if session['user_id'] != owner_id:
                create_notification(
                    owner_id,
                    'project_comment',
                    {
                        'project_id': project_id,
                        'project_title': project_data.get('title', 'Project'),
                        'comment_text': comment_text.strip()
                    },
                    sender_id=session['user_id'],
                    related_id=comment_ref[1].id
                )

            # Format response
            response_data = {
                'success': True,
                'comment': {
                    'id': comment_ref[1].id,
                    **comment_data,
                    'author_avatar': generate_avatar_url(author_data),
                    'author_verified': author_data.get('verified', False),
                    'author_verification_type': author_data.get('verification_type'),
                    'likes_count': 0,
                    'is_liked': False
                }
            }

            return jsonify(response_data), 201
        
        # GET request - fetch comments
        else:
            comments_ref = db.collection('projects').document(project_id).collection('comments')
            comments_query = comments_ref.order_by('created_at', direction=firestore.Query.DESCENDING)

            all_comments = comments_query.stream()
            
            # Process comments and replies
            main_comments = {}
            replies = {}

            for comment_doc in all_comments:
                comment = comment_doc.to_dict()
                comment['id'] = comment_doc.id

                # Get author info
                author_doc = db.collection('users').document(comment['author_id']).get()
                author_data = author_doc.to_dict() if author_doc.exists else None
                
                comment['author_avatar'] = generate_avatar_url(author_data) if author_data else url_for('static', filename='default-avatar.png')
                comment['likes_count'] = len(comment.get('likes', []))
                comment['is_liked'] = session.get('user_id') in comment.get('likes', [])
                comment['can_delete'] = session.get('user_id') == comment['author_id'] or \
                                        session.get('user_id') == project_doc.to_dict().get('created_by')

                if 'parent_id' in comment:
                    if comment['parent_id'] not in replies:
                        replies[comment['parent_id']] = []
                    replies[comment['parent_id']].append(comment)
                else:
                    main_comments[comment_doc.id] = comment

            # Add replies to main comments
            for comment_id, comment in main_comments.items():
                if comment_id in replies:
                    sorted_replies = sorted(replies[comment_id], key=lambda x: x['created_at'])
                    comment['replies'] = sorted_replies

            return jsonify(list(main_comments.values()))

    except Exception as e:
        print(f"Error in project_comments route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/projects/<project_id>/comments/<comment_id>', methods=['DELETE'])
@login_required
def delete_project_comment(project_id, comment_id):
    try:
        # Get project and comment
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'error': 'Project not found'}), 404
            
        comment_ref = project_ref.collection('comments').document(comment_id)
        comment_doc = comment_ref.get()
        
        if not comment_doc.exists:
            return jsonify({'error': 'Comment not found'}), 404
        
        comment_data = comment_doc.to_dict()
        project_data = project_doc.to_dict()
        
        # Check if user is authorized to delete (comment author or project owner)
        if comment_data['author_id'] != session['user_id'] and project_data.get('created_by') != session['user_id']:
            return jsonify({'error': 'Unauthorized to delete this comment'}), 403
        
        # Delete comment
        comment_ref.delete()
        
        # Also delete any replies to this comment
        replies_query = project_ref.collection('comments').where('parent_id', '==', comment_id).get()
        for reply in replies_query:
            reply.reference.delete()
        
        return jsonify({'success': True, 'message': 'Comment deleted successfully'})
        
    except Exception as e:
        print(f"Error deleting project comment: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/projects/<project_id>/comments/<comment_id>/like', methods=['POST'])
@login_required
def like_project_comment(project_id, comment_id):
    try:
        # Check if project exists
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'error': 'Project not found'}), 404
            
        # Get comment
        comment_ref = project_ref.collection('comments').document(comment_id)
        comment_doc = comment_ref.get()
        
        if not comment_doc.exists:
            return jsonify({'error': 'Comment not found'}), 404
            
        comment_data = comment_doc.to_dict()
        current_user_id = session['user_id']
        
        # Toggle like status
        likes = comment_data.get('likes', [])
        
        if current_user_id in likes:
            likes.remove(current_user_id)
        else:
            likes.append(current_user_id)
            
            # Notify comment author about like (if it's not their own comment)
            if current_user_id != comment_data['author_id']:
                create_notification(
                    comment_data['author_id'],
                    'like_comment',
                    {
                        'project_id': project_id,
                        'project_title': project_doc.to_dict().get('title', 'Project'),
                        'comment_text': comment_data['text']
                    },
                    sender_id=current_user_id,
                    related_id=comment_id
                )
        
        # Update comment
        comment_ref.update({'likes': likes})
        
        return jsonify({
            'success': True,
            'likes_count': len(likes),
            'is_liked': current_user_id in likes
        })
        
    except Exception as e:
        print(f"Error liking project comment: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/projects/<project_id>/comments/<comment_id>/reply', methods=['POST'])
@login_required
def reply_to_project_comment(project_id, comment_id):
    try:
        # Check if project exists
        project_ref = db.collection('projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'error': 'Project not found'}), 404
            
        # Get original comment
        comment_ref = project_ref.collection('comments').document(comment_id)
        comment_doc = comment_ref.get()
        
        if not comment_doc.exists:
            return jsonify({'error': 'Comment not found'}), 404
            
        # Get reply text
        reply_text = request.form.get('reply')
        if not reply_text or len(reply_text.strip()) == 0:
            return jsonify({'error': 'Reply text is required'}), 400
            
        if len(reply_text) > 500:
            return jsonify({'error': 'Reply is too long'}), 400
            
        # Get original comment data
        comment_data = comment_doc.to_dict()
        
        # Prepare reply chain
        reply_chain = comment_data.get('reply_chain', [])
        if 'parent_id' in comment_data:
            # This is a reply to a reply
            reply_chain = comment_data.get('reply_chain', [])
        
        # Add current comment to the chain
        reply_chain.append(comment_id)
        
        # Create reply data
        reply_data = {
            'author_id': session['user_id'],
            'author_username': session['username'],
            'text': reply_text.strip(),
            'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
            'parent_id': comment_id,
            'reply_chain': reply_chain,
            'reply_to_username': comment_data['author_username'],
            'reply_level': len(reply_chain),
            'likes': []
        }
        
        # Add reply to project's comments collection
        reply_ref = project_ref.collection('comments').add(reply_data)
        
        # Get author info for response
        author_doc = db.collection('users').document(session['user_id']).get()
        author_data = author_doc.to_dict()
        
        # Notify the original comment author if not the same person
        if session['user_id'] != comment_data['author_id']:
            create_notification(
                comment_data['author_id'],
                'reply_comment',
                {
                    'project_id': project_id,
                    'project_title': project_doc.to_dict().get('title', 'Project'),
                    'original_comment_text': comment_data['text'],
                    'reply_text': reply_text.strip(),
                    'reply_chain': reply_chain
                },
                sender_id=session['user_id'],
                related_id=comment_id
            )
        
        # Format response
        response_data = {
            'success': True,
            'reply': {
                'id': reply_ref[1].id,
                **reply_data,
                'author_avatar': generate_avatar_url(author_data),
                'author_verified': author_data.get('verified', False),
                'author_verification_type': author_data.get('verification_type'),
                'likes_count': 0,
                'is_liked': False
            }
        }
        
        return jsonify(response_data), 201
        
    except Exception as e:
        print(f"Error in reply_to_project_comment: {e}")
        return jsonify({'error': str(e)}), 500
# Configure longer session lifetime (120 days)
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=120)
@app.route('/notifications-page')
@login_required
def notifications_page():
    user_id = session['user_id']
    
    try:
        # Get the active tab from the query parameter, default to 'notifications'
        tab = request.args.get('tab', 'notifications')
        
        # Handle different tabs
        if tab == 'mentions':
            # Get mentions (users who mentioned this user)
            items = []
            # For now, this is placeholder logic
            # Later you can implement a system to track when users are mentioned with @username
        
        elif tab == 'launch_tags':
            # Get launch tags (similar to Product Hunt's launch notifications)
            items = []
            # For now, this is placeholder logic
            # Later you can implement tagging users in project launches
        
        else:  # Default to notifications tab
            # Fetch notifications from the user's collection
            notifications_ref = db.collection('users').document(user_id).collection('notifications')
            notifications = notifications_ref.order_by('created_at', direction=firestore.Query.DESCENDING).limit(50).stream()
            
            items = []
            for notification_doc in notifications:
                notification = notification_doc.to_dict()
                notification['id'] = notification_doc.id
                
                # Ensure consistent notification structure
                if 'content' not in notification:
                    notification['content'] = {}
                
                # Process notification type info
                notification_type = notification.get('type', 'system')
                notification_type_info = ENHANCED_NOTIFICATION_TYPES.get(notification_type, {})
                
                notification['icon'] = notification_type_info.get('icon', '🔔')
                notification['type_label'] = notification_type_info.get('label', 'Notification')
                
                # Handle sender info for different notification types
                if notification_type in ['system', 'admin_message', 'important']:
                    sender_config = notification_type_info.get('sender', {})
                    notification['sender_info'] = {
                        'username': sender_config.get('username', 'Jinaq'),
                        'verified': sender_config.get('verified', True),
                        'verification_type': sender_config.get('verification_type', 'system'),
                        'avatar_url': url_for('static', filename='jinaq_logo.svg')
                    }
                
                # Process avatar URLs for sender
                if notification.get('sender_info'):
                    sender_info = notification['sender_info']
                    if not sender_info.get('avatar_url'):
                        sender_info['avatar_url'] = generate_avatar_url({
                            'username': sender_info.get('username', 'User')
                        })
                
                # Add items to the list
                items.append(notification)
            
            # Mark notifications as read in batches
            try:
                unread_items = [item for item in items if not item.get('is_read')]
                if unread_items:
                    batch = db.batch()
                    for item in unread_items:
                        notification_ref = db.collection('users').document(user_id).collection('notifications').document(item['id'])
                        batch.update(notification_ref, {'is_read': True})
                    batch.commit()
                    
                    # Update the items in memory too
                    for item in items:
                        item['is_read'] = True
            except Exception as e:
                print(f"Error marking notifications as read: {e}")
        
        # Get unread count for the badge (this will be 0 after viewing, but needed for other pages)
        unread_notifications_count = 0
        try:
            unread_notifications = db.collection('users').document(user_id).collection('notifications').where('is_read', '==', False).get()
            unread_notifications_count = len(list(unread_notifications))
        except Exception as e:
            print(f"Error getting unread count: {e}")
        
        # Get current user info for the template
        current_user_avatar = get_current_user_avatar()
        current_username = get_current_username()
        
        return render_template(
            'notifications_page.html',
            items=items,
            current_tab=tab,
            unread_notifications_count=unread_notifications_count,
            current_user_avatar=current_user_avatar,
            current_username=current_username
        )
    
    except Exception as e:
        print(f"Error in notifications page: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading notifications')
        return redirect(url_for('index'))

def get_unread_notifications_count(user_id):
    try:
        if not user_id:
            return 0
            
        notifications_ref = db.collection('users').document(user_id).collection('notifications')
        unread_notifications = notifications_ref.where('is_read', '==', False).get()
        return len(list(unread_notifications))
    except Exception as e:
        print(f"Error getting unread notifications count: {e}")
        return 0



# Вспомогательная функция для создания слагов
def slugify(text):
    """Преобразует текст в URL-дружественный формат (слаг)"""
    import re
    import unicodedata
    
    # Преобразуем в нижний регистр и заменяем пробелы дефисами
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    text = re.sub(r'^-+|-+$', '', text)
    
    return text








# Kazakhstan cities list for dropdown
KAZAKHSTAN_CITIES = [
    "Almaty", "Astana", "Shymkent", "Aktobe", "Karaganda", "Taraz", 
    "Pavlodar", "Ust-Kamenogorsk", "Semey", "Atyrau", "Kostanay", 
    "Kyzylorda", "Uralsk", "Petropavlovsk", "Aktau", "Temirtau", 
    "Kokshetau", "Taldykorgan", "Ekibastuz", "Rudny", "Zhanaozen"
]

# Define new database schemas for schools
def create_school_collections():
    """Create necessary Firestore collections for school management"""
    try:
        schools_ref = db.collection('schools')
        return True
    except Exception as e:
        print(f"Error creating school collections: {e}")
        return False

# Routes for school management system
@app.route('/admin/schools', methods=['GET'])
@login_required
def admin_schools():
    # Ensure only admin can access
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    try:
        # Fetch all schools
        schools_ref = db.collection('schools')
        schools = schools_ref.stream()
        
        school_list = []
        for school_doc in schools:
            school_data = school_doc.to_dict()
            school_data['id'] = school_doc.id
            
            # Get student count for this school
            students_count = len(list(db.collection('users').where('school_id', '==', school_doc.id).get()))
            school_data['students_count'] = students_count
            
            school_list.append(school_data)
        
        return render_template('admin_schools.html', 
            schools=school_list,
            cities=KAZAKHSTAN_CITIES
        )
    except Exception as e:
        print(f"Error in admin_schools: {e}")
        flash('Error loading schools')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/generate_school_account', methods=['POST'])
@login_required
def generate_school_account():
    # Ensure only admin can access
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Generate random username and password
        username = 'school_' + ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
        password = ''.join(random.choices(string.ascii_letters + string.digits + string.punctuation, k=12))
        
        # Create account in Firebase Auth
        user = auth.create_user(
            email=f"{username}@placeholder.edu",
            password=password,
            display_name=f"School Admin {username}"
        )
        
        # Create school document in Firestore with password stored (encrypted)
        # Note: In production, use a more secure method for sensitive data
        school_data = {
            'username': username,
            'email': f"{username}@placeholder.edu",
            'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
            'uid': user.uid,
            'is_school': True,
            'onboarding_completed': False,
            'created_by_admin': session['user_id'],
            'password': password  # Store the password for admin access
        }
        
        db.collection('schools').document(user.uid).set(school_data)
        
        # Log action
        log_admin_action('create_school_account', {
            'school_username': username,
            'school_uid': user.uid
        })
        
        return jsonify({
            'success': True,
            'username': username,
            'password': password,
            'school_id': user.uid
        })
    
    except Exception as e:
        print(f"Error generating school account: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/school/onboarding', methods=['GET', 'POST'])
@login_required
def school_onboarding():
    user_id = session['user_id']
    
    # Check if this is a school account
    school_doc = db.collection('schools').document(user_id).get()
    if not school_doc.exists:
        flash('This account is not authorized for school management')
        return redirect(url_for('index'))
    
    school_data = school_doc.to_dict()
    
    # If already completed onboarding, redirect to dashboard
    if school_data.get('onboarding_completed', False):
        return redirect(url_for('school_dashboard'))
    
    if request.method == 'POST':
        try:
            school_name = request.form.get('school_name')
            city = request.form.get('city')
            
            if not school_name or not city:
                flash('School name and city are required')
                return redirect(url_for('school_onboarding'))
            
            # Handle logo upload
            logo_url = None
            if 'school_logo' in request.files:
                logo_file = request.files['school_logo']
                if logo_file and logo_file.filename:
                    # Create a unique filename
                    file_extension = logo_file.filename.rsplit('.', 1)[1].lower()
                    filename = f"schools/logos/{user_id}/logo.{file_extension}"
                    
                    # Upload to Firebase Storage
                    blob = bucket.blob(filename)
                    blob.upload_from_string(
                        logo_file.read(),
                        content_type=logo_file.content_type
                    )
                    
                    # Make public and get URL
                    blob.make_public()
                    logo_url = blob.public_url
            
            # Update school document
            db.collection('schools').document(user_id).update({
                'name': school_name,
                'city': city,
                'logo_url': logo_url,
                'onboarding_completed': True,
                'updated_at': datetime.datetime.now(tz=datetime.timezone.utc)
            })
            
            flash('School registration completed successfully')
            return redirect(url_for('school_dashboard'))
            
        except Exception as e:
            print(f"Error in school onboarding: {e}")
            flash(f'Error: {str(e)}')
            return redirect(url_for('school_onboarding'))
    
    return render_template('school_onboarding.html', 
                          cities=KAZAKHSTAN_CITIES, 
                          school_data=school_data)
                          
@app.route('/school/dashboard')
@login_required
def school_dashboard():
    user_id = session['user_id']
    
    # Check if this is a school account
    school_doc = db.collection('schools').document(user_id).get()
    if not school_doc.exists:
        flash('This account is not authorized for school management')
        return redirect(url_for('index'))
    
    school_data = school_doc.to_dict()
    
    # If onboarding not completed, redirect to onboarding
    if not school_data.get('onboarding_completed', False):
        return redirect(url_for('school_onboarding'))
    
    try:
        # Get all students for this school
        students_ref = db.collection('users').where('school_id', '==', user_id)
        students = list(students_ref.stream())
        
        total_students = len(students)
        
        # Initialize statistics variables
        class_distribution = {}
        average_gpa = 0
        total_gpa_count = 0
        total_certificates = 0
        profession_distribution = {}
        language_distribution = {}
        total_languages = 0
        achievement_counts = {}
        total_achievements = 0
        
        # For checking highest performing students
        highest_gpa_student = {'gpa': 0, 'name': 'None', 'class': 'None'}
        most_certs_student = {'count': 0, 'name': 'None', 'class': 'None'}
        most_langs_student = {'count': 0, 'name': 'None', 'class': 'None'}
        most_achievements_student = {'count': 0, 'name': 'None', 'class': 'None'}
        
        # Process student data
        for student_doc in students:
            student_data = student_doc.to_dict()
            student_name = student_data.get('original_full_name', student_data.get('full_name', 'Unknown'))
            
            # Normalize and count by class
            student_class = normalize_class_name(student_data.get('class', 'Unassigned'))
            
            # Update student class to normalized version if needed
            if student_data.get('class') != student_class:
                db.collection('users').document(student_doc.id).update({
                    'class': student_class
                })
                
            if student_class in class_distribution:
                class_distribution[student_class] += 1
            else:
                class_distribution[student_class] = 1
            
            # Calculate GPA average
            if student_data.get('academic_info', {}).get('gpa'):
                try:
                    gpa = float(student_data['academic_info']['gpa'])
                    average_gpa += gpa
                    total_gpa_count += 1
                    
                    # Check if highest GPA
                    if gpa > highest_gpa_student['gpa']:
                        highest_gpa_student = {
                            'gpa': gpa,
                            'name': student_name,
                            'class': student_class
                        }
                except:
                    pass
            
            # Count certificates
            certs_query = db.collection('users').document(student_doc.id).collection('certificates').stream()
            student_cert_count = len(list(certs_query))
            total_certificates += student_cert_count
            
            # Check if most certificates
            if student_cert_count > most_certs_student['count']:
                most_certs_student = {
                    'count': student_cert_count,
                    'name': student_name,
                    'class': student_class
                }
            
            # Count profession choices
            specialty = student_data.get('specialty', 'Not specified')
            if specialty in profession_distribution:
                profession_distribution[specialty] += 1
            else:
                profession_distribution[specialty] = 1
            
            # Process languages
            student_languages = student_data.get('academic_info', {}).get('languages', [])
            languages_count = len(student_languages)
            total_languages += languages_count
            
            # Check if most languages
            if languages_count > most_langs_student['count']:
                most_langs_student = {
                    'count': languages_count,
                    'name': student_name,
                    'class': student_class
                }
            
            # Count languages distribution
            for lang_data in student_languages:
                lang_name = lang_data.get('name', 'Unknown')
                if lang_name in language_distribution:
                    language_distribution[lang_name] += 1
                else:
                    language_distribution[lang_name] = 1
            
            # Process achievements
            student_achievements = student_data.get('academic_info', {}).get('achievements', [])
            achievements_count = len(student_achievements)
            total_achievements += achievements_count
            
            # Check if most achievements
            if achievements_count > most_achievements_student['count']:
                most_achievements_student = {
                    'count': achievements_count,
                    'name': student_name,
                    'class': student_class
                }
            
            # Count achievement types
            for achievement in student_achievements:
                achievement_title = achievement.get('title', 'Unknown')
                # Simplify by taking just the first word (category)
                achievement_category = achievement_title.split()[0] if achievement_title.split() else 'Unknown'
                if achievement_category in achievement_counts:
                    achievement_counts[achievement_category] += 1
                else:
                    achievement_counts[achievement_category] = 1
        
        # Calculate averages
        if total_gpa_count > 0:
            average_gpa = average_gpa / total_gpa_count
        average_certificates = total_certificates / total_students if total_students > 0 else 0
        average_languages = total_languages / total_students if total_students > 0 else 0
        average_achievements = total_achievements / total_students if total_students > 0 else 0
        
        # Get class-specific statistics
        class_stats = {}
        for class_name, count in class_distribution.items():
            class_students = db.collection('users').where('school_id', '==', user_id).where('class', '==', class_name).stream()
            
            class_gpa_total = 0
            class_gpa_count = 0
            class_cert_total = 0
            class_languages_total = 0
            class_achievements_total = 0
            class_specialties = {}
            
            for student in class_students:
                student_data = student.to_dict()
                
                # Calculate class GPA
                if student_data.get('academic_info', {}).get('gpa'):
                    try:
                        gpa = float(student_data['academic_info']['gpa'])
                        class_gpa_total += gpa
                        class_gpa_count += 1
                    except:
                        pass
                
                # Count class certificates
                certs = db.collection('users').document(student.id).collection('certificates').stream()
                student_certs = len(list(certs))
                class_cert_total += student_certs
                
                # Count languages
                student_languages = student_data.get('academic_info', {}).get('languages', [])
                class_languages_total += len(student_languages)
                
                # Count achievements
                student_achievements = student_data.get('academic_info', {}).get('achievements', [])
                class_achievements_total += len(student_achievements)
                
                # Track specialties
                specialty = student_data.get('specialty', 'Not specified')
                if specialty in class_specialties:
                    class_specialties[specialty] += 1
                else:
                    class_specialties[specialty] = 1
            
            class_avg_gpa = class_gpa_total / class_gpa_count if class_gpa_count > 0 else 0
            class_avg_certs = class_cert_total / count if count > 0 else 0
            class_avg_languages = class_languages_total / count if count > 0 else 0
            class_avg_achievements = class_achievements_total / count if count > 0 else 0
            
            # Get top specialty for this class
            top_specialty = max(class_specialties.items(), key=lambda x: x[1]) if class_specialties else ('None', 0)
            
            class_stats[class_name] = {
                'count': count,
                'avg_gpa': class_avg_gpa,
                'avg_certificates': class_avg_certs,
                'avg_languages': class_avg_languages,
                'avg_achievements': class_avg_achievements,
                'top_specialty': top_specialty[0]
            }

        # Prepare chart data for class distribution
        class_chart_data = []
        for class_name, count in sorted(class_distribution.items()):
            class_chart_data.append({
                'class': class_name,
                'count': count
            })
        
        # Create chart data for profession distribution (top 5)
        profession_chart_data = []
        for profession, count in sorted(profession_distribution.items(), key=lambda x: x[1], reverse=True)[:5]:
            profession_chart_data.append({
                'profession': profession if profession != 'Not specified' else 'Undecided',
                'count': count
            })
        
        # Create language distribution chart data
        language_chart_data = []
        for language, count in sorted(language_distribution.items(), key=lambda x: x[1], reverse=True)[:5]:
            language_chart_data.append({
                'language': language,
                'count': count
            })
        
        # Create achievement distribution chart data
        achievement_chart_data = []
        for achievement, count in sorted(achievement_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
            achievement_chart_data.append({
                'achievement': achievement,
                'count': count
            })
        
        # Calculate top performing class
        if class_stats:
            top_gpa_class = max(class_stats.items(), key=lambda x: x[1]['avg_gpa'])
            top_certs_class = max(class_stats.items(), key=lambda x: x[1]['avg_certificates'])
            top_langs_class = max(class_stats.items(), key=lambda x: x[1]['avg_languages'])
            top_achievements_class = max(class_stats.items(), key=lambda x: x[1]['avg_achievements'])
        else:
            top_gpa_class = ('None', {'avg_gpa': 0})
            top_certs_class = ('None', {'avg_certificates': 0})
            top_langs_class = ('None', {'avg_languages': 0})
            top_achievements_class = ('None', {'avg_achievements': 0})
        
        return render_template('school_dashboard.html',
                             school_data=school_data,
                             total_students=total_students,
                             class_distribution=class_distribution,
                             average_gpa=average_gpa,
                             average_certificates=average_certificates,
                             average_languages=average_languages,
                             average_achievements=average_achievements,
                             class_stats=class_stats,
                             class_chart_data=class_chart_data,
                             profession_chart_data=profession_chart_data,
                             language_chart_data=language_chart_data,
                             achievement_chart_data=achievement_chart_data,
                             highest_gpa_student=highest_gpa_student,
                             most_certs_student=most_certs_student,
                             most_langs_student=most_langs_student,
                             most_achievements_student=most_achievements_student,
                             top_gpa_class=top_gpa_class,
                             top_certs_class=top_certs_class,
                             top_langs_class=top_langs_class,
                             top_achievements_class=top_achievements_class)
                             
    except Exception as e:
        print(f"Error in school dashboard: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading dashboard: {str(e)}')
        return redirect(url_for('index'))


@app.route('/admin/delete_school/<school_id>', methods=['POST'])
@login_required
def delete_school(school_id):
    # Ensure only admin can access
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Get school document
        school_doc = db.collection('schools').document(school_id).get()
        
        if not school_doc.exists:
            return jsonify({'success': False, 'error': 'School not found'}), 404
            
        school_data = school_doc.to_dict()
        school_name = school_data.get('name', 'Unknown School')
        
        # Get all students associated with this school
        students_ref = db.collection('users').where('school_id', '==', school_id)
        students = list(students_ref.stream())
        
        student_count = len(students)
        deleted_count = 0
        
        # Delete all student accounts
        for student_doc in students:
            student_id = student_doc.id
            student_data = student_doc.to_dict()
            
            try:
                # Delete student from Firebase Auth
                auth.delete_user(student_id)
                
                # Delete student's certificates collection
                certificates_ref = db.collection('users').document(student_id).collection('certificates')
                delete_collection(certificates_ref, 100)
                
                # Delete student's comments collection
                comments_ref = db.collection('users').document(student_id).collection('comments')
                delete_collection(comments_ref, 100)
                
                # Delete student's notifications collection
                notifications_ref = db.collection('users').document(student_id).collection('notifications')
                delete_collection(notifications_ref, 100)
                
                # Delete student document
                db.collection('users').document(student_id).delete()
                
                deleted_count += 1
            except Exception as e:
                print(f"Error deleting student {student_id}: {e}")
                continue
        
        # Delete school's logo from Storage if it exists
        if school_data.get('logo_url'):
            try:
                # Extract filename from URL
                logo_path = f"schools/logos/{school_id}"
                blobs = bucket.list_blobs(prefix=logo_path)
                for blob in blobs:
                    blob.delete()
            except Exception as e:
                print(f"Error deleting school logo: {e}")
        
        # Delete the school from Firebase Auth
        try:
            auth.delete_user(school_id)
        except Exception as e:
            print(f"Error deleting school from Firebase Auth: {e}")
        
        # Delete school document
        db.collection('schools').document(school_id).delete()
        
        # Log the action
        log_admin_action('delete_school', {
            'school_id': school_id,
            'school_name': school_name,
            'students_deleted': deleted_count,
            'total_students': student_count
        })
        
        return jsonify({
            'success': True,
            'message': f"School '{school_name}' deleted successfully with {deleted_count} associated student accounts."
        })
        
    except Exception as e:
        print(f"Error deleting school: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Helper function to delete a collection
def delete_collection(collection_ref, batch_size):
    docs = collection_ref.limit(batch_size).stream()
    deleted = 0
    
    for doc in docs:
        doc.reference.delete()
        deleted += 1
        
    if deleted >= batch_size:
        return delete_collection(collection_ref, batch_size)
@app.route('/school/students')
@login_required
def school_students():
    user_id = session['user_id']
    
    # Check if this is a school account
    school_doc = db.collection('schools').document(user_id).get()
    if not school_doc.exists:
        flash('This account is not authorized for school management')
        return redirect(url_for('index'))
    
    school_data = school_doc.to_dict()
    
    # If onboarding not completed, redirect to onboarding
    if not school_data.get('onboarding_completed', False):
        return redirect(url_for('school_onboarding'))
    
    try:
        # Check if filtering by class
        class_filter = request.args.get('class')
        
        # Get students for this school
        students_ref = db.collection('users').where('school_id', '==', user_id)
        
        # Apply class filter if provided
        if class_filter:
            normalized_class = normalize_class_name(class_filter)
            students_ref = students_ref.where('class', '==', normalized_class)
            
        students = students_ref.stream()
        
        student_list = []
        for student_doc in students:
            student_data = student_doc.to_dict()
            student_data['id'] = student_doc.id
            
            # Always use original name for school dashboard
            if 'original_full_name' in student_data:
                student_data['school_display_name'] = student_data['original_full_name']
            else:
                student_data['school_display_name'] = student_data.get('full_name', '')
            
            # Get certificate count
            certs_query = db.collection('users').document(student_doc.id).collection('certificates').stream()
            student_data['certificate_count'] = len(list(certs_query))
            
            student_list.append(student_data)
        
        # Sort students by class and last name
        student_list.sort(key=lambda x: (x.get('class', ''), x.get('school_display_name', '')))
        
        return render_template('school_students.html',
                             school_data=school_data,
                             students=student_list,
                             class_filter=class_filter)
                             
    except Exception as e:
        print(f"Error in school students: {e}")
        flash(f'Error loading students: {str(e)}')
        return redirect(url_for('school_dashboard'))  
@app.route('/school/normalize-classes', methods=['POST'])
@login_required
def normalize_school_classes():
    user_id = session['user_id']
    
    # Check if this is a school account
    school_doc = db.collection('schools').document(user_id).get()
    if not school_doc.exists:
        flash('This account is not authorized for school management')
        return redirect(url_for('index'))
    
    try:
        # Get all students for this school
        students_ref = db.collection('users').where('school_id', '==', user_id)
        students = students_ref.stream()
        
        updated_count = 0
        class_mapping = {}  # Keep track of class name changes
        
        for student_doc in students:
            student_data = student_doc.to_dict()
            original_class = student_data.get('class', 'Unassigned')
            normalized_class = normalize_class_name(original_class)
            
            # Only update if there's a difference
            if original_class != normalized_class:
                db.collection('users').document(student_doc.id).update({
                    'class': normalized_class
                })
                updated_count += 1
                
                # Track changes for reporting
                if (original_class, normalized_class) in class_mapping:
                    class_mapping[(original_class, normalized_class)] += 1
                else:
                    class_mapping[(original_class, normalized_class)] = 1
        
        # Prepare result message
        if updated_count > 0:
            message = f"Updated {updated_count} student records. Changes made:<br>"
            for (old_class, new_class), count in class_mapping.items():
                message += f"• {old_class} → {new_class}: {count} students<br>"
            flash(message)
        else:
            flash("All class names are already normalized.")
        
        return redirect(url_for('school_dashboard'))
    
    except Exception as e:
        print(f"Error normalizing class names: {e}")
        flash(f'Error: {str(e)}')
        return redirect(url_for('school_dashboard'))    
@app.route('/school/import-students', methods=['GET', 'POST'])
@login_required
def import_students():
    user_id = session['user_id']
    
    # Check if this is a school account
    school_doc = db.collection('schools').document(user_id).get()
    if not school_doc.exists:
        flash('This account is not authorized for school management')
        return redirect(url_for('index'))
    
    school_data = school_doc.to_dict()
    
    # If onboarding not completed, redirect to onboarding
    if not school_data.get('onboarding_completed', False):
        return redirect(url_for('school_onboarding'))
    
    if request.method == 'POST':
        try:
            # Check if file was uploaded
            if 'students_file' not in request.files:
                flash('No file uploaded')
                return redirect(url_for('import_students'))
            
            file = request.files['students_file']
            if not file or file.filename == '':
                flash('No file selected')
                return redirect(url_for('import_students'))
            
            # Verify it's an Excel file
            if not file.filename.endswith(('.xlsx', '.xls')):
                flash('File must be an Excel spreadsheet (.xlsx or .xls)')
                return redirect(url_for('import_students'))
            
            # Read Excel file
            df = pd.read_excel(file)
            
            # Verify required columns
            required_columns = ['Full Name', 'Class']
            for col in required_columns:
                if col not in df.columns:
                    flash(f'Missing required column: {col}')
                    return redirect(url_for('import_students'))
            
            # Get optional columns if they exist
            has_languages = 'Languages' in df.columns
            has_achievements = 'Achievements' in df.columns
            
            # Process students
            school_name = school_data.get('name', 'Unknown School')
            created_students = []
            class_groups = {}
            
            for _, row in df.iterrows():
                full_name = row['Full Name'].strip()
                # Normalize class name here
                student_class = normalize_class_name(str(row['Class']).strip())
                
                if not full_name:
                    continue
                
                # Process languages if available
                languages = []
                if has_languages and not pd.isna(row['Languages']):
                    # Format: "English:B2,Russian:Native,Kazakh:C1"
                    lang_str = str(row['Languages']).strip()
                    if lang_str:
                        for lang_entry in lang_str.split(','):
                            if ':' in lang_entry:
                                lang, level = lang_entry.split(':', 1)
                                languages.append({
                                    'name': lang.strip(),
                                    'level': level.strip()
                                })
                
                # Process achievements if available
                achievements = []
                if has_achievements and not pd.isna(row['Achievements']):
                    # Format: "Math Olympiad 2022|First Place|2022-05-15,Science Fair|Honorable Mention|2021-11-10"
                    achieve_str = str(row['Achievements']).strip()
                    if achieve_str:
                        for achieve_entry in achieve_str.split(','):
                            parts = achieve_entry.split('|')
                            if len(parts) >= 1:
                                achievement = {
                                    'title': parts[0].strip()
                                }
                                if len(parts) >= 2:
                                    achievement['description'] = parts[1].strip()
                                if len(parts) >= 3:
                                    achievement['date'] = parts[2].strip()
                                achievements.append(achievement)
                
                # Generate username from full name - handling Cyrillic properly
                transliterated_name = transliterate_name(full_name)
                username_base = transliterated_name
                display_username = full_name  # Keep original name for display
                
                # Generate random suffix to ensure uniqueness
                random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=7))
                username = f"{username_base}{random_suffix}"
                
                # Generate random password
                password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                
                try:
                    # Create user in Firebase Auth
                    user = auth.create_user(
                        email=f"{username}@student.{school_name.lower().replace(' ', '')}.edu",
                        password=password,
                        display_name=full_name
                    )
                    
                    # Create user document with enhanced data
                    user_data = {
                        'username': username.lower(),
                        'display_username': display_username,
                        'email': f"{username}@student.{school_name.lower().replace(' ', '')}.edu",
                        'full_name': full_name,
                        'original_full_name': full_name,  # Store original name for school tracking
                        'school_id': user_id,
                        'school_name': school_name,
                        'school_logo_url': school_data.get('logo_url'),
                        'class': student_class,  # Normalized class name
                        'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
                        'uid': user.uid,
                        'education': school_name,
                        'location': {
                            'city': school_data.get('city', 'Unknown'),
                            'country': 'Kazakhstan'
                        },
                        'academic_info': {
                            'gpa': '',
                            'sat_score': '',
                            'toefl_score': '',
                            'ielts_score': '',
                            'languages': languages,
                            'achievements': achievements
                        },
                        'student_password': password  # Store password for school admin access
                    }
                    
                    db.collection('users').document(user.uid).set(user_data)
                    
                    # Add student to result
                    student_info = {
                        'full_name': full_name,
                        'class': student_class,
                        'username': username,
                        'password': password,
                        'id': user.uid
                    }
                    created_students.append(student_info)
                    
                    # Group by class
                    if student_class not in class_groups:
                        class_groups[student_class] = []
                    class_groups[student_class].append(student_info)
                    
                except Exception as e:
                    print(f"Error creating student {full_name}: {e}")
                    continue
            
            # Generate master Excel file
            master_excel = create_student_excel(created_students, school_name)
            
            # Generate class-specific Excel files
            class_excel_files = {}
            for class_name, students in class_groups.items():
                class_excel_files[class_name] = create_student_excel(students, school_name, class_name)
            
            # Return the results page with download links
            return render_template('import_results.html', 
                                 school_data=school_data,
                                 created_students=created_students,
                                 class_groups=class_groups)
            
        except Exception as e:
            print(f"Error importing students: {e}")
            flash(f'Error: {str(e)}')
            return redirect(url_for('import_students'))
    
    return render_template('import_students.html', school_data=school_data)

@app.route('/school/download-excel/master', methods=['GET'])
@login_required
def download_excel_master():
    user_id = session['user_id']
    
    # Check if this is a school account
    school_doc = db.collection('schools').document(user_id).get()
    if not school_doc.exists:
        flash('This account is not authorized for school management')
        return redirect(url_for('index'))
    
    school_data = school_doc.to_dict()
    school_name = school_data.get('name', 'School')
    
    try:
        # Fetch all students
        students_ref = db.collection('users').where('school_id', '==', user_id)
        students = students_ref.stream()
        
        student_list = []
        for student_doc in students:
            student_data = student_doc.to_dict()
            student_list.append({
                'full_name': student_data.get('original_full_name', student_data.get('full_name', '')),
                'class': student_data.get('class', ''),
                'username': student_data.get('username', ''),
                'password': student_data.get('student_password', '********')  # Use stored password if available
            })
        
        # Create Excel
        excel_file = create_student_excel(student_list, school_name)
        
        # Return the Excel file
        return send_file(
            BytesIO(excel_file.getvalue()),
            as_attachment=True,
            download_name=f"{school_name}_students.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error downloading Excel: {e}")
        flash(f'Error: {str(e)}')
        return redirect(url_for('school_students'))
    

@app.route('/school/download-excel/class/<class_name>', methods=['GET'])
@login_required
def download_excel_class(class_name):
    user_id = session['user_id']
    
    # Check if this is a school account
    school_doc = db.collection('schools').document(user_id).get()
    if not school_doc.exists:
        flash('This account is not authorized for school management')
        return redirect(url_for('index'))
    
    school_data = school_doc.to_dict()
    school_name = school_data.get('name', 'School')
    
    try:
        # Normalize the class name parameter
        normalized_class_name = normalize_class_name(class_name)
        
        # Fetch students for this class
        students_ref = db.collection('users').where('school_id', '==', user_id).where('class', '==', normalized_class_name)
        students = students_ref.stream()
        
        student_list = []
        for student_doc in students:
            student_data = student_doc.to_dict()
            student_list.append({
                'full_name': student_data.get('original_full_name', student_data.get('full_name', '')),
                'class': student_data.get('class', ''),
                'username': student_data.get('username', ''),
                'password': student_data.get('student_password', '********')  # Use stored password if available
            })
        
        # Create Excel
        excel_file = create_student_excel(student_list, school_name, normalized_class_name)
        
        # Return the Excel file
        return send_file(
            BytesIO(excel_file.getvalue()),
            as_attachment=True,
            download_name=f"{school_name}_class_{normalized_class_name}_students.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error downloading Excel: {e}")
        flash(f'Error: {str(e)}')
        return redirect(url_for('school_students'))   

@app.route('/school/download-import-template', methods=['GET'])
def download_import_template():
    """Provide an enhanced template Excel file for student imports with languages and achievements"""
    try:
        # Create an Excel file in memory
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # Get the active worksheet
        worksheet = workbook.active
        worksheet.title = "Students Import Template"
        
        # Add headers
        headers = ["Full Name", "Class", "Languages", "Achievements"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ["Иван Иванов", "10A", "English:B2,Russian:Native,Kazakh:C1", "Math Olympiad|First Place|2022-05-15"],
            ["Анна Смирнова", "11B", "English:C1,Russian:Native", "Science Fair|Honorable Mention|2021-11-10"],
            ["Айнур Ахметова", "9C", "Kazakh:Native,Russian:C1,English:B1", "Spelling Bee|Second Place|2023-03-25"]
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        
        # Add instructions in a separate sheet
        instructions_sheet = workbook.create_sheet(title="Instructions")
        
        instructions = [
            ["Student Import Instructions"],
            [""],
            ["1. Fill out the 'Students Import Template' sheet with your students' information."],
            ["2. Required Columns:"],
            ["   - Full Name: Student's full name (can be in Cyrillic)"],
            ["   - Class: Student's class (e.g., '10A', '11B')"],
            ["3. Optional Columns:"],
            ["   - Languages: Format as 'Language:Level,Language:Level' (e.g., 'English:B2,Russian:Native')"],
            ["   - Achievements: Format as 'Title|Description|Date' (e.g., 'Math Olympiad|First Place|2022-05-15')"],
            ["4. Usernames will be automatically generated using transliterated names + random string for uniqueness."],
            ["5. Do not change the column headers."],
            ["6. Save the file and upload it to the student import page."]
        ]
        
        for row_num, row_data in enumerate(instructions, 1):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = instructions_sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)
        
        # Adjust column widths
        for worksheet in workbook.worksheets:
            for col in worksheet.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
        
        # Save the workbook to the BytesIO object
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="student_import_template.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"Error creating template: {e}")
        flash('Error creating template file')
        return redirect(url_for('import_students'))
def create_student_excel(students, school_name, class_name=None):
    """Create an Excel file with student account information - now with real passwords"""
    output = io.BytesIO()
    
    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # Set worksheet title
    if class_name:
        worksheet.title = f"Class {class_name} Students"
    else:
        worksheet.title = "All Students"
    
    # Add title
    if class_name:
        worksheet.merge_cells('A1:D1')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"{school_name} - Class {class_name} Student Accounts"
    else:
        worksheet.merge_cells('A1:D1')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"{school_name} - Student Accounts"
    
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ["Full Name", "Class", "Username", "Password"]
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data - now including real passwords
    for row_num, student in enumerate(students, header_row + 1):
        worksheet.cell(row=row_num, column=1).value = student.get('full_name', '')
        worksheet.cell(row=row_num, column=2).value = student.get('class', '')
        worksheet.cell(row=row_num, column=3).value = student.get('username', '')
        
        # Include the real password instead of asterisks
        password_value = student.get('password', '')
        worksheet.cell(row=row_num, column=4).value = password_value
    
    # Add instructions
    instruction_row = len(students) + header_row + 2
    worksheet.merge_cells(f'A{instruction_row}:D{instruction_row}')
    instruction_cell = worksheet.cell(row=instruction_row, column=1)
    instruction_cell.value = "Instructions: Share the username and password with each student. Students should log in at the website and change their password."
    instruction_cell.font = Font(italic=True)
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
    
    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)
    
    return output



# Helper function to check if user is a school admin
def is_school_admin():
    """Check if current user is a school admin"""
    if 'user_id' not in session:
        return False
    
    try:
        school_doc = db.collection('schools').document(session['user_id']).get()
        return school_doc.exists
    except:
        return False

# Inject school admin status into all templates
@app.context_processor
def inject_school_admin_status():
    return {
        'is_school_admin': is_school_admin()
    }
@app.route('/admin/view_school_password/<school_id>', methods=['GET'])
@login_required
def view_school_password(school_id):
    # Ensure only admin can access
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Get school document
        school_doc = db.collection('schools').document(school_id).get()
        
        if not school_doc.exists:
            return jsonify({'success': False, 'error': 'School not found'}), 404
            
        school_data = school_doc.to_dict()
        
        # Return password if available
        if 'password' in school_data:
            return jsonify({
                'success': True,
                'username': school_data.get('username', ''),
                'email': school_data.get('email', ''),
                'password': school_data.get('password', '')
            })
        else:
            return jsonify({'success': False, 'error': 'Password not available'}), 404
            
    except Exception as e:
        print(f"Error viewing school password: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


def transliterate_name(name):
    try:
        # Check if the name contains Cyrillic characters
        if any(ord(char) > 127 for char in name):
            # Transliterate to Latin alphabet
            latin_name = transliterate.translit(name, 'ru', reversed=True)
            # Remove spaces and special characters
            latin_name = ''.join(c for c in latin_name if c.isalnum())
            return latin_name.lower()
        else:
            # Name is already in Latin alphabet
            return ''.join(c for c in name if c.isalnum()).lower()
    except Exception as e:
        print(f"Transliteration error: {e}")
        # Return simplified version of name if transliteration fails
        return ''.join(c for c in name if c.isalnum() and ord(c) < 128).lower()

# Add this helper function to normalize class names
def normalize_class_name(class_name):
    """Normalize class name to ensure consistent formatting"""
    if not class_name:
        return ""
    
    # Convert string representation if needed
    class_name = str(class_name).strip()
    
    # Split into numeric and letter parts (e.g., "10B" -> "10" and "B")
    import re
    match = re.match(r'(\d+)([a-zA-Zа-яА-Я]+)', class_name)
    
    if match:
        number, letter = match.groups()
        # Keep the number as is and capitalize the letter part
        return f"{number}{letter.upper()}"
    else:
        # If it doesn't match the pattern, just uppercase it
        return class_name.upper()
    


# Add these routes to your app.py

@app.route('/career-test', methods=['GET'])
@login_required
def career_test():
    """Landing page for the career test system"""
    user_id = session['user_id']
    
    try:
        # Check if user has any saved progress
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        if test_progress.exists:
            progress_data = test_progress.to_dict()
            current_stage = progress_data.get('current_stage', 1)
            completed_stages = progress_data.get('completed_stages', [])
        else:
            # No existing progress, start at stage 1
            current_stage = 1
            completed_stages = []
        
        # Get user data for personalization
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict() if user_doc.exists else {}
        
        return render_template(
            'career_test/landing.html',
            current_stage=current_stage,
            completed_stages=completed_stages,
            user_data=user_data,
            current_user_avatar=get_current_user_avatar(),
            current_username=get_current_username(),
            total_stages=2  # Updated to 2 stages
        )
        
    except Exception as e:
        print(f"Error loading career test: {e}")
        flash('Error loading career test system')
        return redirect(url_for('profile'))


@app.route('/career-test/analyzing')
@login_required
def career_test_analyzing():
    """Show the analyzing animation before displaying results"""
    user_id = session['user_id']
    
    try:
        # Check if the user has completed all test stages
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        if not test_progress.exists:
            return redirect(url_for('career_test'))
            
        progress_data = test_progress.to_dict()
        completed_stages = progress_data.get('completed_stages', [])
        
        # FIXED: Check explicitly for stages 1 AND 2, not just the length
        if not (1 in completed_stages and 2 in completed_stages):
            print(f"User attempted to access analysis without completing both stages. Completed: {completed_stages}")
            # Find which stage is missing and redirect there
            missing_stages = [stage for stage in [1, 2] if stage not in completed_stages]
            next_stage = missing_stages[0] if missing_stages else 1
            print(f"Redirecting to stage {next_stage}")
            return redirect(url_for('career_test_stage', stage=next_stage))
        
        print(f"User has completed both stages. Showing analysis. Completed: {completed_stages}")
        # Render the analyzing template
        return render_template('career_test/analyzing.html')
        
    except Exception as e:
        print(f"Error in career test analyzing: {e}")
        flash('Error processing test data')
        return redirect(url_for('career_test'))

@app.route('/career-test/stage/<int:stage>', methods=['GET', 'POST'])
@login_required
def career_test_stage(stage):
    """Handle a specific stage of the career test with improved option handling"""
    user_id = session['user_id']
    
    # Validate stage
    if stage < 1 or stage > 2:
        flash('Invalid test stage')
        return redirect(url_for('career_test'))
    
    # Get previous progress
    try:
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        if test_progress.exists:
            progress_data = test_progress.to_dict()
            completed_stages = progress_data.get('completed_stages', [])
            saved_answers = progress_data.get('answers', {})
            
            # Get current question from database
            stage_key = f'stage_{stage}'
            current_question = progress_data.get('current_question', {}).get(stage_key, 1)
            
            print(f"Loaded current_question from DB: {current_question}")
        else:
            completed_stages = []
            saved_answers = {}
            current_question = 1
            
            # Create initial progress record
            initial_progress = {
                'current_stage': stage,
                'current_question': {f'stage_{stage}': 1},
                'completed_stages': [],
                'answers': {}
            }
            test_progress_ref.set(initial_progress)
            
        # Get test questions
        with open('static/js/career_test_questions.json', 'r', encoding='utf-8') as f:
            all_questions = json.load(f)
            
        stage_key = f'stage_{stage}'
        if stage_key not in all_questions:
            flash('Test questions not found')
            return redirect(url_for('career_test'))
            
        stage_questions = all_questions[stage_key]
        total_questions = len(stage_questions['questions'])
        
        # Ensure current_question is valid
        if current_question < 1:
            current_question = 1
        if current_question > total_questions:
            current_question = total_questions
        
        # Handle form submission (POST)
        if request.method == 'POST':
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            action = request.form.get('action')
            
            # Get question number from form
            form_question_id = int(request.form.get('question_id', 1))
            print(f"Form question_id: {form_question_id}, Action: {action}")
            
            if action == 'next':
                # Get current question data - with index validation
                question_index = form_question_id - 1
                if 0 <= question_index < len(stage_questions['questions']):
                    question_data = stage_questions['questions'][question_index]
                    
                    # DEBUG: Log the full question data structure
                    print(f"Current question data: {json.dumps(question_data, ensure_ascii=False)}")
                    
                    # Get available options for this question
                    options = question_data.get('options', [])
                    
                    # DEBUG: Log option data
                    print(f"Available options: {json.dumps(options, ensure_ascii=False)}")
                else:
                    print(f"WARNING: Question index out of range: {question_index}")
                    question_data = {}
                    options = []
                
                # Get the actual selected answers with their values
                raw_answers = request.form.getlist('answer')
                
                # Create a set to eliminate duplicates
                unique_answers = set()
                for answer in raw_answers:
                    if answer.strip():  # Only add non-empty answers
                        unique_answers.add(answer)
                
                # Convert back to list for storage
                selected_answers = list(unique_answers)
                
                # Debug log to check the raw form data
                print(f"Form data for question {form_question_id}: {dict(request.form)}")
                print(f"Raw answer values: {raw_answers}")
                print(f"Unique selected answer values: {selected_answers}")
                
                # Only save non-empty answers
                if selected_answers and any(selected_answers):
                    if stage_key not in saved_answers:
                        saved_answers[stage_key] = {}
                    
                    # Save the selected values (IDs or values from form)
                    saved_answers[stage_key][str(form_question_id)] = selected_answers
                    print(f"Saved answers for {stage_key} question {form_question_id}: {selected_answers}")
                    
                    # Store the answer text (this is what we'll send to Gemini)
                    answer_texts = []
                    option_texts = {}
                    
                    # Create mappings from different option identifiers to text
                    for opt in options:
                        opt_id = str(opt.get('id', ''))
                        opt_value = str(opt.get('value', opt_id))
                        opt_text = opt.get('text', '')
                        
                        # Map both ID and value to text
                        option_texts[opt_id] = opt_text
                        option_texts[opt_value] = opt_text
                        
                        # Also map the text itself as a key for direct text matching
                        option_texts[opt_text] = opt_text
                    
                    # DEBUG: Log the mapping table
                    print(f"Option ID/value to text mapping: {option_texts}")
                    
                    # Try to map each selected answer to its text
                    for answer in selected_answers:
                        answer_str = str(answer)
                        if answer_str in option_texts:
                            answer_texts.append(option_texts[answer_str])
                        else:
                            # Special case for RIASEC numeric ratings
                            if question_data.get('type') == 'riasec' and answer_str.isdigit():
                                rating = int(answer_str)
                                if 1 <= rating <= 5:
                                    answer_texts.append(f"Оценка {rating} из 5")
                                else:
                                    answer_texts.append(f"Рейтинг: {answer_str}")
                            else:
                                # If we can't find a mapping, use the raw answer as text
                                answer_texts.append(f"{answer_str}")
                                print(f"WARNING: Could not find text for option '{answer_str}' in question {form_question_id}")
                    
                    # Save the text values
                    if 'answer_texts' not in saved_answers:
                        saved_answers['answer_texts'] = {}
                    if stage_key not in saved_answers['answer_texts']:
                        saved_answers['answer_texts'][stage_key] = {}
                        
                    saved_answers['answer_texts'][stage_key][str(form_question_id)] = answer_texts
                    print(f"Saved answer texts: {answer_texts}")
                    
                else:
                    print(f"WARNING: No answers selected for question {form_question_id}")
                
                # Calculate next question
                next_question = form_question_id + 1
                print(f"Calculated next_question: {next_question}")
                
                if next_question <= total_questions:
                    # Update progress in DB
                    update_data = {
                        'current_stage': stage,
                        'current_question': {stage_key: next_question},
                        'answers': saved_answers,
                    }
                    test_progress_ref.update(update_data)
                    print(f"Updated DB with next_question: {next_question}")
                    print(f"Current saved answers: {saved_answers}")
                    
                    if is_ajax:
                        # Render next question - IMPROVED VERSION
                        question_data = stage_questions['questions'][next_question-1]
                        selected_answers = []
                        if stage_key in saved_answers and str(next_question) in saved_answers[stage_key]:
                            selected_answers = saved_answers[stage_key][str(next_question)]
                            
                        html = render_template(
                            'career_test/question_partial.html',
                            stage=stage,
                            stage_name=stage_questions['name'],
                            question=question_data,
                            question_number=next_question,
                            total_questions=total_questions,
                            selected_answers=selected_answers,
                            total_stages=2
                        )
                        
                        # IMPROVED JS to update ALL elements with question number
                        fix_all_counters_js = f"""
                            // JavaScript to synchronize all question counters on the page
                            (function() {{
                                // Find and update ALL elements with question number
                                const counters = document.querySelectorAll('.text-gray-600');
                                
                                counters.forEach(function(counter) {{
                                    // Check if this element contains "Question" text
                                    if (counter.textContent.includes('Вопрос')) {{
                                        counter.textContent = 'Вопрос {next_question} из {total_questions}';
                                        console.log('Updated counter: ' + counter.textContent);
                                    }}
                                }});
                                
                                // Force update the header, if present
                                const header = document.querySelector('h1.text-2xl');
                                if (header && header.textContent.includes('Вопрос')) {{
                                    header.textContent = 'Вопрос {next_question} из {total_questions}';
                                    console.log('Updated header: ' + header.textContent);
                                }}
                            }})();
                        """
                        
                        return jsonify({
                            'html': html,
                            'question_number': next_question,
                            'fixed_counter_js': fix_all_counters_js,
                            'debug_info': {
                                'form_question_id': form_question_id,
                                'calculated_next': next_question,
                                'total_questions': total_questions,
                                'saved_answers': saved_answers.get(stage_key, {})
                            }
                        })
                    else:
                        # Traditional redirect
                        return redirect(url_for('career_test_stage', stage=stage))
                else:
                    # STAGE COMPLETION LOGIC - IMPORTANT FIXES HERE
                    print(f"User finished all questions in stage {stage}.")
                    print(f"Current completed_stages before update: {completed_stages}")
                    
                    # Explicitly mark only the current stage as completed
                    if stage not in completed_stages:
                        completed_stages.append(stage)
                        print(f"Added stage {stage} to completed_stages. Now: {completed_stages}")
                    
                    # CRITICAL FIX: Logic for determining what happens after stage completion
                    if stage == 1:
                        # We just completed stage 1, must go to stage 2
                        next_stage = 2
                        
                        # Update database with stage completion
                        update_data = {
                            'current_stage': next_stage,
                            'current_question': {f'stage_{next_stage}': 1},  # Start at first question of stage 2
                            'completed_stages': completed_stages,
                            'answers': saved_answers,
                        }
                        
                        print(f"User completed stage 1. Setting current_stage to {next_stage} in DB.")
                        test_progress_ref.update(update_data)
                        
                        # Force redirect to stage 2
                        print("Redirecting user to stage 2")
                        if is_ajax:
                            return jsonify({'redirect': url_for('career_test_stage', stage=2)})
                        return redirect(url_for('career_test_stage', stage=2))
                    
                    elif stage == 2:
                        # We just completed stage 2
                        # Check if both stages are now completed
                        if 1 in completed_stages and 2 in completed_stages:
                            # Both stages are complete, we can proceed to analysis
                            update_data = {
                                'current_stage': 3,  # Mark as beyond normal stages
                                'completed_stages': completed_stages,
                                'answers': saved_answers,
                            }
                            
                            print("User completed both stages 1 and 2. Marking as ready for analysis.")
                            test_progress_ref.update(update_data)
                            
                            print("Redirecting to analysis page")
                            if is_ajax:
                                return jsonify({'redirect': url_for('career_test_analyzing')})
                            return redirect(url_for('career_test_analyzing'))
                        else:
                            # Stage 2 is complete, but stage 1 is not (unusual)
                            # Redirect to stage 1
                            update_data = {
                                'current_stage': 1,
                                'current_question': {'stage_1': 1},
                                'completed_stages': completed_stages,
                                'answers': saved_answers,
                            }
                            
                            print("User completed stage 2 but not stage 1. Redirecting to stage 1.")
                            test_progress_ref.update(update_data)
                            
                            if is_ajax:
                                return jsonify({'redirect': url_for('career_test_stage', stage=1)})
                            return redirect(url_for('career_test_stage', stage=1))
            
            elif action == 'prev':
                # Handle previous question logic
                prev_question = form_question_id - 1
                
                if prev_question >= 1:
                    # Update DB with previous question
                    update_data = {
                        'current_stage': stage,
                        'current_question': {stage_key: prev_question},
                    }
                    test_progress_ref.update(update_data)
                    
                    if is_ajax:
                        # Render previous question
                        question_data = stage_questions['questions'][prev_question-1]
                        selected_answers = []
                        if stage_key in saved_answers and str(prev_question) in saved_answers[stage_key]:
                            selected_answers = saved_answers[stage_key][str(prev_question)]
                            
                        html = render_template(
                            'career_test/question_partial.html',
                            stage=stage,
                            stage_name=stage_questions['name'],
                            question=question_data,
                            question_number=prev_question,
                            total_questions=total_questions,
                            selected_answers=selected_answers,
                            total_stages=2
                        )
                        
                        # Similar JS update for previous question
                        fix_all_counters_js = f"""
                            // JavaScript to synchronize all question counters on the page
                            (function() {{
                                // Find and update ALL elements with question number
                                const counters = document.querySelectorAll('.text-gray-600');
                                
                                counters.forEach(function(counter) {{
                                    // Check if this element contains "Question" text
                                    if (counter.textContent.includes('Вопрос')) {{
                                        counter.textContent = 'Вопрос {prev_question} из {total_questions}';
                                        console.log('Updated counter: ' + counter.textContent);
                                    }}
                                }});
                                
                                // Force update the header, if present
                                const header = document.querySelector('h1.text-2xl');
                                if (header && header.textContent.includes('Вопрос')) {{
                                    header.textContent = 'Вопрос {prev_question} из {total_questions}';
                                    console.log('Updated header: ' + header.textContent);
                                }}
                            }})();
                        """
                        
                        return jsonify({
                            'html': html,
                            'question_number': prev_question,
                            'fixed_counter_js': fix_all_counters_js
                        })
                    else:
                        return redirect(url_for('career_test_stage', stage=stage))
                else:
                    # Go to previous stage if available
                    if stage > 1:
                        prev_stage = stage - 1
                        # Get the last question of the previous stage
                        prev_stage_key = f'stage_{prev_stage}'
                        
                        with open('static/js/career_test_questions.json', 'r', encoding='utf-8') as f:
                            all_questions = json.load(f)
                            
                        if prev_stage_key in all_questions:
                            prev_stage_questions = all_questions[prev_stage_key]
                            last_question = len(prev_stage_questions['questions'])
                            
                            # Update DB to point to the last question of previous stage
                            update_data = {
                                'current_stage': prev_stage,
                                'current_question': {prev_stage_key: last_question},
                            }
                            test_progress_ref.update(update_data)
                        
                        if is_ajax:
                            return jsonify({'redirect': url_for('career_test_stage', stage=prev_stage)})
                        return redirect(url_for('career_test_stage', stage=prev_stage))
                    else:
                        # Can't go back further, redirect to landing page
                        if is_ajax:
                            return jsonify({'redirect': url_for('career_test')})
                        return redirect(url_for('career_test'))
        
        # GET request - display current question
        # Ensure current_question is within range of 1 to total_questions
        current_question = max(1, min(current_question, total_questions))
        question_data = stage_questions['questions'][current_question-1]
        
        # Get previously saved answer for this question if any
        selected_answers = []
        if stage_key in saved_answers and str(current_question) in saved_answers[stage_key]:
            selected_answers = saved_answers[stage_key][str(current_question)]
        
        return render_template(
            'career_test/question.html',
            stage=stage,
            stage_name=stage_questions['name'],
            question=question_data,
            question_number=current_question,
            total_questions=total_questions,
            selected_answers=selected_answers,
            completed_stages=completed_stages,
            current_user_avatar=get_current_user_avatar(),
            current_username=get_current_username(),
            total_stages=2
        )
            
    except Exception as e:
        print(f"Error in career test stage {stage}: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading test stage: {str(e)}')
        return redirect(url_for('career_test'))
    
    
@app.route('/career-test/results')
@login_required
def career_test_results():
    """Display career test results with analysis animation and cached recommendations"""
    user_id = session['user_id']
    
    try:
        # Get test progress
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        if not test_progress.exists:
            flash('Результаты теста не найдены')
            return redirect(url_for('career_test'))
            
        progress_data = test_progress.to_dict()
        
        # Check if all stages are completed - updated for 2 stages
        completed_stages = progress_data.get('completed_stages', [])
        if len(completed_stages) < 2:
            flash('Пожалуйста, завершите все этапы теста')
            next_stage = min([i for i in range(1, 3) if i not in completed_stages])
            return redirect(url_for('career_test_stage', stage=next_stage))
        
        # Check if recommendations already exist - KEY FIX
        from_cache = False
        recommendations = progress_data.get('recommendations', [])
        
        # If no recommendations exist or they're empty, redirect to analyzing page
        if not recommendations:
            return redirect(url_for('career_test_analyzing'))
        else:
            # Using existing recommendations from database
            from_cache = True
            print(f"Using cached recommendations for user {user_id}")
        
        # Render the results template with recommendations
        return render_template(
            'career_test/results.html',
            recommendations=recommendations,
            user_data=progress_data,
            current_user_avatar=get_current_user_avatar(),
            current_username=get_current_username(),
            from_cache=from_cache,
            total_stages=2
        )
        
    except Exception as e:
        print(f"Error generating career test results: {e}")
        import traceback
        traceback.print_exc()
        flash('Ошибка при генерации рекомендаций по карьере')
        return redirect(url_for('career_test'))
    
@app.route('/career-test/profession/<profession_slug>')
@login_required
def career_profession_detail(profession_slug):
    """Display detailed information about a specific profession"""
    user_id = session['user_id']
    
    try:
        # Check if cached profession data exists
        profession_ref = db.collection('profession_data').document(profession_slug)
        profession_doc = profession_ref.get()
        
        if profession_doc.exists:
            # Use cached data
            profession_data = profession_doc.to_dict()
        else:
            # Generate profession details using AI
            profession_name = profession_slug.replace('-', ' ').title()
            profession_data = generate_profession_details(profession_name)
            
            # Cache the results
            profession_ref.set(profession_data)
        
        # Get user data for personalization (compatibility score)
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        compatibility = None
        if test_progress.exists:
            progress_data = test_progress.to_dict()
            recommendations = progress_data.get('recommendations', [])
            
            # Find this profession in recommendations to get compatibility
            for rec in recommendations:
                if rec.get('slug') == profession_slug:
                    compatibility = rec.get('compatibility')
                    break
        
        return render_template(
            'career_test/profession_detail.html',
            profession=profession_data,
            compatibility=compatibility,
            current_user_avatar=get_current_user_avatar(),
            current_username=get_current_username()
        )
        
    except Exception as e:
        print(f"Error loading profession details: {e}")
        flash('Error loading profession details')
        return redirect(url_for('career_test_results'))


def analyze_career_test_results(answers):
    """Analyze test answers using the stored text values of options instead of IDs"""
    try:
        user_id = session.get('user_id')
        
        # Проверка, что answers не пустой
        if not answers or not isinstance(answers, dict):
            print("WARNING: Answers is empty or not a dictionary")
            answers = {"stage_1": {}, "stage_2": {}}
        
        # Проверка формата ответов для отладки
        print(f"DEBUG: Raw answers received by analyzer: {json.dumps(answers, ensure_ascii=False, indent=2)}")
        
        # First, load the full question data
        try:
            with open('static/js/career_test_questions.json', 'r', encoding='utf-8') as f:
                all_questions = json.load(f)
        except Exception as e:
            print(f"ERROR: Could not load questions data: {e}")
            all_questions = {}
        
        # Prepare enriched answers with full question and answer text
        enriched_data = {
            "stage_1": [],
            "stage_2": []
        }
        
        for stage_name in ["stage_1", "stage_2"]:
            if stage_name not in answers or stage_name not in all_questions:
                print(f"WARNING: Missing {stage_name} in answers or questions")
                continue
                
            # Get the questions for this stage
            stage_questions = all_questions[stage_name].get("questions", [])
            stage_answers = answers[stage_name]
            
            # Create a lookup by question ID
            question_lookup = {str(q["id"]): q for q in stage_questions}
            
            # IMPORTANT: Check if we have text answers stored directly
            have_text_answers = 'answer_texts' in answers and stage_name in answers['answer_texts']
            
            # Process each answer with full context
            for question_id, selected_options in stage_answers.items():
                if question_id not in question_lookup:
                    print(f"WARNING: Question ID {question_id} not found in question_lookup")
                    continue
                    
                question_data = question_lookup[question_id]
                question_text = question_data.get("text", "")
                question_type = question_data.get("type", "")
                
                # Get the selected option text - preferring stored text values if available
                if have_text_answers and question_id in answers['answer_texts'][stage_name]:
                    # Use the stored text values directly
                    selected_texts = answers['answer_texts'][stage_name][question_id]
                    selected_option_ids = selected_options  # Keep the raw IDs for reference
                    
                    print(f"Using pre-stored text values for question {question_id}: {selected_texts}")
                else:
                    # Fallback to looking up text from option data
                    options = question_data.get("options", [])
                    option_lookup = {}
                    
                    # Build a comprehensive lookup map for options
                    for opt in options:
                        opt_id = str(opt.get("id", ""))
                        opt_value = str(opt.get("value", opt_id))
                        opt_text = opt.get("text", "")
                        
                        # Map both ID and value to text
                        option_lookup[opt_id] = opt_text
                        option_lookup[opt_value] = opt_text
                        
                        # Also map the text itself for direct matching
                        option_lookup[opt_text] = opt_text
                    
                    selected_texts = []
                    selected_option_ids = selected_options
                    
                    # Try to map each option to its text
                    for option_id in selected_options:
                        option_str = str(option_id)
                        
                        if option_str in option_lookup:
                            selected_texts.append(option_lookup[option_str])
                        elif question_type == 'riasec' and option_str.isdigit():
                            # Special case for RIASEC test ratings
                            rating = int(option_str)
                            selected_texts.append(f"Оценка {rating} из 5")
                        else:
                            # If we can't find a mapping, use the raw value
                            selected_texts.append(option_str)
                            print(f"WARNING: Could not find text for option '{option_str}' in question {question_id}")
                
                # Add the enriched data with both question and selected answer text
                enriched_data[stage_name].append({
                    "question_id": question_id,
                    "question_text": question_text,
                    "type": question_type,
                    "selected_options": selected_texts,  # This now contains the actual text of options
                    "selected_option_ids": selected_option_ids  # This contains the raw IDs/values
                })
        
        # Initialize Gemini AI model - THIS WAS MISSING
        model = genai.GenerativeModel('gemini-1.5-pro-latest')
        
        # Create enhanced prompt with text focus
        prompt = f"""Как карьерный консультант со специализацией на рынке труда Казахстана, проанализируй эти результаты тестов и рекомендуй 10 наиболее подходящих профессий для этого человека.

ОБОГАЩЕННЫЕ ДАННЫЕ ТЕСТА (полный текст вопросов и ответов):
{json.dumps(enriched_data, ensure_ascii=False, indent=2)}

Тест включает два раздела:
1. Базовый тест с вопросами об интересах, предпочтениях в работе и мотивации
2. Тест Голланда (RIASEC) для определения профессионального типа личности

Выполни тщательный анализ следующим образом:
1. Определи ключевые сильные стороны и предпочтения на основе ответов всех 2 тестов
2. Выяви профессиональные типы по Голланду на основе теста RIASEC
3. Определи основные мотивационные факторы из ответов первого теста
4. Оцени конкретные интересы и склонности на основе выбранных вариантов ответа
5. Сопоставь полученные данные с требованиями профессий на рынке труда Казахстана

ВАЖНО: Анализируй смысл текста выбранных ответов, а не их числовые коды или идентификаторы.

Для каждой рекомендуемой профессии предоставь:
1. Точное название профессии на русском языке
2. Процент совместимости (от 60% до 98%) на основе всестороннего анализа ответов
3. Подробное описание (2-3 предложения) с акцентом на то, почему эта профессия подходит данному человеку
4. Реалистичную среднюю зарплату в Казахстане (в тенге) и в мире (в долларах США)

Формат ответа должен быть в виде массива JSON с такой структурой:
[
  {{
    "name": "Название профессии",
    "slug": "название-профессии",
    "compatibility": 92,
    "description": "Детальное описание профессии с учетом личных качеств человека",
    "salary": {{
      "kz": "X,XXX,XXX тенге в год",
      "global": "$XX,XXX в год"
    }}
  }},
  ...
]

Важно: 
- Используй ТОЛЬКО АКТУАЛЬНЫЕ И ТОЧНЫЕ данные о зарплатах и требованиях для Казахстана
- Учитывай региональную специфику рынка труда Казахстана
- Рекомендации должны быть ПЕРСОНАЛИЗИРОВАННЫМИ на основе конкретных ответов тестируемого
- Процент совместимости должен точно отражать, насколько профессия соответствует всему комплексу характеристик тестируемого
- Все описания должны быть на русском языке и включать конкретные причины, почему данная профессия рекомендуется
- ОТВЕЧАЙ ТОЛЬКО В ФОРМАТЕ JSON, без дополнительного текста
"""

        # Print the AI prompt to console
        print("\n===== CAREER TEST AI PROMPT WITH TEXT VALUES =====")
        print(prompt)
        print("=================================\n")

        # Get recommendations from Gemini
        print("Sending request to Gemini API...")
        response = model.generate_content(prompt)
        print("Received response from Gemini API")
        result_text = response.text
        
        # Try to extract JSON from response
        try:
            import re
            json_match = re.search(r'(\[[\s\S]*\])', result_text)
            if json_match:
                json_str = json_match.group(1)
                recommendations = json.loads(json_str)
            else:
                recommendations = json.loads(result_text)
                
            # Sort by compatibility score (highest first)
            recommendations.sort(key=lambda x: x.get('compatibility', 0), reverse=True)
            
            # Always use fresh results, but still cache them for reference
            if user_id:
                cache_data = {
                    'recommendations': recommendations,
                    'generated_at': datetime.datetime.now(tz=datetime.timezone.utc),
                    'user_id': user_id,
                    'raw_answers': answers,
                    'enriched_answers': enriched_data
                }
                db.collection('career_recommendations').document(user_id).set(cache_data)
            
            # Return top 10 recommendations
            return recommendations[:10]
            
        except json.JSONDecodeError:
            # Fallback with generic recommendations in Russian if parsing fails
            print("Error parsing JSON recommendations from AI")
            return [
                {
                    "name": "Программист",
                    "slug": "programmist",
                    "compatibility": 92,
                    "description": "Разрабатывает программное обеспечение и приложения. Исходя из ваших аналитических способностей и интереса к технологиям, эта профессия отлично соответствует вашему профилю.",
                    "salary": {
                        "kz": "4,800,000 тенге в год",
                        "global": "$85,000 в год"
                    }
                },
                {
                    "name": "Специалист по анализу данных",
                    "slug": "specialist-po-analizu-dannyh",
                    "compatibility": 88,
                    "description": "Анализирует большие наборы данных для получения информации и решения сложных проблем. Ваша способность к логическому мышлению и интерес к исследованиям делают эту профессию высоко релевантной для вас.",
                    "salary": {
                        "kz": "5,200,000 тенге в год",
                        "global": "$95,000 в год"
                    }
                },
                {
                    "name": "Инженер-нефтяник",
                    "slug": "inzhener-neftyanik",
                    "compatibility": 86,
                    "description": "Разрабатывает и совершенствует методы добычи нефти и газа из земли. Учитывая ваш интерес к прикладным наукам и предпочтение работы со сложными техническими системами, эта профессия может стать отличным выбором.",
                    "salary": {
                        "kz": "7,500,000 тенге в год",
                        "global": "$110,000 в год"
                    }
                }
            ]
            
    except Exception as e:
        print(f"Error analyzing career test results: {e}")
        import traceback
        traceback.print_exc()
        
        # Return fallback recommendations in Russian
        return [
            {
                "name": "Разработчик программного обеспечения",
                "slug": "razrabotchik-programmnogo-obespecheniya",
                "compatibility": 85,
                "description": "Проектирует и разрабатывает компьютерное программное обеспечение и приложения. Эта профессия хорошо соответствует вашим техническим навыкам и аналитическому складу ума.",
                "salary": {
                    "kz": "4,800,000 тенге в год",
                    "global": "$85,000 в год"
                }
            },
            {
                "name": "Учитель",
                "slug": "uchitel",
                "compatibility": 82,
                "description": "Обучает учеников различным предметам и на разных уровнях образования. Ваши коммуникативные навыки и желание помогать другим делают эту профессию перспективным вариантом для вас.",
                "salary": {
                    "kz": "2,000,000 тенге в год",
                    "global": "$45,000 в год"
                }
            }
        ]

@app.route('/university-finder')
@login_required
def university_finder():
    user_id = session['user_id']
    
    try:
        # Получаем данные пользователя
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict() or {}
        
        # Получаем академическую информацию
        academic_info = user_data.get('academic_info', {})
        
        # Получаем сертификаты
        certificates = []
        certs_query = db.collection('users').document(user_id).collection('certificates').stream()
        for cert_doc in certs_query:
            cert_data = cert_doc.to_dict()
            cert_data['id'] = cert_doc.id
            certificates.append(cert_data)
        
        # Получаем историю рекомендаций
        recommendations_history = []
        recommendations_query = db.collection('university_recommendations').where('user_id', '==', user_id).order_by('created_at', direction=firestore.Query.DESCENDING).limit(5).stream()
        
        for rec_doc in recommendations_query:
            rec_data = rec_doc.to_dict()
            recommendations_history.append({
                'id': rec_doc.id,
                'country': rec_data.get('country', 'Неизвестно'),
                'created_at': rec_data.get('created_at'),
                'universities': list(rec_data.get('images', {}).keys())[:3]  # Только первые 3 университета
            })
        
        return render_template(
            'university_finder.html',
            user_data=user_data,
            academic_info=academic_info,
            certificates=certificates,
            recommendations_history=recommendations_history,
            current_user_avatar=get_current_user_avatar(),
            current_username=get_current_username()
        )
        
    except Exception as e:
        print(f"Ошибка на странице подбора университетов: {e}")
        import traceback
        traceback.print_exc()
        flash('Произошла ошибка при загрузке страницы')
        return redirect(url_for('profile'))


def generate_university_recommendations(profile, country):
    """Генерировать рекомендации университетов с использованием русскоязычного промпта"""
    model = genai.GenerativeModel('gemini-1.5-pro-latest')
    
    # Упаковываем данные профиля для промпта
    gpa = profile.get('academic', {}).get('gpa', 'Не указан')
    sat_score = profile.get('academic', {}).get('sat_score', 'Не указан')
    toefl_score = profile.get('academic', {}).get('toefl_score', 'Не указан') 
    ielts_score = profile.get('academic', {}).get('ielts_score', 'Не указан')
    skills = ", ".join(profile.get('skills', []))
    languages = [f"{lang.get('name')} ({lang.get('level')})" for lang in profile.get('academic_info', {}).get('languages', [])]
    achievements = [a.get('title') for a in profile.get('academic_info', {}).get('achievements', [])]
    specialty = profile.get('specialty', 'Не указана')
    goals = profile.get('goals', 'Не указаны')
    
    # Подготовка промпта на русском языке
    prompt = f"""Как эксперт по международному образованию со специализацией на университетах {country}, предоставь подробные рекомендации по 5 лучшим университетам, которые идеально подходят для этого студента.

ПРОФИЛЬ СТУДЕНТА:
- GPA: {gpa} 
- Балл SAT: {sat_score}
- Балл TOEFL: {toefl_score}
- Балл IELTS: {ielts_score}
- Навыки: {skills or "Разнообразные навыки"}
- Языки: {", ".join(languages) if languages else "Не указаны"}
- Достижения: {", ".join(achievements) if achievements else "Не указаны"}
- Специализация: {specialty}
- Академические цели: {goals}

ВАЖНЫЕ ПРАВИЛА - ВНИМАТЕЛЬНО ПРОЧИТАЙ:
- НИКОГДА не пиши "похож на [другой университет]" или "как упомянуто выше" - каждый университет должен иметь 100% оригинальный контент
- НИКОГДА не пиши "посетите сайт университета" или "свяжитесь с приемной комиссией" - предоставь фактическую информацию
- ВСЕГДА указывай ТОЧНЫЕ цифры для стоимости обучения, расходов, сроков и требований (конкретные числа, а не диапазоны)
- ПРЕДОСТАВЛЯЙ уникальную, детальную информацию для КАЖДОЙ категории - без пропусков разделов
- ВСЯ информация должна быть специфичной для КАЖДОГО университета (не копируй/вставляй между рекомендациями)
- Фокусируйся на предоставлении ДЕЙСТВЕННЫХ деталей, которые студенты могут использовать немедленно
- Включай ТОЧНЫЕ названия программ, а не просто общие области обучения
- Для рейтингов предоставляй КОНКРЕТНЫЕ цифры (например, "занимает #25 место в мире, #12 по информатике")

ФОРМАТ ДЛЯ КАЖДОГО УНИВЕРСИТЕТА:
**University: [Название]**
**Location:** [Точный город, регион и страна]
**Relevant Programs:**
* [Список минимум 4-5 конкретных программ обучения с их точными названиями]
* [Включи специализации и уникальные особенности каждой программы]

**Admission Requirements:**
* GPA: [Точный минимум со шкалой, например, "3.5/4.0"]
* TOEFL/IELTS: [Точные требования по баллам]
* [Любые дополнительные требования, такие как портфолио, интервью, предварительные условия]
* [Сборы за подачу заявления и процессы]

**Application Deadlines:**
* [Точные календарные даты для каждого тура приема]
* [Сроки подачи заявок на стипендии]
* [Приоритетные и обычные сроки]

**Estimated Costs:**
* Обучение: [Точная годовая стоимость в местной валюте И долларах США]
* Проживание: [Диапазон вариантов с конкретными ценами]
* Расходы на жизнь: [Детальная разбивка по категориям]
* Финансовая помощь: [Доступные стипендии с суммами]

**Notable Strengths:**
* [Список 4-5 уникальных преимуществ, особенно актуальных для этого студента]
* [Исследовательские учреждения, специальные программы, возможности стажировки]
* [Связи с индустрией и статистика трудоустройства]
* [Услуги для студентов и академическая поддержка]

**QS World Ranking:** [Конкретный общий рейтинг И рейтинги по предметам]

Помни: каждая рекомендация университета ДОЛЖНА быть ПОЛНОСТЬЮ УНИКАЛЬНОЙ с КОНКРЕТНОЙ, ДЕТАЛЬНОЙ информацией в КАЖДОМ разделе. Твой ответ будет использоваться непосредственно для заявлений в университеты, поэтому убедись, что вся информация исчерпывающая, точная и немедленно полезная.
пиши информацию на русском
"""

    # Получаем ответ от модели
    response = model.generate_content(prompt)
    return response.text


@app.route('/get-university-details', methods=['POST'])
@login_required
def get_university_details():
    """API эндпоинт для получения подробной информации об университете"""
    user_id = session['user_id']
    data = request.json
    
    try:
        # Проверяем наличие обязательных данных
        if not data or 'university' not in data:
            return jsonify({'error': 'Необходимо указать название университета'}), 400
            
        university_name = data.get('university')
        country = data.get('country', 'Не указано')
        
        # Проверяем, есть ли кэшированные данные для этого университета
        university_slug = slugify(university_name)
        cache_ref = db.collection('university_details').document(f"{university_slug}_{slugify(country)}")
        cache_doc = cache_ref.get()
        
        if cache_doc.exists:
            # Проверяем, не устарели ли данные (старше 30 дней)
            cache_data = cache_doc.to_dict()
            timestamp = cache_data.get('timestamp')
            
            if isinstance(timestamp, datetime.datetime):
                age = datetime.datetime.now(tz=datetime.timezone.utc) - timestamp
                if age.days < 30:
                    print(f"Используем кэшированные данные для университета {university_name}")
                    return jsonify({
                        'details': cache_data.get('details', {}),
                        'from_cache': True
                    })
        
        # Инициализируем модель Gemini
        model = genai.GenerativeModel('gemini-1.5-pro-latest')
        
        # Создаем промпт для получения подробной информации
        prompt = f"""Как эксперт по высшему образованию в {country}, создай подробное и точное описание университета "{university_name}" специально для российских и казахстанских студентов.

Включи следующую информацию:
1. Общий обзор университета (основные функции, обязанности, роль в образовательной системе)
2. Необходимое образование и квалификации (конкретные требования, специальности)
3. Точный диапазон стоимости обучения в стране (в местной валюте) и глобально (в долларах США)
4. Конкретные перспективы для поступления с учетом текущих тенденций
5. Подробный список ключевых академических навыков и личностных качеств для поступления
6. Типичные учебные задачи, условия и среда обучения
7. Четкие пути развития с указанием конкретных специализаций
8. Смежные направления обучения и возможности смены специализации
9. Список ведущих работодателей для выпускников данного университета

Формат ответа - объект JSON со следующими ключами:
{{
  "name": "Название университета",
  "overview": "Подробный обзор...",
  "education": ["Пункт об образовании 1", "Пункт 2", ...],
  "salary": {{
    "kz": {{
      "range": "X,XXX,XXX - X,XXX,XXX тенге в год",
      "average": "X,XXX,XXX тенге в год"
    }},
    "global": {{
      "range": "$XX,XXX - $XXX,XXX в год",
      "average": "$XX,XXX в год"
    }}
  }},
  "job_outlook": "Подробные перспективы обучения...",
  "skills": ["Навык 1", "Навык 2", ...],
  "personality": ["Качество 1", "Качество 2", ...],
  "daily_work": ["Задача 1", "Задача 2", ...],
  "advancement": ["Путь 1", "Путь 2", ...],
  "related_professions": ["Родственная специальность 1", "Родственная специальность 2", ...],
  "top_employers": ["Работодатель 1", "Работодатель 2", ...]
}}

КРИТИЧЕСКИ ВАЖНО:
- Вся информация ДОЛЖНА быть конкретной, актуальной и основываться на реальных данных об университете
- Информация о стоимости должна отражать реальную ситуацию на 2023-2024 годы
- Образовательные требования должны соответствовать реальным требованиям университета
- Перечень работодателей должен включать реальные компании
- ВСЯ ИНФОРМАЦИЯ ДОЛЖНА БЫТЬ ИСКЛЮЧИТЕЛЬНО НА РУССКОМ ЯЗЫКЕ
- Ответ должен содержать ТОЛЬКО JSON-объект без дополнительного текста
"""

        # Получаем данные от Gemini
        response = model.generate_content(prompt)
        result_text = response.text
        
        # Пытаемся извлечь JSON из ответа
        try:
            import re
            json_match = re.search(r'(\{[\s\S]*\})', result_text)
            if json_match:
                json_str = json_match.group(1)
                university_data = json.loads(json_str)
            else:
                university_data = json.loads(result_text)
                
            # Добавляем метаданные
            university_data['generated_at'] = datetime.datetime.now(tz=datetime.timezone.utc).isoformat()
            university_data['slug'] = university_slug
            
            # Сохраняем в кэш
            cache_data = {
                'details': university_data,
                'timestamp': datetime.datetime.now(tz=datetime.timezone.utc),
                'university': university_name,
                'country': country
            }
            cache_ref.set(cache_data)
            
            return jsonify({
                'details': university_data,
                'from_cache': False
            })
            
        except json.JSONDecodeError:
            # Создаем запасные данные на русском, если парсинг не удался
            print(f"Ошибка парсинга JSON для университета {university_name}")
            fallback_data = {
                "name": university_name,
                "overview": f"Университет {university_name} - это учебное заведение, которое предлагает различные программы обучения. Для получения более подробной информации рекомендуется посетить официальный сайт университета или связаться с приемной комиссией.",
                "education": [
                    "Для поступления требуется аттестат о среднем образовании", 
                    "Необходимо сдать вступительные экзамены или предоставить результаты международных тестов",
                    "Для иностранных студентов может потребоваться подтверждение знания языка"
                ],
                "salary": {
                    "kz": {
                        "range": "3,000,000 - 7,000,000 тенге в год",
                        "average": "5,000,000 тенге в год"
                    },
                    "global": {
                        "range": "$10,000 - $30,000 в год",
                        "average": "$20,000 в год"
                    }
                },
                "job_outlook": "Выпускники данного университета обычно востребованы на рынке труда. Конкретные перспективы зависят от выбранной специальности и индивидуальных достижений студента.",
                "skills": [
                    "Академические навыки", 
                    "Навыки критического мышления", 
                    "Коммуникативные навыки",
                    "Навыки самоорганизации",
                    "Технические навыки по выбранной специальности"
                ],
                "personality": [
                    "Целеустремленность", 
                    "Ответственность", 
                    "Адаптивность",
                    "Стрессоустойчивость",
                    "Любознательность"
                ],
                "daily_work": [
                    "Посещение лекций и семинаров", 
                    "Работа над проектами и заданиями", 
                    "Проведение исследований",
                    "Участие в студенческих мероприятиях",
                    "Сдача экзаменов и тестов"
                ],
                "advancement": [
                    "Бакалавриат → Магистратура → Докторантура", 
                    "Возможность получения двойного диплома", 
                    "Участие в программах обмена с другими университетами"
                ],
                "related_professions": [
                    "Различные специальности в зависимости от факультета", 
                    "Возможность перевода между факультетами",
                    "Междисциплинарные программы"
                ],
                "top_employers": [
                    "Ведущие компании в соответствующей отрасли", 
                    "Международные корпорации", 
                    "Государственные учреждения",
                    "Исследовательские институты",
                    "Образовательные учреждения"
                ],
                "generated_at": datetime.datetime.now(tz=datetime.timezone.utc).isoformat(),
                "slug": university_slug
            }
            
            # Сохраняем запасные данные в кэш
            cache_data = {
                'details': fallback_data,
                'timestamp': datetime.datetime.now(tz=datetime.timezone.utc),
                'university': university_name,
                'country': country,
                'is_fallback': True
            }
            cache_ref.set(cache_data)
            
            return jsonify({
                'details': fallback_data,
                'from_cache': False,
                'is_fallback': True
            })
            
    except Exception as e:
        print(f"Ошибка получения деталей университета {university_name}: {e}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'error': str(e),
            'details': {
                "name": university_name,
                "overview": "Не удалось получить информацию об этом университете. Пожалуйста, попробуйте позже или обратитесь к официальным источникам.",
                "education": ["Информация временно недоступна"],
                "salary": {
                    "kz": {
                        "range": "Информация недоступна",
                        "average": "Информация недоступна"
                    },
                    "global": {
                        "range": "Информация недоступна",
                        "average": "Информация недоступна"
                    }
                }
            }
        }), 500
@app.route('/university-recommendation')
@login_required
def university_recommendation():
    try:
        user_id = session['user_id']
        user_doc = db.collection('users').document(user_id).get()
        
        # Преобразуем DocumentSnapshot в словарь
        user_data = user_doc.to_dict() or {}
        
        # Преобразуем все вложенные данные в базовые типы
        academic_info = {}
        if 'academic_info' in user_data:
            academic_info = {
                'gpa': user_data['academic_info'].get('gpa', ''),
                'sat_score': user_data['academic_info'].get('sat_score', ''),
                'toefl_score': user_data['academic_info'].get('toefl_score', ''),
                'ielts_score': user_data['academic_info'].get('ielts_score', ''),
                'languages': list(user_data['academic_info'].get('languages', [])),
                'achievements': list(user_data['academic_info'].get('achievements', []))
            }

        skills = list(user_data.get('skills', []))
        education = str(user_data.get('education', ''))
        goals = str(user_data.get('goals', ''))
        
        # Получаем и преобразуем сертификаты
        certificates_data = []
        certs_query = db.collection('users').document(user_id).collection('certificates').stream()
        
        for cert_doc in certs_query:
            cert_dict = cert_doc.to_dict()
            # Преобразуем timestamp в строку, если он есть
            if 'uploaded_at' in cert_dict:
                cert_dict['uploaded_at'] = cert_dict['uploaded_at'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(cert_dict['uploaded_at'], datetime.datetime) else str(cert_dict['uploaded_at'])
            cert_dict['id'] = cert_doc.id
            # Убеждаемся, что все значения сериализуемы
            cert_dict = {
                'id': cert_doc.id,
                'title': cert_dict.get('title', ''),
                'file_url': cert_dict.get('file_url', ''),
                'uploaded_at': cert_dict.get('uploaded_at', '')
            }
            certificates_data.append(cert_dict)

        # Подготавливаем безопасный для JSON объект с данными пользователя
        safe_user_data = {
            'username': user_data.get('username', ''),
            'display_username': user_data.get('display_username', ''),
            'email': user_data.get('email', ''),
            'academic_info': academic_info,
            'skills': skills,
            'education': education,
            'goals': goals
        }

        return render_template('university_recommendation.html',
                             user_data=safe_user_data,
                             academic_info=academic_info,
                             skills=skills,
                             education=education,
                             goals=goals,
                             certificates=certificates_data)
                             
    except Exception as e:
        print(f"Error in university recommendation route: {e}")
        # Логируем полную ошибку для отладки
        import traceback
        traceback.print_exc()
        flash('Error loading profile data')
        return redirect(url_for('profile'))
@app.route('/get-university-recommendations', methods=['POST'])
@login_required
def get_university_recommendations():
    user_id = session['user_id']
    data = request.json
    
    # Check for required data
    if not data or 'country' not in data:
        return jsonify({'error': 'Необходимо указать страну для рекомендаций'}), 400
    
    # Extract data from request
    country = data.get('country')
    profession = data.get('profession', '')  # Get profession if provided
    
    try:
        # Get user data directly from database
        user_doc = db.collection('users').document(user_id).get()
        user_data = user_doc.to_dict() or {}
        
        # Get academic information
        academic_info = user_data.get('academic_info', {})
        if not isinstance(academic_info, dict):
            academic_info = {}
            
        # Extract values from form
        gpa = data.get('academic', {}).get('gpa') or academic_info.get('gpa', '')
        sat_score = data.get('academic', {}).get('satScore') or academic_info.get('sat_score', '')
        toefl_score = data.get('academic', {}).get('toeflScore') or academic_info.get('toefl_score', '')
        ielts_score = data.get('academic', {}).get('ieltsScore') or academic_info.get('ielts_score', '')
        
        # Convert to strings
        gpa = str(gpa) if gpa else ''
        sat_score = str(sat_score) if sat_score else ''
        toefl_score = str(toefl_score) if toefl_score else ''
        ielts_score = str(ielts_score) if ielts_score else ''
        
        # Get other profile data
        specialty = data.get('profile', {}).get('specialty') or user_data.get('specialty', '')
        education = data.get('profile', {}).get('education') or user_data.get('education', '')
        goals = data.get('profile', {}).get('goals') or user_data.get('goals', '')
        bio = user_data.get('bio', '')
        
        # If profession is provided, use it to enhance the specialty
        if profession and not specialty:
            specialty = profession
        elif profession:
            specialty = f"{specialty}, {profession}"
        
        # Get skills
        skills = data.get('profile', {}).get('skills') or user_data.get('skills', [])
        if not isinstance(skills, list):
            skills = []
            
        # Get certificates
        certificates = []
        certs_query = db.collection('users').document(user_id).collection('certificates').stream()
        for cert_doc in certs_query:
            cert_data = cert_doc.to_dict()
            certificates.append(cert_data)
            
        # Get certificate titles for analysis
        certificate_titles = [cert.get('title', '') for cert in certificates]
        
        # Get languages
        languages = []
        if academic_info and 'languages' in academic_info:
            for lang in academic_info.get('languages', []):
                if isinstance(lang, dict) and 'name' in lang:
                    languages.append(f"{lang['name']} ({lang.get('level', 'Conversational')})")
                    
        # Get achievements
        achievements = []
        if academic_info and 'achievements' in academic_info:
            for achievement in academic_info.get('achievements', []):
                if isinstance(achievement, dict) and 'title' in achievement:
                    achievements.append(achievement['title'])
                    
        # Determine interests based on certificates, specialty, and skills
        interests = []
        interest_mapping = {
            'Информатика': ['python', 'программирование', 'код', 'software', 'разработчик', 'web', 'app', 'computer science', 'data science', 'алгоритм'],
            'Инженерия': ['инженер', 'engineering', 'механический', 'электрический', 'гражданский', 'аэрокосмический'],
            'Бизнес': ['бизнес', 'management', 'финансы', 'экономика', 'маркетинг'],
            'Медицина': ['медицина', 'medical', 'здоровье', 'биология', 'доктор'],
            'Искусство': ['дизайн', 'art', 'музыка', 'творчество', 'media'],
            'Право': ['право', 'юрист', 'адвокат', 'justice'],
            'Образование': ['преподавание', 'образование', 'педагогика']
        }
        
        # Check certificates, specialty, skills, and profession for keywords
        all_text = ' '.join([' '.join(certificate_titles), specialty.lower(), ' '.join(skills), goals.lower(), profession.lower()]).lower()
        
        for field, keywords in interest_mapping.items():
            if any(keyword in all_text for keyword in keywords):
                interests.append(field)
                
        # If profession is provided, add it to interests directly
        if profession and profession not in interests:
            interests.append(profession)
                
        # If no interests detected, use specialty or profession
        if not interests:
            if specialty:
                interests.append(specialty)
            elif profession:
                interests.append(profession)
            
        # Create full profile for recommendations
        profile = {
            'gpa': gpa,
            'sat_score': sat_score,
            'toefl_score': toefl_score,
            'ielts_score': ielts_score,
            'skills': skills,
            'education': education,
            'goals': goals,
            'bio': bio,
            'interests': interests[:3],  # Top 3 interests
            'languages': languages,
            'achievements': achievements,
            'certificates': certificate_titles,
            'target_country': country,
            'academic_info': academic_info,
            'target_profession': profession  # Include target profession
        }
        
        # Generate recommendations using our Russian-language prompt
        recommendations = generate_university_recommendations(profile, country)
        
        # Extract university names for images
        universities = extract_universities(recommendations)
        
        # Calculate admission chances based on profile
        admission_chances = calculate_admission_chances(universities, profile)
        
        # Get images for universities
        images = {}
        for university in universities:
            # Get image via Google API or use cache
            image_url = get_university_image(university, country)
            if image_url:
                images[university] = image_url
            else:
                # If API failed, use a universal image based on name
                images[university] = f"https://ui-avatars.com/api/?name={university.replace(' ', '+')}&size=256&background=random"
        
        # Save recommendation to database
        recommendation_data = {
            'user_id': user_id,
            'country': country,
            'recommendations': recommendations,
            'images': images,
            'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
            'profile_snapshot': profile,
            'admission_chances': admission_chances,
            'universities': universities,
            'profession': profession  # Store the profession
        }
        
        # Save to recommendations collection
        rec_ref = db.collection('university_recommendations').add(recommendation_data)
        
        # Update last recommendation in user profile
        user_ref = db.collection('users').document(user_id)
        user_ref.update({
            'last_university_recommendation': {
                'created_at': datetime.datetime.now(tz=datetime.timezone.utc),
                'country': country,
                'universities': universities,
                'recommendation_id': rec_ref[1].id,
                'profession': profession  # Include profession
            }
        })
        
        return jsonify({
            'recommendations': recommendations,
            'images': images,
            'admission_chances': admission_chances,
            'universities': universities,
            'profession': profession,  # Return profession in response
            'profile_analyzed': {
                'gpa': profile['gpa'],
                'toefl': profile['toefl_score'],
                'ielts': profile['ielts_score'],
                'sat': profile['sat_score'],
                'certificates_count': len(profile['certificates']),
                'interests_identified': profile['interests'],
                'skills_considered': len(profile['skills']),
                'languages_considered': len(profile['languages']),
                'achievements_considered': len(profile['achievements']),
                'target_profession': profession
            }
        })
        
    except Exception as e:
        print(f"Error in university recommendation process: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
def calculate_admission_chances(universities, profile):
    """Calculate realistic admission chances based on student profile and university selectivity"""
    chances = {}
    
    # Get basic academic metrics
    try:
        gpa = float(profile.get('gpa', 0) or 0)
        if gpa > 4.0:  # Adjust if using different scale
            gpa = gpa / 10 * 4.0 if gpa <= 10 else 3.0
    except (ValueError, TypeError):
        gpa = 3.0
    
    try:
        toefl = int(profile.get('toefl_score', 0) or 0)
    except (ValueError, TypeError):
        toefl = 90
    
    try:
        ielts = float(profile.get('ielts_score', 0) or 0)
    except (ValueError, TypeError):
        ielts = 6.5
        
    try:
        sat = int(profile.get('sat_score', 0) or 0)
        if sat < 400 or sat > 1600:  # Old SAT format or invalid
            sat = 1200
    except (ValueError, TypeError):
        sat = 1200
    
    # University selectivity tiers
    university_tiers = {
        # Ultra-selective (Tier 1)
        'Harvard University': 95,
        'Stanford University': 94,
        'Massachusetts Institute of Technology': 95,
        'California Institute of Technology': 94,
        'Princeton University': 93,
        'Yale University': 93,
        'University of Oxford': 92,
        'University of Cambridge': 92,
        'Columbia University': 91,
        'University of Chicago': 90,
        
        # Very selective (Tier 2)
        'University of Toronto': 85,
        'University of British Columbia': 84,
        'McGill University': 83,
        'University of California, Berkeley': 88,
        'University of California, Los Angeles': 87,
        'UCL': 86,
        'University of Edinburgh': 85,
        'University of Michigan': 84,
        'Cornell University': 88,
        'University of Pennsylvania': 89,
        
        # Selective (Tier 3)
        'University of Washington': 80,
        'University of Wisconsin-Madison': 79,
        'Pennsylvania State University': 78,
        'University of Minnesota': 78,
        'Purdue University': 77,
        'University of Manchester': 76,
        'University of Amsterdam': 75,
        'University of Melbourne': 78,
        'University of Sydney': 77,
        
        # Moderately selective (Tier 4)
        'Arizona State University': 70,
        'Michigan State University': 71,
        'University of Alberta': 71,
        'Simon Fraser University': 70,
        'University of Calgary': 69,
        'University of Sheffield': 72,
        'University of Leeds': 72,
        'Monash University': 73
    }
    
    # Calculate chances for each university
    for university in universities:
        # Base chances depend on university selectivity
        difficulty = university_tiers.get(university, 75)  # Default level if unknown
        
        # Initial value with difficulty adjustment
        base_chance = max(100 - difficulty, 25)
        
        # GPA adjustment (significant factor)
        if gpa >= 3.9:
            base_chance += 20
        elif gpa >= 3.7:
            base_chance += 15
        elif gpa >= 3.5:
            base_chance += 10
        elif gpa >= 3.3:
            base_chance += 5
        elif gpa < 3.0:
            base_chance -= 15
        
        # Language score adjustment
        if toefl >= 110 or ielts >= 8.0:
            base_chance += 10
        elif toefl >= 100 or ielts >= 7.5:
            base_chance += 7
        elif toefl >= 90 or ielts >= 7.0:
            base_chance += 5
        elif toefl < 80 or ielts < 6.0:
            base_chance -= 15
            
        # SAT adjustment
        if sat >= 1550:
            base_chance += 15
        elif sat >= 1500:
            base_chance += 10
        elif sat >= 1450:
            base_chance += 8
        elif sat >= 1400:
            base_chance += 5
        elif sat < 1200:
            base_chance -= 10
        
        # Other factors
        skills_count = len(profile.get('skills', []))
        base_chance += min(skills_count * 1.5, 7.5)  # Up to 7.5% bonus
        
        certificates_count = len(profile.get('certificates', []))
        base_chance += min(certificates_count * 2, 10)  # Up to 10% bonus
        
        languages_count = len(profile.get('languages', []))
        base_chance += min(languages_count * 2, 6)  # Up to 6% bonus
        
        # Add some variability (±5%)
        import random
        variability = random.uniform(-5, 5)
        base_chance += variability
        
        # Ensure chances stay within realistic bounds
        if difficulty >= 90:  # Top tier
            max_chance = 80
        elif difficulty >= 85:  # Very selective
            max_chance = 85
        elif difficulty >= 80:  # Selective
            max_chance = 90
        else:  # Less selective
            max_chance = 95
            
        # Final chance calculation
        final_chance = min(max(int(base_chance), 25), max_chance)
        
        chances[university] = final_chance
    
    return chances
def use_fallback_image(university_name, country):
    """Improved fallback mechanism for university images"""
    # First try university-specific images
    university_images = {
        'Harvard University': 'https://images.unsplash.com/photo-1610816732435-69c37d5498d6',
        'Stanford University': 'https://images.unsplash.com/photo-1455582916367-25f75bfc6710',
        'MIT': 'https://images.unsplash.com/photo-1508231101580-eed9c8913fb4',
        # Your existing university mappings...
    }
    
    if university_name in university_images:
        return university_images[university_name]
    
    # Next try country-specific images
    country_images = {
        'USA': 'https://images.unsplash.com/photo-1498243691581-b145c3f54a5a',
        'UK': 'https://images.unsplash.com/photo-1607237138185-eedd9c632b0b',
        # Your existing country mappings...
    }
    
    if country in country_images:
        return country_images[country]
    
    # Use deterministic fallback based on university name
    generic_images = [
        'https://images.unsplash.com/photo-1541339907198-e08756dedf3f',
        'https://images.unsplash.com/photo-1607237138185-eedd9c632b0b',
        'https://images.unsplash.com/photo-1598826867442-11c9a9221cfa',
        'https://images.unsplash.com/photo-1523050854058-8df90110c9f1',
        'https://images.unsplash.com/photo-1592280771190-3e2e4d571952',
        'https://images.unsplash.com/photo-1544531585-9847b68c8c86',
        'https://images.unsplash.com/photo-1519452575417-564c1401ecc0',
        'https://images.unsplash.com/photo-1580332449505-682ff38a8b45'
    ]
    
    # Use first character of university name to deterministically select an image
    if university_name:
        index = ord(university_name[0].lower()) % len(generic_images)
        return generic_images[index]
    
    return generic_images[0]

def cache_university_image(university_name, image_url):
    """Cache university image URLs to reduce API calls"""
    try:
        # Get existing cache or create new one
        cache_ref = db.collection('image_cache').document('universities')
        cache_doc = cache_ref.get()
        
        if cache_doc.exists:
            cache_data = cache_doc.to_dict()
        else:
            cache_data = {}
        
        # Update with new image
        cache_data[university_name] = {
            'url': image_url,
            'timestamp': datetime.datetime.now(tz=datetime.timezone.utc)
        }
        
        # Save cache
        cache_ref.set(cache_data)
    except Exception as e:
        print(f"Error caching image: {e}")
# Улучшенная функция для получения изображения университета
def get_university_image(university_name, country):
    """Get university image using specific Google Custom Search API URL"""
    try:
        # Format the university name for the query
        formatted_name = ""
        for char in university_name:
            if char == ' ' or char == '+':
                formatted_name += '+'
            else:
                formatted_name += char
        
        # Create the search query with additional terms for better results
        search_query = f"{formatted_name}+university+campus"
        
        # Use the exact URL provided with hardcoded credentials
        search_url = f"https://www.googleapis.com/customsearch/v1?key=AIzaSyDptyzxGJg-aR5IldozvISzjNgF2_TISJo&cx=e1cac863f07bf4f8b&q={search_query}&searchType=image"
        
        # Make the request
        print(f"Searching for image: {university_name}")
        response = requests.get(search_url)
        
        # Check if we got a valid response
        if response.status_code != 200:
            print(f"Error response from Google API: {response.status_code} - {response.text}")
            return None
        
        # Parse the response JSON
        image_data = response.json()
        
        # Check if we have image results
        if 'items' in image_data and len(image_data['items']) > 0:
            # Get the first image URL
            image_url = image_data['items'][0]['link']
            print(f"Found image for {university_name}: {image_url}")
            return image_url
        else:
            print(f"No image results found for {university_name}")
            return None
            
    except Exception as e:
        print(f"Error getting image for {university_name}: {e}")
        return None



@app.route('/career-test/generate-results', methods=['POST'])
@login_required
def generate_career_results():
    """API endpoint to actually generate the results after showing animation"""
    user_id = session['user_id']
    
    try:
        # Get test answers
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        test_progress = test_progress_ref.get()
        
        if not test_progress.exists:
            return jsonify({'error': 'Результаты теста не найдены'}), 404
            
        progress_data = test_progress.to_dict()
        answers = progress_data.get('answers', {})
        
        # Print the actual answers for debugging
        print(f"User answers before analysis: {json.dumps(answers, ensure_ascii=False, indent=2)}")
        
        # Always analyze results using AI, ignoring cache
        recommendations = analyze_career_test_results(answers)
        
        # Update the test_progress document with the new recommendations
        test_progress_ref.update({
            'recommendations': recommendations
        })
        
        # Return the results
        return jsonify({
            'success': True,
            'recommendations': recommendations,
            'fresh_results': True
        })
        
    except Exception as e:
        print(f"Error generating results: {e}")
        return jsonify({'error': str(e)}), 500

def generate_profession_details(profession_name):
    """Generate detailed profession information in Russian with central profession cache"""
    try:
        # Generate a consistent slug for caching
        import re
        # Replace all non-alphanumeric characters with hyphens and convert to lowercase
        slug = re.sub(r'[^a-zA-Zа-яА-Я0-9]+', '-', profession_name.lower()).strip('-')
        
        # Check if cached profession data exists in central profession cache
        profession_ref = db.collection('profession_details').document(slug)
        profession_doc = profession_ref.get()
        
        if profession_doc.exists:
            # Use cached data
            print(f"Using cached profession data for {profession_name}")
            profession_data = profession_doc.to_dict()
            return profession_data
        
        # Initialize Gemini AI model
        model = genai.GenerativeModel('gemini-1.5-pro-latest')
        
        # Create prompt for profession details - RUSSIAN VERSION with enhanced detail
        prompt = f"""Как специалист по карьерам в Казахстане, создай исчерпывающее и точное описание профессии "{profession_name}" конкретно для казахстанского рынка труда.

Включи следующую информацию:
1. Детальный обзор профессии (основные функции, обязанности, роль в организации)
2. Необходимое образование и квалификации в контексте Казахстана (конкретные вузы, специальности)
3. Точный диапазон зарплат в Казахстане (в тенге) с учетом региональных различий и в мире (в долларах США)
4. Конкретные перспективы трудоустройства и роста в Казахстане с учетом текущих экономических тенденций
5. Подробный список ключевых профессиональных навыков и личностных качеств
6. Типичные рабочие задачи, условия труда и рабочая среда
7. Чёткие пути карьерного роста с указанием конкретных должностей и временных рамок
8. Смежные профессии и возможности смены специализации
9. Список ведущих работодателей в Казахстане с акцентом на региональное распределение

Формат ответа - объект JSON со следующими ключами:
{{
  "name": "Название профессии",
  "overview": "Подробный обзор...",
  "education": ["Пункт об образовании 1", "Пункт 2", ...],
  "salary": {{
    "kz": {{
      "range": "X,XXX,XXX - X,XXX,XXX тенге в год",
      "average": "X,XXX,XXX тенге в год"
    }},
    "global": {{
      "range": "$XX,XXX - $XXX,XXX в год",
      "average": "$XX,XXX в год"
    }}
  }},
  "job_outlook": "Подробные перспективы работы в Казахстане...",
  "skills": ["Навык 1", "Навык 2", ...],
  "personality": ["Качество 1", "Качество 2", ...],
  "daily_work": ["Задача 1", "Задача 2", ...],
  "advancement": ["Путь 1", "Путь 2", ...],
  "related_professions": ["Родственная профессия 1", "Родственная профессия 2", ...],
  "top_employers": ["Работодатель 1", "Работодатель 2", ...]
}}

КРИТИЧЕСКИ ВАЖНО:
- Вся информация ДОЛЖНА быть конкретной, актуальной и основываться на реальных данных рынка труда Казахстана
- Информация о зарплатах должна отражать реальную ситуацию в Казахстане на 2023-2024 годы
- Образовательные требования должны соответствовать казахстанской системе образования
- Перечень работодателей должен включать реальные компании, действующие в Казахстане
- ВСЯ ИНФОРМАЦИЯ ДОЛЖНА БЫТЬ ИСКЛЮЧИТЕЛЬНО НА РУССКОМ ЯЗЫКЕ
- Ответ должен содержать ТОЛЬКО JSON-объект без дополнительного текста
"""

        # Get details from Gemini
        response = model.generate_content(prompt)
        result_text = response.text
        
        # Try to extract JSON from response
        try:
            import re
            json_match = re.search(r'(\{[\s\S]*\})', result_text)
            if json_match:
                json_str = json_match.group(1)
                profession_data = json.loads(json_str)
            else:
                profession_data = json.loads(result_text)
                
            # Add timestamp and slug
            profession_data['generated_at'] = datetime.datetime.now(tz=datetime.timezone.utc).isoformat()
            profession_data['slug'] = slug
            
            # Store in central profession cache
            db.collection('profession_details').document(slug).set(profession_data)
            
            return profession_data
            
        except json.JSONDecodeError:
            # Fallback with generic data in Russian if parsing fails
            print(f"Error parsing JSON profession details from AI for {profession_name}")
            fallback_data = {
                "name": profession_name,
                "overview": "Эта профессия включает работу в специализированной области, требующей конкретных навыков и знаний. Специалисты данного профиля востребованы на рынке труда Казахстана и вносят значительный вклад в развитие соответствующей отрасли.",
                "education": [
                    "Высшее образование по специальности в одном из ведущих вузов Казахстана (КазНУ, ЕНУ, КБТУ)", 
                    "Профессиональные сертификаты от международных организаций",
                    "Регулярное повышение квалификации через специализированные курсы"
                ],
                "salary": {
                    "kz": {
                        "range": "3,000,000 - 7,000,000 тенге в год",
                        "average": "5,000,000 тенге в год"
                    },
                    "global": {
                        "range": "$40,000 - $100,000 в год",
                        "average": "$70,000 в год"
                    }
                },
                "job_outlook": "Перспективы для этой профессии в Казахстане умеренные с ожидаемым стабильным ростом. Спрос на квалифицированных специалистов особенно высок в крупных городах, таких как Алматы и Астана. С развитием отрасли ожидается увеличение количества вакансий и возможностей для карьерного роста.",
                "skills": [
                    "Коммуникабельность", 
                    "Решение проблем", 
                    "Технические знания",
                    "Работа в команде",
                    "Аналитическое мышление"
                ],
                "personality": [
                    "Внимание к деталям", 
                    "Умение работать в команде", 
                    "Адаптивность",
                    "Стрессоустойчивость",
                    "Ответственность"
                ],
                "daily_work": [
                    "Командные встречи", 
                    "Работа над проектами", 
                    "Профессиональное развитие",
                    "Взаимодействие с клиентами и партнерами",
                    "Решение текущих рабочих задач"
                ],
                "advancement": [
                    "Специалист → Старший специалист → Руководитель группы", 
                    "Специализация в узком профиле с повышением экспертности", 
                    "Переход в смежные области с расширением компетенций"
                ],
                "related_professions": [
                    "Похожая профессия 1", 
                    "Похожая профессия 2",
                    "Альтернативная карьерная траектория"
                ],
                "top_employers": [
                    "Казмунайгаз", 
                    "Казахтелеком", 
                    "Air Astana",
                    "Kaspi Bank",
                    "Самрук-Казына"
                ],
                "generated_at": datetime.datetime.now(tz=datetime.timezone.utc).isoformat(),
                "slug": slug
            }
            
            # Still cache the fallback data
            db.collection('profession_details').document(slug).set(fallback_data)
            
            return fallback_data
            
    except Exception as e:
        print(f"Error generating profession details for {profession_name}: {e}")
        import traceback
        traceback.print_exc()
        
        # Return fallback data in Russian
        return {
            "name": profession_name,
            "overview": "Подробная информация об этой профессии временно недоступна. Профессия представляет собой специализированную область деятельности, требующую определенных навыков и квалификации.",
            "education": ["Высшее образование", "Профессиональное обучение"],
            "salary": {
                "kz": {
                    "range": "3,000,000 - 6,000,000 тенге в год",
                    "average": "4,500,000 тенге в год"
                },
                "global": {
                    "range": "$40,000 - $80,000 в год",
                    "average": "$60,000 в год"
                }
            },
            "job_outlook": "Информация о перспективах работы в Казахстане временно недоступна.",
            "skills": ["Профессиональные навыки", "Технические знания"],
            "personality": ["Адаптивность", "Внимание к деталям"],
            "daily_work": ["Варьируется в зависимости от рабочего места"],
            "advancement": ["Карьерный рост от начальных до руководящих позиций"],
            "related_professions": ["Смежные профессиональные роли"],
            "top_employers": ["Ведущие компании Казахстана в соответствующей отрасли"],
            "generated_at": datetime.datetime.now(tz=datetime.timezone.utc).isoformat(),
            "slug": slug
        }

@app.route('/admin/test-career-prompt', methods=['GET'])
@login_required
def admin_test_career_prompt():
    """Route for administrators to test the career test system with random answers"""
    user_id = session['user_id']
    
    # Verify this is an admin
    if user_id != 'vVbXL4LKGidXtrKnvqa21gWRY3V2':  # Your admin ID
        return "Unauthorized", 403
    
    try:
        # Load test questions
        with open('static/js/career_test_questions.json', 'r', encoding='utf-8') as f:
            all_questions = json.load(f)
        
        # Generate random answers for all questions
        random_answers = {
            "stage_1": {},
            "stage_2": {}
        }
        
        # Stage 1 random answers
        stage1_questions = all_questions['stage_1']['questions']
        for question in stage1_questions:
            question_id = str(question['id'])
            options = question['options']
            
            # For radio buttons, select one random answer
            if question['type'] == 'radio':
                random_answers['stage_1'][question_id] = [random.choice([opt['id'] for opt in options])]
            
            # For checkboxes, select between 1 and 3 random answers
            elif question['type'] == 'checkbox':
                num_selections = random.randint(1, min(3, len(options)))
                random_answers['stage_1'][question_id] = random.sample([opt['id'] for opt in options], num_selections)
        
        # Stage 2 random answers
        stage2_questions = all_questions['stage_2']['questions']
        for question in stage2_questions:
            question_id = str(question['id'])
            # For Holland test, select a random answer (1-5)
            random_answers['stage_2'][question_id] = [str(random.randint(1, 5))]
        
        # Print the random answers for verification
        print("\n===== RANDOM TEST ANSWERS =====")
        print(json.dumps(random_answers, indent=2))
        print("===============================\n")
        
        # Analyze the results - using the original function
        # This will print the prompt to the console
        recommendations = analyze_career_test_results(random_answers)
        
        # Save these random answers to the user's test progress (for testing)
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        
        test_progress_ref.set({
            'current_stage': 2,
            'completed_stages': [1, 2],
            'answers': random_answers,
            'recommendations': recommendations,
            'is_test_data': True
        })
        
        return render_template('admin/career_test_debug.html', 
                             random_answers=random_answers,
                             recommendations=recommendations,
                             all_questions=all_questions)
                             
    except Exception as e:
        print(f"Error in admin test: {e}")
        import traceback
        traceback.print_exc()
        return f"Error: {str(e)}", 500
@app.route('/career-test/restart', methods=['GET'])
@login_required
def restart_career_test():
    """Reset career test progress and start fresh"""
    user_id = session['user_id']
    
    try:
        # Get reference to test progress document
        test_progress_ref = db.collection('users').document(user_id).collection('test_progress').document('career_test')
        
        # Check if progress exists
        test_progress = test_progress_ref.get()
        
        if test_progress.exists:
            # Reset the progress data but maintain the document
            test_progress_ref.update({
                'current_stage': 1,
                'current_question': {'stage_1': 1},
                'answers': {},  # Clear all answers
                'recommendations': [],  # IMPORTANT FIX: Clear recommendations
                'reset_count': firestore.Increment(1),
                'last_reset_at': datetime.datetime.now(tz=datetime.timezone.utc)
            })
            
            print(f"Career test progress reset for user: {user_id}")
        else:
            # Create initial progress document
            test_progress_ref.set({
                'current_stage': 1,
                'current_question': {'stage_1': 1},
                'completed_stages': [],
                'answers': {},
                'recommendations': [],
                'reset_count': 0
            })
            
            print(f"New career test progress created for user: {user_id}")
            
        # Redirect to the first stage of the test
        return redirect(url_for('career_test_stage', stage=1))
        
    except Exception as e:
        print(f"Error resetting career test progress: {e}")
        flash('Произошла ошибка при перезапуске теста. Попробуйте еще раз.')
        return redirect(url_for('career_test'))


def calculate_profile_completeness(user_info):
    score = 0
    
    if user_info.get('bio') and len(user_info.get('bio', '')) > 10:
        score += 15
    
    if user_info.get('specialty'):
        score += 15
    
    if user_info.get('education'):
        score += 10
    

    
    skills = user_info.get('skills', [])
    if skills:
        skill_score = min(len(skills) * 4, 20)
        score += skill_score
    
    projects = user_info.get('projects', [])
    if projects:
        project_score = min(len(projects) * 5, 15)
        score += project_score
    
    certificates = user_info.get('certificates', [])
    if certificates:
        cert_score = min(len(certificates) * 2.5, 5)
        score += cert_score
    
    return score

def create_semantic_search_prompt(query, user_data_list):

    
    cert_count_match = re.search(r'(\d+)\s+certificates?', query.lower())
    if cert_count_match:
        count = int(cert_count_match.group(1))
        cert_filtered_users = [user for user in user_data_list if len(user.get('certificates', [])) == count]
        if cert_filtered_users:
            user_data_list = cert_filtered_users
    
    project_count_match = re.search(r'(\d+)\s+projects?', query.lower())
    if project_count_match:
        count = int(project_count_match.group(1))

        project_filtered_users = [user for user in user_data_list if len(user.get('projects', [])) == count]
        if project_filtered_users:
            user_data_list = project_filtered_users
    
    user_profiles = []
    for user in user_data_list:
        projects_info = []
        for project in user.get('projects', []):
            projects_info.append({
                "title": project.get('title', ''),
                "description": project.get('description', '')[:100] if project.get('description') else '',
                "tags": project.get('tags', [])
            })
            
        certificates_info = [cert.get('title', '') for cert in user.get('certificates', [])]
        
        profile = {
            "id": user.get('id'),
            "username": user.get('display_username', user.get('username', '')),
            "full_name": user.get('full_name', ''),
            "bio": user.get('bio', ''),
            "specialty": user.get('specialty', ''),
            "skills": user.get('skills', []),
            "education": user.get('education', ''),
            "goals": user.get('goals', ''),
            "projects": projects_info,
            "certificates": certificates_info,
            "certificate_count": len(user.get('certificates', [])),
            "project_count": len(user.get('projects', [])),
            "profile_completeness": user.get('completeness_score', 0),
            "searchable_text": user.get('searchable_text', '')
        }
        user_profiles.append(profile)
    
    prompt = f"""
You are a precise search engine that understands user queries and finds exact matches.

SEARCH QUERY: "{query}"

USER PROFILES DATABASE:
```json
{json.dumps(user_profiles, ensure_ascii=False, indent=2)}
```

IMPORTANT SEARCH INSTRUCTIONS:
1. For SPECIFIC searches like "someone who loves X", ONLY return users who explicitly mention loving X in their profile
2. For NUMERICAL searches like "X certificates" or "Y projects", ONLY return users with EXACTLY that number
3. For SKILL-BASED searches, ONLY return users who have that specific skill listed or in project descriptions
4. QUALITY over QUANTITY - return ONLY truly relevant users, even if it's just 1-2 results
5. For programming searches, require actual programming evidence (skills, projects, etc.)
6. NEVER INCLUDE users with empty or mostly empty profiles
7. If FEWER THAN 3 users match the query well, that's fine - return only the genuinely matching users

CRITICAL: For this query "{query}" - you MUST verify that returned profiles contain explicit evidence matching the query. Irrelevant results should NOT be included just to have more results.

RESPONSE FORMAT:
Return ONLY a JSON array of the most relevant user IDs, ordered by relevance:
["most_relevant_user_id", "second_most_relevant_id", ...]

Return ONLY the JSON array with no explanations.
"""
    
    return prompt

def parse_query_with_openai(query):
    try:
        user_data_list = get_users_data(limit=100)
        
        if not user_data_list:
            print("No users found in database")
            return []
            
        search_prompt = create_semantic_search_prompt(query, user_data_list)
        
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a search engine that understands any query and returns only a JSON array of user IDs."},
                {"role": "user", "content": search_prompt}
            ],
            temperature=0.2,
            max_tokens=400
        )
        
        result = response.choices[0].message.content.strip()
        
        try:
            json_pattern = re.compile(r'(\[[\s\S]*\])')
            match = json_pattern.search(result)
            
            if match:
                json_str = match.group(1)
                matching_user_ids = json.loads(json_str)
            else:
                matching_user_ids = json.loads(result)
                
            if not isinstance(matching_user_ids, list):
                print(f"API didn't return a list: {matching_user_ids}")
                return []
            
            matching_users = []
            for user_id in matching_user_ids:
                for user in user_data_list:
                    if user['id'] == user_id:
                        matching_users.append(user)
                        break
            
            return matching_users
            
        except json.JSONDecodeError:
            print(f"Couldn't parse JSON from OpenAI response: {result}")
            return []
            
    except Exception as error:
        print(f"Search error: {str(error)}")
        return []
    
def a_star_search(query, user_data_list, max_results=5):
    query_tokens = set(query.lower().split())
    
    user_scores = []
    
    for user in user_data_list:
        g_score = 0
        

        matching_skills = 0
        for skill in user.get('skills', []):
            if any(token in skill.lower() for token in query_tokens):
                matching_skills += 1
                g_score += 10 
        

        matching_projects = 0
        for project in user.get('projects', []):
            project_title = project.get('title', '').lower()
            project_desc = project.get('description', '').lower()
            project_tags = [tag.lower() for tag in project.get('tags', [])]

            if any(token in project_title for token in query_tokens):
                matching_projects += 1
                g_score += 8  
                
            if any(token in project_desc for token in query_tokens):
                matching_projects += 0.5
                g_score += 5 
                
            if any(token in tag for token in query_tokens for tag in project_tags):
                matching_projects += 0.5
                g_score += 6 
        

        matching_certs = 0
        for cert in user.get('certificates', []):
            cert_title = cert.get('title', '').lower()
            if any(token in cert_title for token in query_tokens):
                matching_certs += 1
                g_score += 7  
        

        specialty = user.get('specialty', '').lower()
        if any(token in specialty for token in query_tokens):
            g_score += 9  

        h_score = 0

        bio = user.get('bio', '').lower()
        education = user.get('education', '').lower()
        goals = user.get('goals', '').lower()
      
        token_density = 0
        for token in query_tokens:
            token_density += bio.count(token) * 1.5
            token_density += education.count(token) * 1.2
            token_density += goals.count(token) * 1.0
            token_density += specialty.count(token) * 2.0
        
        h_score = token_density * 2

        h_score += len(user.get('projects', [])) * 1.5
        h_score += len(user.get('certificates', [])) * 2
        
        f_score = g_score + h_score

        user_scores.append({
            'id': user.get('id'),
            'f_score': f_score,
            'g_score': g_score,
            'h_score': h_score,
            'display_username': user.get('display_username')
        })
    

    user_scores.sort(key=lambda x: x['f_score'], reverse=True)

    return [user['id'] for user in user_scores[:max_results]]


def parse_query_with_openai(query):
    try:
        user_data_list = get_users_data(limit=50)
        
        if not user_data_list:
            print("No users found in database")
            return []
            
        search_prompt = create_semantic_search_prompt(query, user_data_list)
        
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a search engine that finds users by query meaning and returns only a JSON array of user IDs."},
                {"role": "user", "content": search_prompt}
            ],
            temperature=0.2,
            max_tokens=400
        )
        
        result = response.choices[0].message.content.strip()
        
        try:
            json_pattern = re.compile(r'(\[[\s\S]*\])')
            match = json_pattern.search(result)
            
            if match:
                json_str = match.group(1)
                matching_user_ids = json.loads(json_str)
            else:
                matching_user_ids = json.loads(result)
                
            if not isinstance(matching_user_ids, list):
                print(f"API didn't return a list: {matching_user_ids}")
                matching_user_ids = a_star_search(query, user_data_list)
        except json.JSONDecodeError:
            print(f"Couldn't parse JSON from OpenAI response: {result}")
            matching_user_ids = a_star_search(query, user_data_list)

        matching_users = []
        for user_id in matching_user_ids:
            for user in user_data_list:
                if user['id'] == user_id:
                    matching_users.append(user)
                    break
        
        return matching_users
            
    except Exception as error:
        print(f"Semantic search error: {str(error)}")
        try:
            matching_user_ids = a_star_search(query, user_data_list)
            
            matching_users = []
            for user_id in matching_user_ids:
                for user in user_data_list:
                    if user['id'] == user_id:
                        matching_users.append(user)
                        break
            
            return matching_users
        except Exception as e:
            print(f"A* search fallback error: {str(e)}")
            return []


def get_all_users_from_firebase():
    try:
        users_ref = db.collection('users')
        users = users_ref.stream()
        
        all_users = []
        
        for user_doc in users:
            user_data = user_doc.to_dict()
            user_data['id'] = user_doc.id
            
            if user_data.get('blocked', False):
                continue
                
            all_users.append(user_data)
            
        print(f"Retrieved {len(all_users)} user profiles from Firebase")
        return all_users
        
    except Exception as e:
        print(f"Error retrieving users from Firebase: {e}")
        import traceback
        traceback.print_exc()
        return []    
def get_users_data(limit=100):
    users_ref = db.collection('users').limit(limit)
    users = users_ref.stream()
    
    user_data_list = []
    for user_doc in users:
        user_data = user_doc.to_dict()
        user_id = user_doc.id
        
        if user_data.get('blocked', False):
            continue
            
        created_projects = []
        try:
            projects_ref = db.collection('projects').where('created_by', '==', user_id)
            for project_doc in projects_ref.stream():
                project_data = project_doc.to_dict()
                created_projects.append({
                    'id': project_doc.id,
                    'title': project_data.get('title', ''),
                    'description': project_data.get('description', ''),
                    'tags': project_data.get('tags', [])
                })
        except Exception as e:
            print(f"Error getting projects for user {user_id}: {e}")
            
        collaborative_projects = []
        try:
            collab_projects_ref = db.collection('projects')
            for project_doc in collab_projects_ref.stream():
                project_data = project_doc.to_dict()
                collaborators = project_data.get('collaborators', [])
                
                if any(c.get('user_id') == user_id for c in collaborators):
                    collaborative_projects.append({
                        'id': project_doc.id,
                        'title': project_data.get('title', ''),
                        'description': project_data.get('description', ''),
                        'tags': project_data.get('tags', [])
                    })
        except Exception as e:
            print(f"Error getting collaborative projects for user {user_id}: {e}")
            
        certificates = []
        try:
            certs_ref = db.collection('users').document(user_id).collection('certificates')
            for cert_doc in certs_ref.stream():
                cert_data = cert_doc.to_dict()
                certificates.append({
                    'id': cert_doc.id,
                    'title': cert_data.get('title', '')
                })
        except Exception as e:
            print(f"Error getting certificates for user {user_id}: {e}")
            
        user_info = {
            'id': user_id,
            'display_username': user_data.get('display_username', user_data.get('username', '')),
            'bio': user_data.get('bio', ''),
            'specialty': user_data.get('specialty', ''),
            'skills': user_data.get('skills', []),
            'education': user_data.get('education', ''),
            'goals': user_data.get('goals', ''),
            'full_name': user_data.get('full_name', ''),
            'projects': created_projects + collaborative_projects,
            'certificates': certificates
        }
        
        completeness_score = calculate_profile_completeness(user_info)
        user_info['completeness_score'] = completeness_score
        
        user_info['searchable_text'] = (
            f"{user_info['display_username']} {user_info['full_name']} "
            f"{user_info['bio']} {user_info['specialty']} {user_info['education']} "
            f"{user_info['goals']} {' '.join(user_info['skills'])}"
        ).lower()
        
        project_texts = []
        for project in user_info['projects']:
            project_texts.append(project.get('title', ''))
            project_texts.append(project.get('description', ''))
            project_texts.extend(project.get('tags', []))
        
        user_info['project_text'] = ' '.join(project_texts).lower()
        
        cert_texts = [cert.get('title', '') for cert in user_info['certificates']]
        user_info['certificate_text'] = ' '.join(cert_texts).lower()
        
        user_data_list.append(user_info)
        
    return user_data_list

@app.route('/test-openai-search', methods=['GET'])
def test_openai_search():
    try:
        test_query = "someone who loves alibek"
        
        matching_users = parse_query_with_openai(test_query)
        
        results_html = ""
        if matching_users:
            results_html += "<h3>Found users:</h3><ul>"
            for user in matching_users:
                results_html += f"""
                <li class="user-result">
                    <strong>{user['display_username']}</strong>
                    <p><em>{user['specialty']}</em></p>
                    <p>{user['bio'][:200]}{"..." if len(user['bio']) > 200 else ""}</p>
                    <p>Skills: {', '.join(user['skills'])}</p>
                </li>
                """
            results_html += "</ul>"
        else:
            results_html = "<p>No matching users found</p>"

        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Semantic Search Test</title>
            <style>
                body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }}
                h1 {{ color: #2c3e50; }}
                .search-query {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                .user-result {{ background-color: #f1f8ff; padding: 12px; margin-bottom: 10px; border-radius: 4px; border-left: 4px solid #4285f4; }}
                .stats {{ color: #666; font-size: 14px; }}
                ul {{ list-style-type: none; padding: 0; }}
            </style>
        </head>
        <body>
            <h1>Semantic Search Test</h1>
            <div class="search-query">
                <strong>Query:</strong> "{test_query}"
            </div>
            <div class="results">
                {results_html}
            </div>
            <div class="stats">
                <p>Found: {len(matching_users)} users</p>
            </div>
        </body>
        </html>
        """
    except Exception as e:
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Search Error</title>
        </head>
        <body>
            <h1>An error occurred</h1>
            <p>{str(e)}</p>
        </body>
        </html>
        """



@app.errorhandler(404)
def page_not_found(e):

    print(f"404 Error: {e}")
    

    print(f"Request URL: {request.url}")
    print(f"Request Method: {request.method}")
    
    return render_template('404.html'), 404
@app.errorhandler(500)
def internal_server_error(e):

    print(f"500 Error: {e}")
    print(f"Request URL: {request.url}")
    print(f"Request Method: {request.method}")
    
    return render_template('500.html'), 500


app.config['DEBUG'] = True

if check_openai_key_on_start():
    app.run(debug=True)
else:
    print("server aint launching cuz no openai key")
@app.route('/test_404')
def test_404():
    abort(404)

@app.route('/test_500')
def test_500():
    abort(500)
if __name__ == '__main__':
    app.run(debug=True)